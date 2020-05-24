# -*- coding: utf-8 -*-
"""
Created on Mon Mar  9 22:09:39 2020

@author: dell
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Jan 27 18:17:39 2020

pip install mysql-connector

@author: Bhavani
"""
#from PIL import Image, ImageTk
import re 
from itertools import cycle
import datetime
import webbrowser
from openpyxl import Workbook
from tkinter import messagebox
import os
import glob
from fpdf import FPDF
from tkinter import *
import tkinter as tk
import tkinter.ttk as ttk 
from ttkthemes import ThemedStyle
#import cv2
import mysql.connector
db_cur=mysql.connector.connect(host='localhost',user='root',passwd='',database='studentform')
db=db_cur.cursor()
class MyFirstGui(tk.Tk):
       def __init__(self):
                tk.Tk.__init__(self)
                self._frame = None
                
                self.switch_frame(StartPage)
                self.state("zoomed")
       
       def switch_frame(self, frame_class):
                new_frame = frame_class(self)
                if self._frame is not None:
                    self._frame.destroy()
                style = ThemedStyle(self)
                style.set_theme("plastik")
               
                style.configure('W.TButton', font =
                   ('calibri', 20, 'bold'))  
                
                style.configure('Wbb.TButton', font =
                   ('calibri', 10, 'bold'))  
                style.configure('B.TButton', font =
                   ('calibri', 5, 'bold')) 
                style.configure('L.TButton', font =
                   ('calibri', 20, 'bold')) 
                style.configure('Wild.TRadiobutton', font=('calibri',20,'bold'),   # First argument is the name of style. Needs to end with: .TRadiobutton
                 background='darkorchid',foreground='black') 
                self._frame = new_frame
                self._frame.pack()
                self.state("zoomed")
                self.configure(background='cadetBlue1')
               
        
         
class StartPage(tk.Frame): 
           
      def __init__(self, master):
                    tk.Frame.__init__(self,master)
                    self.configure(background='cadetBlue1')
                    self.Frame1 = tk.Frame(self.master)
                    self.Frame1.pack(side="top",  pady=10,padx=10,expand=True )
                    self.Frame1.configure(background='cadetBlue1')
                    photo=tk.PhotoImage(file="ss.png")  
                    label1221=tk.Label(self.Frame1,text="Alagappa chettiar Government College of Engineering and Technology,Karaikudi",font =
                   ('calibri', 30, 'bold'))
                    label1221.grid(row=0,rowspan=1,columnspan=40)
                    label12221=tk.Label(self.Frame1,text="(An autonomous government institution permanently affilitated to anna university)",font =
                   ('calibri', 20, 'bold'))
                    label12221.grid(row=1,rowspan=1,columnspan=40)
                    
                    #photo = photo.zoom(1)
                    label =tk.Label(self,image = photo)#,width=1680,height=1080)
                    label.image = photo # keep a reference!
                    label.grid(row=2,column=0,columnspan=5,rowspan=20)
                    self.Frame2 = tk.Frame(self.master)
                    
                    self.Frame2.pack(side="right",  pady=80,padx=10, expand=True )
                    self.Frame2.configure(background='cadetBlue1')
                    wer=tk.Label(self.Frame2,text='VISION',font = ('calibri', 22, 'bold'))
                    wer.grid(row=2,column=0,sticky='W')
                    le2=tk.Label(self.Frame2,text='Our  Commitment  as  a  Centre   of  ',font = ('calibri', 18, 'bold'))
                    le2.grid(row=3,column=0,sticky='W')#olumnspan=40,rowspan=4)
                    let2=tk.Label(self.Frame2,text='Engineering  Educations   to  impart ',font = ('calibri', 18, 'bold'))
                    let2.grid(row=4,column=0,sticky='W')
                    le12=tk.Label(self.Frame2,text='Technical   Knowledge  par excellence, ',font = ('calibri', 18, 'bold'))
                    le12.grid(row=5,column=0,sticky='W')#olumnspan=40,rowspan=4)
                    lere=tk.Label(self.Frame2,text='motivate the learners   in  Research,    ',font = ('calibri', 18, 'bold'))
                    lere.grid(row=6,column=0,sticky='W')
                    lee=tk.Label(self.Frame2,text='evolve  result  –  oriented,  innovative   ',font = ('calibri',18, 'bold'))
                    lee.grid(row=7,column=0,sticky='W')
                    lee1r=tk.Label(self.Frame2,text='techniques in   Engineering, provide ',font = ('calibri', 18, 'bold'))
                    lee1r.grid(row=8,column=0,sticky='W')
                    lee1=tk.Label(self.Frame2,text='necessary career guidance,and  train  ',font = ('calibri', 18, 'bold'))
                    lee1.grid(row=9,column=0,sticky='W')
                    lee2=tk.Label(self.Frame2,text='our learners in leadership',font = ('calibri',18, 'bold'))
                    lee2.grid(row=10,column=0,sticky='W')
                    lee2=tk.Label(self.Frame2,text='qualities so as to achieve better',font = ('calibri', 18, 'bold'))
                    lee2.grid(row=11,column=0,sticky='W')
                    lee3=tk.Label(self.Frame2,text='productivity and prosperity',font = ('calibri',18, 'bold'))
                    lee3.grid(row=12,column=0,sticky='W')
                    lee3=tk.Label(self.Frame2,text='for our country.',font = ('calibri',18, 'bold'))
                    lee3.grid(row=13,column=0,sticky='W')
                    #self.state("zoomed")
                    global username,passwords
                    self.l1=ttk.Label(self,text="Login here",font=("Times 40 italic"),background="gray69")
                    self.l1.grid(row=3,column=1,columnspan=2,sticky='W',padx=10,pady=50)
                    
                    self.l2=ttk.Label(self,text="Username",font=("Times 30 italic"),background="gray64")
                    self.l2.grid(row=5,column=0,padx=10,pady=50)
                    self.password=ttk.Label(self,text="Password",font=("Times 30 italic"),background="gray43")
                    self.password.grid(row=7,column=0,padx=10,pady=50)
                    self.e1=ttk.Entry(self,font=("Times 18 italic"))
                    self.e1.grid(row=5,column=2,padx=5,pady=5,ipady=5)
                    self.e2=ttk.Entry(self,font=("Times 18 italic"))
                    self.e2.grid(row=7,column=2,padx=5,pady=5,ipady=5)
                    self.b1=ttk.Button(self,text="SUBMIT",style = 'W.TButton',command=self.passd)
                    self.b1.grid(row=9,column=2,padx=30,pady=50)
                    self.b12=ttk.Button(self,text="Register",style = 'W.TButton',command=self.reg)
                    self.b12.grid(row=9,column=1,padx=30,pady=50)
                    self.b2=ttk.Button(self,text="RESET",style = 'W.TButton',command=self.reset)
                    self.b2.grid(row=9,column=0,padx=30,pady=50)
                    
                    #self.button=tk.Button(self,text="->",width=10,command=lambda: master.switch_frame(pageOne))
                    #self.button.grid(row=4,column=0,sticky='W')
                    
            #self.ll1=tk.Button(self, text = 'Click Me !', image = photo1)
      def reset(self):
          self.e1.delete(first=0,last=300)
          self.e2.delete(first=0,last=300)
      def reg(self):
          self.Frame2.destroy()
          self.master.switch_frame(registration)
      def passd(self):
            
            self.username = self.e1.get()
            self.passwords = self.e2.get()
            db.execute("select *from adminlogin where name='%s' and password='%s'"%(self.username,self.passwords))
            l=db.fetchone()
            if(l):
                self.Frame2.destroy()
                self.Frame1.destroy()
                print("check")
                #self.switch_frame(pageOne)
                self.master.switch_frame(pageOne)
                db_cur.commit()
            else:
                   messagebox.showinfo("Login Error", "Invalid Username Or Passord")

                
     # def make_label(parent, img):
      #    label11 = ttk.Label(parent, image=img)
       #   label11.pack()


class registration(tk.Frame):
    def __init__(self,master):
        tk.Frame.__init__(self,master)
        
        #self.configure(background='powderblue')
        self.configure(background='darkorchid')
        photo=tk.PhotoImage(file="ss.png") 
                    
                    #photo = photo.zoom(1)
        label =tk.Label(self,image = photo)#,width=1680,height=1080)
        label.image = photo # keep a reference!
        label.grid(row=0,column=0,columnspan=5,rowspan=20)
        self.Frame3 = tk.Frame(self.master)
                    
        self.Frame3.pack(side="right",  pady=50,padx=10, expand=True )
        self.Frame3.configure(background='cadetBlue1')
        la=tk.Label(self.Frame3,text='MISSION',font = ('calibri',30, 'bold'))
        la.grid(row=0,column=0,sticky='W')
        la1=tk.Label(self.Frame3,text='Constantly updating the departmental resources,',font = ('calibri',18, 'bold'))
        la1.grid(row=1,column=0,sticky='W')
        la2=tk.Label(self.Frame3,text=' facility and other infrastructure ',font = ('calibri',18, 'bold'))
        la2.grid(row=2,column=0,sticky='W')
        la3=tk.Label(self.Frame3,text=' by acquiring state of art equipment.',font = ('calibri',18, 'bold'))
        la3.grid(row=3,column=0,sticky='W')
        la4=tk.Label(self.Frame3,text='')
        la4.grid(row=4,column=0,sticky='W')
        la5=tk.Label(self.Frame3,text='Imparting constant in–service training',font = ('calibri',18, 'bold'))
        la5.grid(row=5,column=0,sticky='W')
        la6=tk.Label(self.Frame3,text='to the faculty and supporting staff.',font = ('calibri',18, 'bold'))
        la6.grid(row=6,column=0,sticky='W')
        la6=tk.Label(self.Frame3,text='',font = ('calibri',18, 'bold'))
        la6.grid(row=7,column=0,sticky='W')
        la6=tk.Label(self.Frame3,text='Inculcating the feeling of oneness responsibility ',font = ('calibri',18, 'bold'))
        la6.grid(row=8,column=0,sticky='W')
        la6=tk.Label(self.Frame3,text='and service to community in the minds',font = ('calibri',18, 'bold'))
        la6.grid(row=9,column=0,sticky='W')
        la6=tk.Label(self.Frame3,text='of students to serve the society better.',font = ('calibri',18, 'bold'))
        la6.grid(row=10,column=0,sticky='W')
        self.lb1=ttk.Label(self,text="Register here",font=("Times 40 italic"),background="gray60")
        self.lb1.grid(row=0,column=0,padx=10,pady=30)
        self.lb2=ttk.Label(self,text="Name",font=("Times 18 italic"),background="gray60")
        self.lb2.grid(row=1,column=0,padx=10,pady=15)
        self.ent1=ttk.Entry(self,font=("Times 18 italic"))
        self.ent1.grid(row=1,column=1,padx=5,pady=5,ipady=5)
        self.lb3=ttk.Label(self,text="Password",font=("Times 18 italic"),background="gray60")
        self.lb3.grid(row=2,column=0,padx=10,pady=15)
        self.ent2=ttk.Entry(self,font=("Times 18 italic"))
        self.ent2.grid(row=2,column=1,padx=5,pady=5,ipady=5)
        self.lb4=ttk.Label(self,text="Department",font=("Times 18 italic"),background="gray60")
        self.lb4.grid(row=3,column=0,padx=10,pady=15)
        self.ent3=ttk.Entry(self,font=("Times 18 italic"))
        self.ent3.grid(row=3,column=1,padx=5,pady=5,ipady=5)
        self.lb5=ttk.Label(self,text="Posting",font=("Times 18 italic"),background="gray60")
        self.lb5.grid(row=4,column=0,padx=10,pady=15)
        self.ent4=ttk.Entry(self,font=("Times 18 italic"))
        self.ent4.grid(row=4,column=1,padx=5,pady=5,ipady=5)
        self.lb6=ttk.Label(self,text="Admin Name",font=("Times 18 italic"),background="gray60")
        self.lb6.grid(row=5,column=0,padx=10,pady=15)
        self.ent5=ttk.Entry(self,font=("Times 18 italic"))
        self.ent5.grid(row=5,column=1,padx=5,pady=5,ipady=5)
        self.lb7=ttk.Label(self,text="Admin Password",font=("Times 18 italic"),background="gray60")
        self.lb7.grid(row=6,column=0,padx=10,pady=15)
        self.ent6=ttk.Entry(self,font=("Times 18 italic"))
        self.ent6.grid(row=6,column=1,padx=5,pady=5,ipady=5)
        self.bt1=ttk.Button(self,text="Register",command=self.regg,style = 'W.TButton')
        self.bt1.grid(row=7,column=1,sticky='S',padx=10,pady=15)
       
        self.bt2=ttk.Button(self,text="RESET",style = 'W.TButton',command=self.rese)
        self.bt2.grid(row=7,column=0,padx=30,pady=15)
    def rese(self):  
        self.ent1.delete(first=0,last=300)
        self.ent2.delete(first=0,last=300)
        self.ent3.delete(first=0,last=300)
        self.ent4.delete(first=0,last=300)
        self.ent5.delete(first=0,last=300)
        self.ent6.delete(first=0,last=300)
    def regg(self):
        self.namee=self.ent1.get()
        self.passsword=self.ent2.get()
        self.depart=self.ent3.get()
        self.post=self.ent4.get()
        self.adminu=self.ent5.get()
        self.adminpass=self.ent6.get()
        db.execute("select *from adminlogin where name='%s' and password='%s'"%(self.adminu,self.adminpass))
        l=db.fetchone()
        if(l):
            sql1="INSERT INTO `adminlogin`(`name`, `password`, `department`, `posting`) VALUES(%s,%s,%s,%s)"
            val1=(self.namee,self.passsword,self.depart,self.post)
            db.execute(sql1, val1)
            db_cur.commit()
            self.Frame3.destroy()
            self.master.switch_frame(StartPage)
        else:
            messagebox.showinfo("Registration Error", "You Are Not a Valid Admin.. Please Get Verification From The Admin")
                 


       
class pageOne(tk.Frame):
      def __init__(self,master):
             tk.Frame.__init__(self,master)
             #self.Frame1 = tk.Frame(self.master)
             #self.Frame1.pack(side="top",  pady=10,padx=10,expand=True )
             #self.Frame1.configure(background='cadetBlue1')
             #photo=tk.PhotoImage(file="ss.png")  
             label1221=tk.Label(self,text="Alagappa chettiar Government College of Engineering and Technology,Karaikudi",font =
                   ('calibri', 30, 'bold'))
             label1221.grid(row=0)
             label12221=tk.Label(self,text="(An autonomous government institution permanently affilitated to anna university)",font =
                   ('calibri', 20, 'bold'))
             label12221.grid(row=1)
             self.configure(background='darkorchid')
             self.menubar = tk.Menu(master,bg="lightgrey", fg="black")
             self.filemenu = tk.Menu(self.menubar, tearoff=0,background='green', foreground='yellow',activebackground='blue', activeforeground='red')
             self.filemenu.add_command(label="Individual Transfer Certificate",command=lambda: master.switch_frame(transfersingle),font=("Times 18 italic"))
             self.filemenu.add_command(label="Batch Transfer Certificate",command=lambda: master.switch_frame(transfergroup),font=("Times 18 italic"))
             self.menubar.add_cascade(label="Transfer Certificate Generation", menu=self.filemenu)
             self.edit = tk.Menu(self.menubar, tearoff=0,relief=tk.SUNKEN, bd=0,background='green', foreground='yellow',activebackground='blue', activeforeground='red')
             self.edit.add_command(label="Edit",command=lambda: master.switch_frame(editspecify),font=("Times 18 italic"))
             
             self.menubar.add_cascade(label='Edit Details',menu=self.edit)
             self.insert = tk.Menu(self.menubar, tearoff=0,relief=tk.SUNKEN, bd=0,background='green', foreground='yellow',activebackground='blue', activeforeground='red')
             self.insert.add_command(label="Insert",command=lambda: master.switch_frame(newreg),font=("Times 18 italic"))
         
             self.menubar.add_cascade(label='Insert Details',menu=self.insert)
             self.search = tk.Menu(self.menubar, tearoff=0,relief=tk.SUNKEN, bd=0,background='green', foreground='yellow',activebackground='blue', activeforeground='red')
             self.search.add_command(label="Search",command=lambda: master.switch_frame(search),font=("Times 18 italic"))
         
             self.menubar.add_cascade(label='search Details',menu=self.search)
             #self.menubar.config("Verdana", 14)
             self.master.config(menu=self.menubar,background="Red")
             
             #self.top =Toplevel(menu=self.menubar, width=500, relief=tk.RAISED,borderwidth=2)
             self.delay = 3000
             image_files = [
            'ss.png',
            '190.png',
            'ss.png',
            'ss.png',
            'ss.png'
            ]
            # allows repeat cycling through the pictures
            # store as (img_object, img_name) tuple
             self.pictures = cycle((tk.PhotoImage(file=image), image)
                                  for image in image_files)
             self.picture_display = tk.Label(self,width=600,height=400)
             self.picture_display.grid(row=3,column=0,sticky='W')
             #self.p = cycle((tk.PhotoImage(file=image), image)for image in image_files)
             self.p_d=tk.Label(self,width=600,height=400)             
             self.p_d.grid(row=3,column=0,sticky='E')
             self.label=tk.Label(self,text='')
             self.label.grid(row=2,column=1)
             self.lee=tk.Label(self)
             self.lee.grid(row=4)
             self.lee2=tk.Label(self,text='HISTORY OF ACGCET',font=("Times 15 italic"))
             self.lee2.grid(row=5,sticky='W')
             self.lw=tk.Label(self,font=("Times 11 "),text='Dr.RM.Alagappa Chettiar, a man of rare wisdom, and forethought, founded  Alagappa Chettiar Educational Trust  with the sole aim of developing the backward area of Karaikudi into a centre for higher education and provided ')
             self.lw.grid(row=6,column=0,sticky='W')
             self.lw1=tk.Label(self,text='necessary funds for the establishment of education and provided necessary funds for the establishment of educational institutions. On the Occasion of laying of the foundation stone of the central Electro Chemical Research Institue',font=("Times 11 "))
             self.lw1.grid(row=7,column=0,sticky='W')
             self.lw2=tk.Label(self,text='by pandit jawaharlal Nehru on 25th july 1948. Dr.Alagappa Chettiar, in his Welcome address, said “It is my hope to start here an Engineering College immediately. A College with Dr.Alagappa Chettiar and the University of Madras ',font=("Times 11 "))
             self.lw2.grid(row=8,column=0,sticky='W')
             self.lw3=tk.Label(self,text='willing start functioning with Civil Engineering by academic year 1949.” In 1952, Dr. Alagappa Chettiar’s dream came true. Alagappa Chettiar College of Engg & Tech., Started functioning from 21st July 1952 with three ',font=("Times 11 "))
             self.lw3.grid(row=9,column=0,sticky='W')
             self.lw4=tk.Label(self,text='faculties-Civil,Mechanical, Electrical & Electronics Engineering.The Foundation tablet for the main building of the college was laid by Dr.Rajendra Prasad, the then President of India on 19 th February 1953.',font=("Times 11 "))
             self.lw4.grid(row=10,column=0,sticky='W')
             img_object, img_name = next(self.pictures)
             self.picture_display.config(image=img_object)
             self.p_d.config(image=img_object)
             # shows the image filename, but could be expanded
             # to show an associated description of the image
             #self.title(img_name)
             self.after(self.delay, self.show_slides)
           
        
      def show_slides(self):
        '''cycle through the images and show them'''
        # next works with Python26 or higher
        img_object, img_name = next(self.pictures)
        self.picture_display.config(image=img_object)
        self.p_d.config(image=img_object)
        # shows the image filename, but could be expanded
        # to show an associated description of the image
        #self.title(img_name)
        self.after(self.delay, self.show_slides)



             #self.configure(menu=menubar)

             #self.configure(background='powderblue')
             
             

        
class transferspecify(tk.Frame):
      def __init__(self, master):
             tk.Frame.__init__(self,master)
             self.configure(background='darkorchid')
             lablle1=ttk.Label(self,text="Transfer Certificate",font=("Times 30 italic"),background="darkorchid")
             lablle1.grid(row=0,column=1,sticky='NSWE',padx=10,pady=90)
             
             button=ttk.Button(self,text="Individual Transfer Certificate",style = 'W.TButton',command=lambda: master.switch_frame(transfersingle))
             button.grid(row=1,column=1,sticky='W',padx=30,pady=20)
             ll=ttk.Label(self)
             ll.grid(row=2,column=1,padx=10,pady=80)
             Button1=ttk.Button(self,text="     Batch Transfer Certificate    ",style = 'W.TButton',command=lambda: master.switch_frame(transfergroup))
             Button1.grid(row=2,column=1,sticky='W',padx=30,pady=20)
             
             
             self.button13=ttk.Button(self,text="BACK",style = 'W.TButton',command=lambda: master.switch_frame(pageOne))
             self.button13.grid(row=4,column=1,sticky='W',padx=30,pady=20)
         
class transfergroup(tk.Frame):
        def __init__(self, master):
                tk.Frame.__init__(self,master)
                self.configure(background='darkorchid')
                self.view=tk.IntVar()
                self.l=ttk.Label(self,text="Enter the Details",font=("Times 22 italic"),background="darkorchid")
                self.l.grid(row=0,padx=10,pady=20)
                
                
                self.l12=ttk.Label(self,text="Branch Of Study",font=("Times 18 italic"),background="darkorchid")
                self.l12.grid(row=1,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.ll12=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll12.grid(row=1,column=2,sticky='W',pady=3,padx=40)
                self.t6 = tk.StringVar(self)
                self.p6 =ttk.Combobox(self, width=25, textvariable=self.t6)
                self.p6['values']=('Civil Engineering','Mechanical Engineering','Electrical and Electronics Engineering','Electronics and Communication Engineering','Computer science and Engineering')  
                self.p6.grid(row=1,column=1,padx=5,pady=5,ipady=5,sticky='W')
                self.p6.current()
                
                
                              
                self.le2=ttk.Label(self,text="Leaving Date",font=("Times 18 italic"),background="darkorchid")
                self.le2.grid(row=2,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.ee2=ttk.Entry(self,font=("Times 18 italic"))
                self.ee2.grid(row=2,column=1,padx=5,pady=5,ipady=5,sticky='W')
                self.le3=ttk.Label(self,text="Year Of PassOut",font=("Times 18 italic"),background="darkorchid")
                self.le3.grid(row=3,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.ee3=ttk.Entry(self,font=("Times 18 italic"))
                self.ee3.grid(row=3,column=1,padx=5,pady=5,ipady=5,sticky='W')
                #print(entr1.get())
                
                self.vieww=tk.IntVar()
                self.viewe=tk.IntVar()
                self.viiew=tk.IntVar()
                
                
                self.l1=ttk.Label(self,text="Whether the student has paid all the fees due to the college?",font=("Times 20 italic"),background='darkorchid')
                self.l1.grid(row=5,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=10)
                self.r10=ttk.Radiobutton(self, text="Yes",variable=self.vieww, value=1,style = 'Wild.TRadiobutton',command=self.feey)
                self.r10.grid(row=5,column=1,sticky='W',padx=20,pady=5,ipadx=20,ipady=10)
                self.r11=ttk.Radiobutton(self, text="No",variable=self.vieww, value=2,style = 'Wild.TRadiobutton',command=self.feex)
                self.r11.grid(row=5,column=2,sticky='W',padx=20,pady=5,ipadx=20,ipady=10)
                
                self.l2=ttk.Label(self,text="Whether the student was in the receipt of any scholarship",font=("Times 20 italic"),background='darkorchid')
                self.l2.grid(row=6,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=10)
                self.r20=ttk.Radiobutton(self, text="Yes",variable=self.viewe, value=1,style = 'Wild.TRadiobutton',command=self.schy)
                self.r20.grid(row=6,column=1,sticky='W',padx=20,pady=5,ipadx=20,ipady=10)
                self.r21=ttk.Radiobutton(self, text="No",variable=self.viewe, value=2,style = 'Wild.TRadiobutton',command=self.schx)
                self.r21.grid(row=6,column=2,sticky='W',padx=20,pady=5,ipadx=20,ipady=10)
                
                self.l3=ttk.Label(self,text="Whether the student has undergone Medical inspection during the year",font=("Times 20 italic"),background='darkorchid')
                self.l3.grid(row=7,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=10)
                self.r30=ttk.Radiobutton(self, text="Yes",variable=self.viiew, value=1,style = 'Wild.TRadiobutton',command=self.medy)
                self.r30.grid(row=7,column=1,sticky='W',padx=10,pady=5,ipadx=20,ipady=10)
                self.r31=ttk.Radiobutton(self, text="No",variable=self.viiew, value=2,style = 'Wild.TRadiobutton',command=self.medx)
                self.r31.grid(row=7,column=2,sticky='W',padx=10,pady=5,ipadx=20,ipady=10)
                                                        
                self.l4=ttk.Label(self,text="Reason for leaving the college",font=("Times 20 italic"),background="darkorchid")
                self.l4.grid(row=8,column=0,padx=5,pady=5,sticky='W')
                self.e40=ttk.Entry(self,font=("Times 15 italic"))
                self.e40.grid(row=8,column=1,padx=5,pady=5,ipady=5,sticky='W')
                
                self.l4=ttk.Label(self,text="Date on which application for Transfer Certificate was ",font=("Times 20 italic"),background="darkorchid")
                self.l4.grid(row=9,column=0,padx=5,pady=5,sticky='W')                
                self.l5=ttk.Label(self,text="made by the student or on his/her behalf by parent/guardian",font=("Times 20 italic"),background="darkorchid")
                self.l5.grid(row=10,column=0,padx=5,pady=5,sticky='W')
                self.e41=ttk.Entry(self,font=("Times 15 italic"))
                self.e41.grid(row=10,column=1,padx=5,pady=5,ipady=5,sticky='W')
                #print(entr1.get())
                
                self.but=ttk.Button(self,text="SUBMIT",style = 'W.TButton',command=self.tranzgrp)
                self.but.grid(row=11,column=1,padx=5,pady=5,sticky='E')
              
                self.button16=ttk.Button(self,text="BACK",style = 'W.TButton',command=lambda: master.switch_frame(transferspecify))
                self.button16.grid(row=11,column=0,sticky='W')
        
        def feey(self):
            self.feeey="yes"
        def feex(self):
            self.feeey="no"
        def schy(self):
            self.schhy="yes"
        def schx(self):
            self.schhy="no"
        def medy(self):
            self.meedy="yes"
        def medx(self):
            self.meedy="no"
        
            
        def tranzgrp(self):
            global sentt,getit,yopo,ppname
            sentt=self.p6.get()
            yopo = self.ee3.get()
            getit=self.ee2.get()
            self.reasn=self.e40.get()
            self.donw=self.e41.get()
            #print(sentt)
            #print(getit)
            db.execute("select *from studentdetails where branch='%s' and Year_Of_Passout ='%s' and flag='0'"%(sentt,yopo))
            my=db.fetchall()
            #j=1
            for o in my:
                reg_no = o[1]
                if(datetime.datetime.strptime(self.donw, '%Y-%m-%d') and self.feeey!="" and reg_no !="" and self.schhy!="" and self.meedy!="" and self.schhy!="" ):
                     sql0="INSERT INTO `addtcinfo`(`reg_no`, `student_bill`, `scholarship`, `medicalinspection`, `reasonforleaving`, `addofaplication`) VALUES(%s,%s,%s,%s,%s,%s)"
                     val0=(reg_no,self.feeey,self.schhy,self.meedy,self.reasn,self.donw)
                     db.execute(sql0,val0)
                     db_cur.commit() 
                     db.execute("update studentdetails set flag='1' where reg_number='%s'"%(reg_no))
                         
                     db_cur.commit()             
                     self.master.switch_frame(pageOne) 
                else:
                    messagebox.showinfo("Transfer Error", "Fill all the fields and Check above details are True") 

              
            db.execute("select *from studentdetails where branch='%s' and Year_Of_Passout ='%s'"%(sentt,yopo))
            myresult=db.fetchall()
            pdf=FPDF()
            #reg=
            i=1
           
            for f in myresult:
                print(i)
                i=i+1
                db.execute("SELECT * FROM `addtcinfo` WHERE `reg_no`= '%s' "%(f[1]))
                mys=db.fetchone()
                if(mys):
                  print("2")
                  pdf.add_page()
                  pdf.set_font("Arial",size=12)
                  name="alagappa.jpg"
                  pdf.image(name,w=200,h=20)
                    
                  pdf.cell(100,10,txt="                                                              TRANSFER CERTIFICATE")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     1.  Name of the student")
                  pdf.cell(100,10,txt="              "+str(f[0]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     2.  Name of the parent/Guardian")
                  pdf.cell(100,10,txt="              "+str(f[3]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     3.  Nationality Religion and coommunity")
                  pdf.cell(100,10,txt="              "+str(f[6]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     4.  Sex")
                  pdf.cell(100,10,txt="              "+str(f[8]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     5.  Date of Birth (in figure and words) ")
                  pdf.cell(100,10,txt="              "+str(f[9]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     as entered in the admission register")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     6.  Course of Study")
                  pdf.cell(100,10,txt="              "+str(f[10]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     7.  Date of Admission to this college")
                  pdf.cell(100,10,txt="              "+str(f[12]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     8.  a) Whether the Student has paid all the  ")
                  pdf.cell(100,10,txt="              "+str(mys[1]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     Fees due to the college ?")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     8.  b) Whether the Student was in receipt of ")
                  pdf.cell(100,10,txt="              "+str(mys[2]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     any scholarship")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     9.  Whether the Student has undergone ")
                    
                  pdf.cell(100,10,txt="              "+str(mys[3]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     Medical inspection during the year")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     10.  Reasons for leaving the College")
                    
                  pdf.cell(100,10,txt="              "+str(mys[4]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     11.  Date of Leaving")
                  pdf.cell(100,10,txt="              "+ getit)
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     12.  Date on which application for Transfer  ")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     Certificate was made by the Student or On ")
                  pdf.cell(100,10,txt="              "+str(mys[5]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     his/her behalf by Parent/Guardian")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     13.  Date of the Transfer Certificate")
                  pdf.cell(100,10,txt="              "+str(f[23]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     14.  Medium")
                  pdf.cell(100,10,txt="              English")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     Seal")
                  pdf.cell(100,10,txt="              PRINCIPAL/VICE-PRINCIPAL")
                  pdf.ln(10)
            r=sentt+".pdf"
            pdf.output(r)
            if os.path.exists(r):
                     os.startfile(r)
            
class transfersingle(tk.Frame):
          
          def __init__(self, master):
                tk.Frame.__init__(self,master)
                self.configure(background='darkorchid')
                self.l=ttk.Label(self,text="Enter the Details",font=("Times 30 italic"),background="darkorchid")
                self.l.grid(row=0,padx=10,pady=10,ipadx=20,ipady=10)
                
                self.l1=ttk.Label(self,text="Register Number",font=("Times 18 italic"),background="darkorchid")
                self.l1.grid(row=1,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.entr1=ttk.Entry(self,font=("Times 15 italic"))
                self.entr1.grid(row=1,column=1,padx=5,pady=1,ipady=5,sticky='W')
                self.l2=ttk.Label(self,text="Leaving Date",font=("Times 18 italic"),background="darkorchid")
                self.l2.grid(row=2,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                
                
                
                self.e2=ttk.Entry(self,font=("Times 15 italic"))
                self.e2.grid(row=2,column=1,padx=5,pady=1,ipady=5,sticky='W')
                self.vieww=tk.IntVar()
                self.viewe=tk.IntVar()
                self.viiew=tk.IntVar()
                
                
                self.l1=ttk.Label(self,text="Whether the student has paid all the fees due to the college?",font=("Times 20 italic"),background='darkorchid')
                self.l1.grid(row=5,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.r10=ttk.Radiobutton(self, text="Yes",variable=self.vieww, value=1,style = 'Wild.TRadiobutton',command=self.feey)
                self.r10.grid(row=5,column=1,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                self.r11=ttk.Radiobutton(self, text="No",variable=self.vieww, value=2,style = 'Wild.TRadiobutton',command=self.feex)
                self.r11.grid(row=5,column=2,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                
                self.l2=ttk.Label(self,text="Whether the student was in the receipt of any scholarship",font=("Times 20 italic"),background='darkorchid')
                self.l2.grid(row=6,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.r20=ttk.Radiobutton(self, text="Yes",variable=self.viewe, value=1,style = 'Wild.TRadiobutton',command=self.schy)
                self.r20.grid(row=6,column=1,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                self.r21=ttk.Radiobutton(self, text="No",variable=self.viewe, value=2,style = 'Wild.TRadiobutton',command=self.schx)
                self.r21.grid(row=6,column=2,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                
                self.l3=ttk.Label(self,text="Whether the student has undergone Medical inspection during the year",font=("Times 20 italic"),background='darkorchid')
                self.l3.grid(row=7,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.r30=ttk.Radiobutton(self, text="Yes",variable=self.viiew, value=1,style = 'Wild.TRadiobutton',command=self.medy)
                self.r30.grid(row=7,column=1,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                self.r31=ttk.Radiobutton(self, text="No",variable=self.viiew, value=2,style = 'Wild.TRadiobutton',command=self.medx)
                self.r31.grid(row=7,column=2,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                                                        
                self.l4=ttk.Label(self,text="Reason for leaving the college",font=("Times 20 italic"),background="darkorchid")
                self.l4.grid(row=8,column=0,padx=10,pady=30,sticky='W')
                self.e40=ttk.Entry(self,font=("Times 15 italic"))
                self.e40.grid(row=8,column=1,padx=5,pady=5,ipady=5,sticky='W')
                
                self.l4=ttk.Label(self,text="Date on which application for Transfer Certificate was ",font=("Times 20 italic"),background="darkorchid")
                self.l4.grid(row=9,column=0,padx=10,pady=10,sticky='W')                
                self.l5=ttk.Label(self,text="made by the student or on his/her behalf by parent/guardian",font=("Times 20 italic"),background="darkorchid")
                self.l5.grid(row=10,column=0,padx=10,pady=1,sticky='W')
                self.e41=ttk.Entry(self,font=("Times 15 italic"))
                self.e41.grid(row=10,column=1,padx=5,pady=1,ipady=5,sticky='W')
                #print(entr1.get())
                self.but=ttk.Button(self,text="Submit",style = 'W.TButton',command=self.tranz)
                self.but.grid(row=11,column=1,padx=35,pady=5,ipady=5,sticky='W')
              
                self.button16=ttk.Button(self,text="BACK",style = 'W.TButton',command=lambda: master.switch_frame(transferspecify))
                self.button16.grid(row=11,column=0,sticky='W')
          def feey(self):
            self.feeey="yes"
          def feex(self):
                self.feeey="no"
          def schy(self):
                self.schhy="yes"
          def schx(self):
                self.schhy="no"
          def medy(self):
                self.meedy="yes"
          def medx(self):
                self.meedy="no"
          
                
        
          def tranz(self):
                self.reasn=self.e40.get()
                self.donw=self.e41.get()
                global setentvar,sete2var
                setentvar=self.entr1.get()
                #self.regnoo=self.e412.get()
                #self.mbso=self.e42.get()
                print(self.feeey)
                print(self.schhy)
                print(self.meedy)
                print(self.reasn)
                print(self.donw)
                #print(self.mbso)
                
                if(datetime.datetime.strptime(self.donw, '%Y-%m-%d') and self.feeey!="" and setentvar!="" and self.schhy!="" and self.meedy!="" and self.schhy!="" ):
                         sql0="INSERT INTO `addtcinfo`(`reg_no`, `student_bill`, `scholarship`, `medicalinspection`, `reasonforleaving`, `addofaplication`) VALUES(%s,%s,%s,%s,%s,%s)"
                         val0=(setentvar,self.feeey,self.schhy,self.meedy,self.reasn,self.donw)
                         db.execute(sql0,val0)
                         db_cur.commit() 
                         db.execute("update studentdetails set flag='1' where reg_number='%s'"%(setentvar))
                         db_cur.commit()             
                         #self.master.switch_frame(pageOne) 
                         
                         sete2var=self.e2.get()
                         #MyFirstGui().switch_frame(transferCertificate)
                         db.execute("select *from studentdetails where reg_number='%s'"%(setentvar))
                         f=db.fetchone()
                         db.execute("SELECT * FROM `addtcinfo` WHERE `reg_no`= '%s' "%(setentvar))
                         mys=db.fetchone()
                         if(f and mys):
                              reg=setentvar
                               
                              pdf=FPDF()
                              pdf.add_page()
                              pdf.set_font("Arial",size=12)
                              name="alagappa.jpg"
                              pdf.image(name,w=200,h=20)
                                
                              pdf.cell(100,10,txt="                                                              TRANSFER CERTIFICATE")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     1.  Name of the student")
                              pdf.cell(100,10,txt="              "+str(f[0]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     2.  Name of the parent/Guardian")
                              pdf.cell(100,10,txt="              "+str(f[3]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     3.  Nationality Religion and coommunity")
                              pdf.cell(100,10,txt="              "+str(f[6]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     4.  Sex")
                              pdf.cell(100,10,txt="              "+str(f[8]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     5.  Date of Birth (in figure and words) ")
                              pdf.cell(100,10,txt="              "+str(f[9]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     as entered in the admission register")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     6.  Course of Study")
                              pdf.cell(100,10,txt="              "+str(f[10]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     7.  Date of Admission to this college")
                              pdf.cell(100,10,txt="              "+str(f[12]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     8.  a) Whether the Student has paid all the  ")
                              pdf.cell(100,10,txt="              "+str(mys[1]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     Fees due to the college ?")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     8.  b) Whether the Student was in receipt of ")
                              pdf.cell(100,10,txt="              "+str(mys[2]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     any scholarship")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     9.  Whether the Student has undergone ")
                                
                              pdf.cell(100,10,txt="              "+str(mys[3]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     Medical inspection during the year")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     10.  Reasons for leaving the College")
                                
                              pdf.cell(100,10,txt="              "+str(mys[4]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     11.  Date of Leaving")
                              pdf.cell(100,10,txt="              "+ sete2var)
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     12.  Date on which application for Transfer  ")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     Certificate was made by the Student or On ")
                              pdf.cell(100,10,txt="              "+str(mys[5]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     his/her behalf by Parent/Guardian")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     13.  Date of the Transfer Certificate")
                              pdf.cell(100,10,txt="              "+str(f[23]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     14.  Medium")
                              pdf.cell(100,10,txt="              English")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     Seal")
                              pdf.cell(100,10,txt="              PRINCIPAL/VICE-PRINCIPAL")
                              pdf.ln(10)
                              r=reg+".pdf"
                              pdf.output(r)
                              if os.path.exists(r):
                                           os.startfile(r)
                              else:
                                  messagebox.showinfo("TC Error", "Enter A Valid Register Number")
                         else:
                             messagebox.showinfo("Transfer Error", "Fill all the fields and Check above details are True") 


                
          
           
class transferCertificate(tk.Frame):
            
   def __init__(self, master):
                tk.Frame.__init__(self,master)
                self.configure(background='darkorchid')
                db.execute("select *from studentdetails where reg_number=%s"%(setentvar))
                f=db.fetchone()
                if(f):
                    print("ok")
                Labe=ttk.Label(self,text="TRANSFER CERTIFICATE",font=("Times 30 italic"),background="darkorchid")
                Labe.grid(row=2,sticky='NSWE',pady=15)
                lab=ttk.Label(self,text="SI.NO :",font=("Times 30 italic"),background="darkorchid")
                lab.grid(row=3,column=0,sticky='W',pady=15)
                la=ttk.Label(self,text="ROLL NO",font=("Times 30 italic"),background="darkorchid")
                la.grid(row=3,column=3,sticky='W',pady=15)
                
                
                
                l1=ttk.Label(self,text="1.  Name of the student",font=("Times 30 italic"),background="darkorchid")
                l1.grid(row=4,column=0,sticky='W',pady=7)
                s1=ttk.Label(self,text=" : ",background="darkorchid")
                s1.grid(row=4,column=1,sticky='W',pady=7)
                a1=ttk.Label(self,text=f[0],font=("Times 30 italic"),background="darkorchid")
                a1.grid(row=4,column=2,sticky='W',pady=7)
                
                l2=ttk.Label(self,text="2.  Name of the parent/Guardian",font=("Times 30 italic"),background="darkorchid")
                l2.grid(row=5,column=0,sticky='W',pady=7)
                s2=ttk.Label(self,text=" : ",background="darkorchid")
                s2.grid(row=5,column=1,sticky='W',pady=7)
                a2=ttk.Label(self,text=f[3],font=("Times 30 italic"),background="darkorchid")
                a2.grid(row=5,column=2,sticky='W',pady=7)
                
                l3=ttk.Label(self,text="3.  Nationality Religion and coommunity",font=("Times 30 italic"),background="darkorchid")
                l3.grid(row=6,column=0,sticky='W',pady=7)
                s3=ttk.Label(self,text=" : ",background="darkorchid")
                s3.grid(row=6,column=1,sticky='W',pady=7)
                a3=ttk.Label(self,text=f[4]+"  "+f[5]+"  "+f[7],font=("Times 30 italic"),background="darkorchid")
                a3.grid(row=6,column=2,sticky='W',pady=7)
                
                l4=ttk.Label(self,text="4.  Sex",font=("Times 30 italic"),background="darkorchid")
                l4.grid(row=7,column=0,sticky='W',pady=7)
                s4=ttk.Label(self,text=" : ",background="darkorchid")
                s4.grid(row=7,column=1,sticky='W',pady=7)
                a4=ttk.Label(self,text=f[8],font=("Times 30 italic"),background="darkorchid")
                a4.grid(row=7,column=2,sticky='W',pady=7)
                
                l5=ttk.Label(self,text="5.  Date of Birth (in figure and words) as entered in the admission register",font=("Times 30 italic"),background="powderblue")
                l5.grid(row=8,column=0,sticky='W',pady=7)
                s5=ttk.Label(self,text=" : ",background="darkorchid")
                s5.grid(row=8,column=1,sticky='W',pady=7)
                a5=ttk.Label(self,text=f[9],font=("Times 30 italic"),background="darkorchid")
                a5.grid(row=8,column=2,sticky='W',pady=7)
                
                l6=ttk.Label(self,text="6.  Course of Study",font=("Times 30 italic"),background="darkorchid")
                l6.grid(row=9,column=0,sticky='W',pady=7)
                s6=ttk.Label(self,text=" : ",background="darkorchid")
                s6.grid(row=9,column=1,sticky='W',pady=7)
                a6=ttk.Label(self,text=f[10],font=("Times 30 italic"),background="darkorchid")
                a6.grid(row=9,column=2,sticky='W',pady=7)
                
                l7=ttk.Label(self,text="7.  Date of Admission to this college",font=("Times 30 italic"),background="darkorchid")
                l7.grid(row=10,column=0,sticky='W',pady=7)
                s7=ttk.Label(self,text=" : ",background="darkorchid")
                s7.grid(row=10,column=1,sticky='W',pady=7)
                a7=ttk.Label(self,text=f[12],font=("Times 30 italic"),background="darkorchid")
                a7.grid(row=10,column=2,sticky='W',pady=7)
                
                l8=ttk.Label(self,text="8.  a) Whether the Student has paid all the Fees due to the college ?",font=("Times 30 italic"),background="powderblue")
                l8.grid(row=11,column=0,sticky='W',pady=7)
                s8=ttk.Label(self,text=" : ",background="darkorchid")
                s8.grid(row=11,column=1,sticky='W',pady=7)
                a8=ttk.Label(self,text="",font=("Times 30 italic"),background="darkorchid")
                a8.grid(row=11,column=2,sticky='W',pady=7)
                
                l9=ttk.Label(self,text="8.  b) Whether the Student was in receipt of any scholarship",font=("Times 30 italic"),background="powderblue")
                l9.grid(row=12,column=0,sticky='W',pady=7)
                s9=ttk.Label(self,text=" : ",background="darkorchid")
                s9.grid(row=12,column=1,sticky='W',pady=7)
                a9=ttk.Label(self,text="",font=("Times 30 italic"),background="darkorchid")
                a9.grid(row=12,column=2,sticky='W',pady=7)
                
                l10=ttk.Label(self,text="9.  Whether the Student has undergone Medical inspection during the year",font=("Times 30 italic"),background="powderblue")
                l10.grid(row=13,column=0,sticky='W',pady=7)
                s10=ttk.Label(self,text=" : ",background="powderblue")
                s10.grid(row=13,column=1,sticky='W',pady=7)
                a10=ttk.Label(self,text="",font=("Times 30 italic"),background="powderblue")
                a10.grid(row=13,column=2,sticky='W',pady=7)
                
                l11=ttk.Label(self,text="10.  Reasons for leaving the College",font=("Times 30 italic"),background="powderblue")
                l11.grid(row=14,column=0,sticky='W',pady=7)
                s11=ttk.Label(self,text=" : ",background="powderblue")
                s11.grid(row=14,column=1,sticky='W',pady=7)
                a11=ttk.Label(self,text="",font=("Times 30 italic"),background="powderblue")
                a11.grid(row=14,column=2,sticky='W',pady=7)
                
                l12=ttk.Label(self,text="11.  Date of Leaving",font=("Times 30 italic"),background="powderblue")
                l12.grid(row=15,column=0,sticky='W',pady=7)
                s12=ttk.Label(self,text=" : ",background="powderblue")
                s12.grid(row=15,column=1,sticky='W',pady=7)
                a12=ttk.Label(self,text=sete2var,font=("Times 30 italic"),background="powderblue")
                a12.grid(row=15,column=2,sticky='W',pady=7)
                
                l13=ttk.Label(self,text="12.  Date on which application for Transfer Certificate was made by the Student or On his/her behalf by Parent/Guardian",font=("Times 30 italic"),background="powderblue")
                l13.grid(row=16,column=0,sticky='W',pady=7)
                s13=ttk.Label(self,text=" : ",background="powderblue")
                s13.grid(row=16,column=1,sticky='W',pady=7)
                a13=ttk.Label(self,text="",font=("Times 30 italic"),background="powderblue")
                a13.grid(row=16,column=2,sticky='W',pady=7)
                
                l14=ttk.Label(self,text="13.  Date of the Transfer Certificate",font=("Times 30 italic"),background="powderblue")
                l14.grid(row=17,column=0,sticky='W',pady=7)
                s14=ttk.Label(self,text=" : ",background="powderblue")
                s14.grid(row=17,column=1,sticky='W',pady=7)
                a14=ttk.Label(self,text="",font=("Times 30 italic"),background="powderblue")
                a14.grid(row=17,column=2,sticky='W',pady=7)
                
                l15=ttk.Label(self,text="14.  MEDIUM",font=("Times 30 italic"),background="powderblue")
                l15.grid(row=18,column=0,sticky='W',pady=7)
                s15=ttk.Label(self,text=" : ",background="powderblue")
                s15.grid(row=18,column=1,sticky='W',pady=7)
                a15=ttk.Label(self,text="ENGLISH",font=("Times 30 italic"),background="powderblue")
                a15.grid(row=18,column=2,sticky='W',pady=7)
                
                l16=ttk.Label(self,text="SEAL",font=("Times 30 italic"),background="powderblue")
                l16.grid(row=19,column=0,sticky='W',pady=7)
                s16=ttk.Label(self,text=" : ",background="powderblue")
                s16.grid(row=19,column=1,sticky='W',pady=7)
                a16=ttk.Label(self,text="",font=("Times 30 italic"),background="powderblue")
                a16.grid(row=19,column=2,sticky='W',pady=7)
                buttt=ttk.Button(self,text="print",style = 'W.TButton',command=self.printpdf)
                buttt.grid(row=20,column=1,sticky='W',pady=7)
                button17=ttk.Button(self,text="<-",style = 'B.TButton',command=lambda: master.switch_frame(transfersingle))
                button17.grid(row=20,column=0,sticky='W',pady=7)
                        
   def printpdf(self):
         
                db.execute("select *from studentdetails where reg_number=%s"%(setentvar))
                f=db.fetchone()
                reg=setentvar
                
                pdf=FPDF()
                pdf.add_page()
                pdf.set_font("Arial",size=12)
                name="alagappa.jpg"
                pdf.image(name,w=200,h=20)
                
                pdf.cell(100,10,txt="                                                              TRANSFER CERTIFICATE")
                pdf.ln(10)
                pdf.cell(100,10,txt="     1.  Name of the student")
                pdf.cell(100,10,txt="              "+str(f[0]))
                pdf.ln(10)
                pdf.cell(100,10,txt="     2.  Name of the parent/Guardian")
                pdf.cell(100,10,txt="              "+str(f[3]))
                pdf.ln(10)
                pdf.cell(100,10,txt="     3.  Nationality Religion and coommunity")
                pdf.cell(100,10,txt="              "+str(f[6]))
                pdf.ln(10)
                pdf.cell(100,10,txt="     4.  Sex")
                pdf.cell(100,10,txt="              "+str(f[8]))
                pdf.ln(10)
                pdf.cell(100,10,txt="     5.  Date of Birth (in figure and words) ")
                pdf.cell(100,10,txt="              "+str(f[9]))
                pdf.ln(10)
                pdf.cell(100,10,txt="     as entered in the admission register")
                pdf.ln(10)
                pdf.cell(100,10,txt="     6.  Course of Study")
                pdf.cell(100,10,txt="              "+str(f[10]))
                pdf.ln(10)
                pdf.cell(100,10,txt="     7.  Date of Admission to this college")
                pdf.cell(100,10,txt="              "+str(f[12]))
                pdf.ln(10)
                pdf.cell(100,10,txt="     8.  a) Whether the Student has paid all the  ")
                pdf.cell(100,10,txt="")
                pdf.ln(10)
                pdf.cell(100,10,txt="     Fees due to the college ?")
                pdf.ln(10)
                pdf.cell(100,10,txt="     8.  b) Whether the Student was in receipt of ")
                pdf.cell(100,10,txt="")
                pdf.ln(10)
                pdf.cell(100,10,txt="     any scholarship")
                pdf.ln(10)
                pdf.cell(100,10,txt="     9.  Whether the Student has undergone ")
                
                pdf.cell(100,10,txt="")
                pdf.ln(10)
                pdf.cell(100,10,txt="     Medical inspection during the year")
                pdf.ln(10)
                pdf.cell(100,10,txt="     10.  Reasons for leaving the College")
                
                pdf.cell(100,10,txt="")
                pdf.ln(10)
                pdf.cell(100,10,txt="     11.  Date of Leaving")
                pdf.cell(100,10,txt="")
                pdf.ln(10)
                pdf.cell(100,10,txt="     12.  Date on which application for Transfer  ")
                pdf.cell(100,10,txt="")
                pdf.ln(10)
                pdf.cell(100,10,txt="     Certificate was made by the Student or On ")
                pdf.ln(10)
                pdf.cell(100,10,txt="     his/her behalf by Parent/Guardian")
                pdf.ln(10)
                pdf.cell(100,10,txt="     13.  Date of the Transfer Certificate")
                pdf.cell(100,10,txt="              "+str(f[23]))
                pdf.ln(10)
                pdf.cell(100,10,txt="     14.  Medium")
                pdf.cell(100,10,txt="              English")
                pdf.ln(10)
                pdf.cell(100,10,txt="     Seal")
                pdf.cell(100,10,txt="")
                pdf.ln(10)
                r=reg+".pdf"
                pdf.output(r)
                if os.path.exists(r):
                           os.startfile(r)


class editspecify(tk.Frame):
      def __init__(self, master):
             tk.Frame.__init__(self,master)
             self.configure(background='darkorchid')
             lablle1=ttk.Label(self,text="EDIT SPECIFICATION",font=("Times 30 italic"),background="darkorchid")
             lablle1.grid(row=0,column=0,sticky='NSWE',pady=50)
             lp=ttk.Label(self,text="Register number",font=("Times 30 italic"),background="darkorchid")
             lp.grid(row=1,column=0,sticky='W',padx=35,pady=30,ipady=3)
             self.ent=ttk.Entry(self,font=("Times 18 italic"))
             self.ent.grid(row=1,column=1,sticky='W',padx=5,pady=30,ipady=5)
             buton1=ttk.Button(self,text="submit",style = 'W.TButton',command=self.editz)
             buton1.grid(row=3,column=1,sticky='W',pady=50)
             
             button18=ttk.Button(self,text="BACK",style = 'W.TButton',command=lambda: master.switch_frame(pageOne))
             button18.grid(row=3,column=0,sticky='W',pady=50,padx=35)
      def editz(self):
          global setvar
          setvar=self.ent.get()
          
          db.execute("select *from studentdetails where reg_number='%s'"%(setvar))
          e=db.fetchone()
          if(e):
              self.master.switch_frame(edit)
          else:
              messagebox.showinfo("Edit Error", "Enter A Valid Register Number")
          
                                                     
class edit(tk.Frame):
              def __init__(self, master):
                    tk.Frame.__init__(self, master)
                    
                    
                    #self._rootwindow.bind('<Configure>', self.onResize)
                    self.configure(background='darkorchid')
                    db.execute("select *from studentdetails where reg_number='%s'"%(setvar))
                    e=db.fetchone()
                    if(e):
                        self.labe=ttk.Label(self,text="EDIT",font=("Times 25 italic"),background="darkorchid")
                        self.labe.grid(row=0,column=2,sticky='W',pady=5,padx=35)
                        
                        self.l1=ttk.Label(self,text="Name Of The Candidate",font=("Times 18 italic"),background="darkorchid")
                        self.l1.grid(row=1,column=0,sticky='W',pady=5,padx=35)
                        self.ll1=ttk.Label(self,text="  :  ",background="darkorchid")
                        self.ll1.grid(row=1,column=1,sticky='W',pady=5,padx=35)
                        self.e1=ttk.Entry(self,font=("Times 18 italic"))
                        self.e1.grid(row=1,column=2,sticky='W',padx=35,pady=3,ipady=3)
                        self.e1.insert(0,e[0])
                        self.l2=ttk.Label(self,text="Reg Number",font=("Times 18 italic"),background="darkorchid")
                        self.l2.grid(row=2,column=0,sticky='W',pady=5,padx=35)
                        self.ll2=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll2.grid(row=2,column=1,sticky='W',pady=5,padx=35)
                        self.e2=ttk.Entry(self,font=("Times 18 italic"))
                        self.e2.grid(row=2,column=2,sticky='W',padx=35,pady=3,ipady=3)
                        self.e2.insert(0,e[1])
                        
                        self.l3=ttk.Label(self,text="Roll Number",font=("Times 18 italic"),background="darkorchid")
                        self.l3.grid(row=3,column=0,sticky='W',pady=5,padx=35)
                        self.ll3=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll3.grid(row=3,column=1,sticky='W',pady=5,padx=35)
                        self.e3=ttk.Entry(self,font=("Times 18 italic"))
                        self.e3.grid(row=3,column=2,sticky='W',padx=35,pady=3,ipady=3)
                        self.e3.insert(0,e[2])
                        self.l4=ttk.Label(self,text="Father's Name",font=("Times 18 italic"),background="darkorchid")
                        self.l4.grid(row=4,column=0,sticky='W',pady=5,padx=35)
                        self.ll4=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll4.grid(row=4,column=1,sticky='W',pady=5,padx=35)
                        self.e4=ttk.Entry(self,font=("Times 18 italic"))
                        self.e4.grid(row=4,column=2,sticky='W',padx=35,pady=3,ipady=3)
                        self.e4.insert(0,e[3])
                        self.l5=ttk.Label(self,text="Nationality",font=("Times 18 italic"),background="darkorchid")
                        self.l5.grid(row=5,column=0,sticky='W',pady=5,padx=35)
                        self.ll5=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll5.grid(row=5,column=1,sticky='W',pady=5,padx=35)
                        self.e5 = tk.StringVar(self)
                        self.pp1 =ttk.Combobox(self, width=25, textvariable=self.e5)
                        self.pp1['values']=(e[4],'Indian','Others')
                        self.pp1.grid(row=5,column=2,sticky='W',padx=38,pady=5,ipadx=33,ipady=5)
                        self.pp1.current(0)
                                
                        
                        
                        
                        self.l6=ttk.Label(self,text="Religion",font=("Times 18 italic"),background="darkorchid")
                        self.l6.grid(row=6,column=0,sticky='W',pady=5,padx=35)
                        self.ll6=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll6.grid(row=6,column=1,sticky='W',pady=5,padx=35)
                        
                        self.e6 = tk.StringVar(self)
                        self.p2 =ttk.Combobox(self, width=25, textvariable=self.e6)
                        self.p2['values']=(e[5],'Hindu','Musim','Christian','Others')
                        self.p2.grid(row=6,column=2,sticky='W',padx=38,pady=5,ipadx=33,ipady=5)
                        self.p2.current(0)
                        
                        
                        self.l7=ttk.Label(self,text="Caste",font=("Times 18 italic"),background="darkorchid")
                        self.l7.grid(row=7,column=0,sticky='W',pady=5,padx=35)
                        self.ll7=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll7.grid(row=7,column=1,sticky='W',pady=5,padx=35)
                        self.e7=ttk.Entry(self,font=("Times 18 italic"))
                        self.e7.grid(row=7,column=2,sticky='W',padx=35,pady=3,ipady=3)
                        self.e7.insert(0,e[7])
                        
                            
                        
                        self.l8=ttk.Label(self,text="CommunityCaste",font=("Times 18 italic"),background="darkorchid")
                        self.l8.grid(row=8,column=0,sticky='W',pady=5,padx=35)
                        self.ll8=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll8.grid(row=8,column=1,sticky='W',pady=5,padx=35)
                        self.e8 = tk.StringVar(self)
                        self.p3 =ttk.Combobox(self, width=25, textvariable=self.e8)
                        self.p3['values']=(e[6],'OC','BC','BCM','MBC','SC','ST','Others')  
                        self.p3.grid(row=8,column=2,sticky='W',padx=38,pady=5,ipadx=33,ipady=5)
                        self.p3.current(0)
                            
                        
                        self.l9=ttk.Label(self,text="Sex",font=("Times 18 italic"),background="darkorchid")
                        self.l9.grid(row=9,column=0,sticky='W',pady=5,padx=35)
                        self.ll9=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll9.grid(row=9,column=1,sticky='W',pady=5,padx=35)
                        self.e9 = tk.StringVar(self)
                        self.p4 =ttk.Combobox(self, width=25, textvariable=self.e9)
                        self.p4['values']=(e[8],'Male','Female','Others')  
                        self.p4.grid(row=9,column=2,sticky='W',padx=38,pady=5,ipadx=33,ipady=5)
                        self.p4.current(0)
                        
                        
                        self.l10=ttk.Label(self,text="Date Of Birth",font=("Times 18 italic"),background="darkorchid")
                        self.l10.grid(row=10,column=0,sticky='W',pady=5,padx=35)
                        self.ll10=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll10.grid(row=10,column=1,sticky='W',pady=5,padx=35)
                        self.e10=ttk.Entry(self,font=("Times 18 italic"))
                        self.e10.grid(row=10,column=2,sticky='W',padx=35,pady=3,ipady=3)
                        self.e10.insert(0,e[9])
                        self.l11=ttk.Label(self,text="Course Of Study",font=("Times 18 italic"),background="darkorchid")
                        self.l11.grid(row=11,column=0,sticky='W',pady=5,padx=35)
                        self.ll11=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll11.grid(row=11,column=1,sticky='W',pady=5,padx=35)
                        self.e11 = tk.StringVar(self)
                        self.p5 =ttk.Combobox(self, width=25, textvariable=self.e11)
                        self.p5['values']=(e[10],'BE','ME','BE Parttime')  
                        self.p5.grid(row=11,column=2,sticky='W',padx=38,pady=5,ipadx=33,ipady=5)
                        self.p5.current(0)
                                
                        
                        
                        self.l12=ttk.Label(self,text="Branch Of Study",font=("Times 18 italic"),background="darkorchid")
                        self.l12.grid(row=12,column=0,sticky='W',pady=5,padx=35)
                        self.ll12=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll12.grid(row=12,column=1,sticky='W',pady=5,padx=35)
                        
                        self.e12 = tk.StringVar(self)
                        self.p6 =ttk.Combobox(self, width=25, textvariable=self.e12)
                        self.p6['values']=(e[11],'Civil Engineering','Mechanical Engineering','Electrical and Electronics Engineering','Electronics and Communication Engineering','Computer science and Engineering')  
                        self.p6.grid(row=12,column=2,sticky='W',padx=38,pady=5,ipadx=33,ipady=5)
                        self.p6.current(0)
                        
                        
                        self.l13=ttk.Label(self,text=" Admitted On",font=("Times 18 italic"),background="darkorchid")
                        self.l13.grid(row=1,column=4,sticky='W',pady=5,padx=35)
                        self.ll13=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll13.grid(row=1,column=5,sticky='W',pady=5,padx=35)
                        self.e13=ttk.Entry(self,font=("Times 18 italic"))
                        self.e13.grid(row=1,column=6,sticky='W',padx=35,pady=3,ipady=3)
                        self.e13.insert(0,e[12])
                        self.l14=ttk.Label(self,text="Receipt Number",font=("Times 18 italic"),background="darkorchid")
                        self.l14.grid(row=2,column=4,sticky='W',pady=5,padx=35)
                        self.ll14=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll14.grid(row=2,column=5,sticky='W',pady=5,padx=35)
                        self.e14=ttk.Entry(self,font=("Times 18 italic"))
                        self.e14.grid(row=2,column=6,sticky='W',padx=35,pady=3,ipady=3)
                        self.e14.insert(0,e[13])
                        self.l15=ttk.Label(self,text="Receipt Date",font=("Times 18 italic"),background="darkorchid")
                        self.l15.grid(row=3,column=4,sticky='W',pady=5,padx=35)
                        self.ll15=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll15.grid(row=3,column=5,sticky='W',pady=5,padx=35)
                        self.e15=ttk.Entry(self,font=("Times 18 italic"))
                        self.e15.grid(row=3,column=6,sticky='W',padx=35,pady=3,ipady=3)
                        self.e15.insert(0,e[14])
                        self.l16=ttk.Label(self,text="Mother Tongue",font=("Times 18 italic"),background="darkorchid")
                        self.l16.grid(row=4,column=4,sticky='W',pady=5,padx=35)
                        self.ll16=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll16.grid(row=4,column=5,sticky='W',pady=5,padx=35)
                        
                        self.e16 = tk.StringVar(self)
                        self.p7 =ttk.Combobox(self, width=25, textvariable=self.e16)
                        self.p7['values']=(e[15],'Tamil','Hindi','Telugu','Kannada','Malayalam','Sourashtra','Others')  
                        self.p7.grid(row=4,column=6,sticky='W',padx=38,pady=5,ipadx=33,ipady=5)
                        self.p7.current(0)
                        
                        
                        self.l17=ttk.Label(self,text="Name Of State",font=("Times 18 italic"),background="darkorchid")
                        self.l17.grid(row=5,column=4,sticky='W',pady=5,padx=35)
                        self.ll17=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll17.grid(row=5,column=5,sticky='W',pady=5,padx=35)
                        
                        self.e17 = tk.StringVar(self)
                        self.p8 =ttk.Combobox(self, width=25, textvariable=self.e17)
                        self.p8['values']=(e[16],'Andhra Pradesh','Arunachal Pradesh','Assam','Bihar','Chhattisgarh','Goa','Gujarat','Haryana','Himachal Pradesh','Jammu and Kashmir','Jharkhand','Karnataka','Kerala','Madya Pradesh','Maharashtra','Manipur','Meghalaya','Mizoram','Nagaland','Orissa','Punjab','Rajasthan','Sikkim','Tamil Nadu','Telagana','Tripura','Uttaranchal','Uttar Pradesh','West Bengal','Andaman and Nicobar','Chandigarh','Dadar and Nagar Haveli','Daman and Diu','Delhi','Lakshadeep','Pondicherry')  
                        self.p8.grid(row=5,column=6,sticky='W',padx=38,pady=5,ipadx=33,ipady=5)
                        self.p8.current(0)
                                
                        
                        self.l18=ttk.Label(self,text="Present address ",font=("Times 18 italic"),background="darkorchid")
                        self.l18.grid(row=6,column=4,sticky='W',pady=5,padx=35)
                        self.ll18=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll18.grid(row=6,column=5,sticky='W',pady=5,padx=35)
                        self.e18=ttk.Entry(self,font=("Times 18 italic"))
                        self.e18.grid(row=6,column=6,sticky='W',padx=35,pady=3,ipady=3)
                        self.e18.insert(0,e[17])
                        self.l19=ttk.Label(self,text="Taluk ",font=("Times 18 italic"),background="darkorchid")
                        self.l19.grid(row=7,column=4,sticky='W',pady=5,padx=35)
                        self.ll19=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll19.grid(row=7,column=5,sticky='W',pady=5,padx=35)
                        self.e19=ttk.Entry(self,font=("Times 18 italic"))
                        self.e19.grid(row=7,column=6,sticky='W',padx=35,pady=3,ipady=3)
                        self.e19.insert(0,e[18])
                        self.l20=ttk.Label(self,text="City",font=("Times 18 italic"),background="darkorchid")
                        self.l20.grid(row=8,column=4,sticky='W',pady=5,padx=35)
                        self.ll20=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll20.grid(row=8,column=5,sticky='W',pady=5,padx=35)
                        self.e20=ttk.Entry(self,font=("Times 18 italic"))
                        self.e20.grid(row=8,column=6,sticky='W',padx=35,pady=3,ipady=3)
                        self.e20.insert(0,e[19])
                        self.l21=ttk.Label(self,text="District",font=("Times 18 italic"),background="darkorchid")
                        self.l21.grid(row=9,column=4,sticky='W',pady=5,padx=35)
                        self.ll21=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.ll21.grid(row=9,column=5,sticky='W',pady=5,padx=35)
                        
                        self.e21= tk.StringVar(self)
                        self.p9 =ttk.Combobox(self, width=25, textvariable=self.e21)
                        self.p9['values']=(e[20],'Ariyalur','Chengalpet','Chennai','Coimbatore','Cuddalore','Dharmapuri','Dindigul','Erode','Kallakurichi','Kancheepuram','Karur','Krishnagiri','Madurai','Nagapattinam','Nilgiris','Kanyakumari','Namakkal','Perambalur','Pudukottai','Ramanathapuram','Ranipet','Salem','Sivagangai','Tenkasi','Thanjavur','Theni','Thiruvallur','Thiruvarur','Tuticorin','Trichirappalli','Thirunelveli','Tirupattur','Tiruppur','Thiruvannamalai','Vellore','Viluppuram','Virudhunagar')  
                        self.p9.grid(row=9,column=6,sticky='W',padx=38,pady=5,ipadx=33,ipady=5)
                        self.p9.current(0)
                        
                        
                        self.la22=ttk.Label(self,text="Cell Number",font=("Times 18 italic"),background="darkorchid")
                        self.la22.grid(row=10,column=4,sticky='W',pady=5,padx=35)
                        self.lla22=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.lla22.grid(row=10,column=5,sticky='W',pady=5,padx=35)
                        self.ea22=ttk.Entry(self,font=("Times 18 italic"))
                        self.ea22.grid(row=10,column=6,sticky='W',padx=35,pady=3,ipady=3)
                        self.ea22.insert(0,e[21])
                        self.la23=ttk.Label(self,text="Aadhar Number",font=("Times 18 italic"),background="darkorchid")
                        self.la23.grid(row=11,column=4,sticky='W',pady=5,padx=35)
                        self.lla23=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.lla23.grid(row=11,column=5,sticky='W',pady=5,padx=35)
                        self.ea23=ttk.Entry(self,font=("Times 18 italic"))
                        self.ea23.grid(row=11,column=6,sticky='W',padx=35,pady=3,ipady=3)
                        self.ea23.insert(0,e[22])
                         
                        self.la24=ttk.Label(self,text="T.C.No",font=("Times 18 italic"),background="darkorchid")
                        self.la24.grid(row=12,column=4,sticky='W',pady=5,padx=35)
                        self.lla24=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.lla24.grid(row=12,column=5,sticky='W',pady=5,padx=35)
                        self.ea24=ttk.Entry(self,font=("Times 18 italic"))
                        self.ea24.grid(row=12,column=6,sticky='W',padx=35,pady=3,ipady=3)
                        self.ea24.insert(0,e[23]) 
                        self.la25=ttk.Label(self,text="Issued ON",font=("Times 18 italic"),background="darkorchid")
                        self.la25.grid(row=13,column=0,sticky='W',pady=5,padx=35)
                        self.lla25=ttk.Label(self,text="  :  ",font=("Times 18 italic"),background="darkorchid")
                        self.lla25.grid(row=13,column=1,sticky='W',pady=5,padx=35)
                        self.ea25=ttk.Entry(self,font=("Times 18 italic"))
                        self.ea25.grid(row=13,column=2,sticky='W',padx=35,pady=3,ipady=3)
                        self.ea25.insert(0,e[24])
                        self.llla24=ttk.Button(self,style = 'W.TButton',text="SUBMIT",command=self.valme)
                        self.llla24.grid(row=13,column=6,sticky='W',padx=35,pady=3,ipady=3)
                        self.button19=ttk.Button(self,text="BACK",style = 'W.TButton',command=lambda: master.switch_frame(editspecify))
                        self.button19.grid(row=13,column=4,sticky='W',pady=5,padx=35)
                        
                        #button19=tk.Button(self,text="<-",style = 'B.TButton',command=lambda: master.switch_frame(editspecify))
                        #button19.grid(row=23,column=0,sticky='W')
                        
        
                        print(self.e1.get())
                    

              def valme(self):
                regex1 = '^\d{4}\d{4}\d{4}$'
               
                regex3 = '(0/91)?[7-9][0-9]{9}'
                print(self.e13.get())
                
                #day,month,year = self.dob.split('-')
                #d,m,y=self.receiptdate.split('-')
                #d1,m1,y1=self.issuedon.split('-')
                try:
                    if(re.search(regex1,self.ea23.get())):  
                         if(re.search(regex3,self.ea22.get())):  
                            datetime.datetime.strptime(self.e10.get(), '%Y-%m-%d')
                            datetime.datetime.strptime(self.e13.get(), '%Y-%m-%d')
                            datetime.datetime.strptime(self.e15.get(), '%Y-%m-%d')
                            if(self.e1.get()!="" and self.e3.get()!="" and self.e4.get()!="" and self.e5.get()!="" and self.e6.get()!="" and self.e7.get()!="" and self.e8.get() !="" and self.e9.get(),self.e10.get(),self.e11.get(),self.e12.get(),self.e13.get(),self.e14.get(),self.e15.get()!="" and self.e16.get()!="" and self.e17.get() !="" and self.e18.get()!="" and self.e19.get()!="" and self.e20.get()!="" and self.e21.get()!="" and self.ea22.get()!="" and self.ea23.get()!="" ):
                                self.callme()
                            else:
                             messagebox.showinfo("Insert Error", "Fill all the fields")
                                
                         else:  
                             messagebox.showinfo("Insert Error", "Not valid phone no")
                    else:  
                        messagebox.showinfo("Insert Error", "Not valid aadhar no") 
                    
                    
                    

                except ValueError :
                    messagebox.showinfo("Insert Error", "Not valid date no")

                
                  
              def callme(self):
                        sql="UPDATE `studentdetails` SET `name`=%s,`father_name`=%s,`nationality`=%s,`religion`=%s,`caste`=%s,`community`=%s,`sex`=%s,`dateofbirth`=%s,`course`=%s,`branch`=%s,`admittedon`=%s,`receiptno`=%s,`receiptdate`=%s,`mothertongue`=%s,`state`=%s,`present_address`=%s,taluk=%s,`city`=%s,`district`=%s,`cell_number`=%s,`aadhar_number`=%s,`tcno`=%s,`issuedon`=%s WHERE reg_number=%s"
                        val=(self.e1.get(),self.e4.get(),self.e5.get(),self.e6.get(),self.e8.get(),self.e7.get(),self.e9.get(),self.e10.get(),self.e11.get(),self.e12.get(),self.e13.get(),self.e14.get(),self.e15.get(),self.e16.get(),self.e17.get(),self.e18.get(),self.e19.get(),self.e20.get(),self.e21.get(),self.ea22.get(),self.ea23.get(),self.ea24.get(),self.ea25.get(),setvar)
                        db.execute(sql, val)
                        db_cur.commit()
                        self.master.switch_frame(pageOne)
                
           
class newreg(tk.Frame):
              def __init__(self, master):
                tk.Frame.__init__(self,master)
                self.configure(background='darkorchid')
                #db.execute("select *from studentdetails where reg_number='%s'"%(setvar))
                #e=db.fetchone()
                labe=tk.Label(self,text="REGISTER",font=("Times 30 italic"),background="darkorchid")
                labe.grid(row=0,column=2,sticky='W',padx=20)
                global name,regno,rollno,taluk,father,nation,religion,caste,community,sex,dob,course,branch,admiton,receiptno,mothertongue,state,address,city,district,cellno,aadharno,tcno,issuedon
                self.l1=ttk.Label(self,text="Name Of The Candidate",font=("Times 15 italic"),background="darkorchid")
                self.l1.grid(row=1,column=0,sticky='W',pady=3,padx=40)
                self.ll1=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll1.grid(row=1,column=1,sticky='W',padx=40)
                self.e1=ttk.Entry(self,font=("Times 15 italic"))
                self.e1.grid(row=1,column=2,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l2=ttk.Label(self,text="Reg Number",font=("Times 15 italic"),background="darkorchid")
                self.l2.grid(row=2,column=0,sticky='W',pady=3,padx=40)
                self.ll2=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll2.grid(row=2,column=1,sticky='W',pady=3,padx=40)
                self.e2=ttk.Entry(self,font=("Times 15 italic"))
                self.e2.grid(row=2,column=2,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l3=ttk.Label(self,text="Roll Number",font=("Times 15 italic"),background="darkorchid")
                self.l3.grid(row=3,column=0,sticky='W',pady=3,padx=40)
                self.ll3=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll3.grid(row=3,column=1,sticky='W',pady=3,padx=40)
                self.e3=ttk.Entry(self,font=("Times 15 italic"))
                self.e3.grid(row=3,column=2,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l4=ttk.Label(self,text="Father's Name",font=("Times 15 italic"),background="darkorchid")
                self.l4.grid(row=4,column=0,sticky='W',pady=3,padx=40)
                self.ll4=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll4.grid(row=4,column=1,sticky='W',pady=3,padx=40)
                self.e4=ttk.Entry(self,font=("Times 15 italic"))
                self.e4.grid(row=4,column=2,sticky='W',padx=40,pady=3,ipady=3)
            
                self.l5=ttk.Label(self,text="Nationality",font=("Times 15 italic"),background="darkorchid")
                self.l5.grid(row=5,column=0,sticky='W',pady=3,padx=40)
                self.ll5=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll5.grid(row=5,column=1,sticky='W',pady=3,padx=40)
                self.t1 = tk.StringVar(self)
                self.p1 =ttk.Combobox(self, width=25, textvariable=self.t1)
                self.p1['values']=('Indian','Others')
                self.p1.grid(row=5,column=2,sticky='W',padx=40,pady=5,ipadx=16,ipady=5)
                self.p1.current()
                
                self.l6=ttk.Label(self,text="Religion",font=("Times 15 italic"),background="darkorchid")
                self.l6.grid(row=6,column=0,sticky='W',pady=3,padx=40)
                self.ll6=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll6.grid(row=6,column=1,sticky='W',pady=3,padx=40)
                self.t2 = tk.StringVar(self)
                self.p2 =ttk.Combobox(self, width=25, textvariable=self.t2)
                self.p2['values']=('Hindu','Musim','Christian','Others')
                self.p2.grid(row=6,column=2,sticky='W',padx=40,pady=5,ipadx=16,ipady=5)
                self.p2.current()
                
                self.l7=ttk.Label(self,text="Caste",font=("Times 15 italic"),background="darkorchid")
                self.l7.grid(row=7,column=0,sticky='W',pady=3,padx=40)
                self.ll7=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll7.grid(row=7,column=1,sticky='W',pady=3,padx=40)
                self.e7=ttk.Entry(self,font=("Times 15 italic"))
                self.e7.grid(row=7,column=2,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l8=ttk.Label(self,text="Community",font=("Times 15 italic"),background="darkorchid")
                self.l8.grid(row=8,column=0,sticky='W',pady=3,padx=40)
                self.ll8=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll8.grid(row=8,column=1,sticky='W',pady=3,padx=40)
                self.t3 = tk.StringVar(self)
                self.p3 =ttk.Combobox(self, width=25, textvariable=self.t3)
                self.p3['values']=('OC','BC','BCM','MBC','SC','ST','Others')  
                self.p3.grid(row=8,column=2,sticky='W',padx=40,pady=5,ipadx=16,ipady=5)
                self.p3.current()
                
                self.l9=ttk.Label(self,text="Gender",font=("Times 15 italic"),background="darkorchid")
                self.l9.grid(row=9,column=0,sticky='W',pady=3,padx=40)
                self.ll9=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll9.grid(row=9,column=1,sticky='W',pady=3,padx=40)
                self.t4 = tk.StringVar(self)
                self.p4 =ttk.Combobox(self, width=25, textvariable=self.t4)
                self.p4['values']=('Male','Female','Others')  
                self.p4.grid(row=9,column=2,sticky='W',padx=40,pady=5,ipadx=16,ipady=5)
                self.p4.current()
                
                self.l10=ttk.Label(self,text="Date Of Birth",font=("Times 15 italic"),background="darkorchid")
                self.l10.grid(row=10,column=0,sticky='W',pady=3,padx=40)
                self.ll10=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll10.grid(row=10,column=1,sticky='W',pady=3,padx=40)
                self.e10=ttk.Entry(self,font=("Times 15 italic"))
                self.e10.grid(row=10,column=2,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l11=ttk.Label(self,text="Course Of Study",font=("Times 15 italic"),background="darkorchid")
                self.l11.grid(row=11,column=0,sticky='W',pady=3,padx=40)
                self.ll11=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll11.grid(row=11,column=1,sticky='W',pady=3,padx=40)
                self.t5 = tk.StringVar(self)
                self.p5 =ttk.Combobox(self, width=25, textvariable=self.t5)
                self.p5['values']=('BE','ME','BE Parttime')  
                self.p5.grid(row=11,column=2,sticky='W',padx=40,pady=5,ipadx=16,ipady=5)
                self.p5.current()
                
                self.l12=ttk.Label(self,text="Branch Of Study",font=("Times 15 italic"),background="darkorchid")
                self.l12.grid(row=1,column=4,sticky='W',pady=3,padx=40)
                self.ll12=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll12.grid(row=1,column=5,sticky='W',pady=3,padx=40)
                self.t6 = tk.StringVar(self)
                self.p6 =ttk.Combobox(self, width=25, textvariable=self.t6)
                self.p6['values']=('Civil Engineering','Mechanical Engineering','Electrical and Electronics Engineering','Electronics and Communication Engineering','Computer science and Engineering')  
                self.p6.grid(row=1,column=6,sticky='W',padx=40,pady=5,ipadx=16,ipady=5)
                self.p6.current()
                
                self.l13=ttk.Label(self,text="Admitted On",font=("Times 15 italic"),background="darkorchid")
                self.l13.grid(row=2,column=4,sticky='W',pady=3,padx=40)
                self.ll13=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll13.grid(row=2,column=5,sticky='W',pady=3,padx=40)
                self.e13=ttk.Entry(self,font=("Times 15 italic"))
                self.e13.grid(row=2,column=6,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l14=ttk.Label(self,text="Receipt Number",font=("Times 15 italic"),background="darkorchid")
                self.l14.grid(row=3,column=4,sticky='W',pady=3,padx=40)
                self.ll14=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll14.grid(row=3,column=5,sticky='W',pady=3,padx=40)
                self.e14=ttk.Entry(self,font=("Times 15 italic"))
                self.e14.grid(row=3,column=6,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l15=ttk.Label(self,text="Receipt Date",font=("Times 15 italic"),background="darkorchid")
                self.l15.grid(row=4,column=4,sticky='W',pady=3,padx=40)
                self.ll15=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll15.grid(row=4,column=5,sticky='W',pady=3,padx=40)
                self.e15=ttk.Entry(self,font=("Times 15 italic"))
                self.e15.grid(row=4,column=6,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l16=ttk.Label(self,text="Mother Tongue",font=("Times 15 italic"),background="darkorchid")
                self.l16.grid(row=5,column=4,sticky='W',pady=3,padx=40)
                self.ll16=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll16.grid(row=5,column=5,sticky='W',pady=3,padx=40)
                self.t7 = tk.StringVar(self)
                self.p7 =ttk.Combobox(self, width=25, textvariable=self.t7)
                self.p7['values']=('Tamil','Hindi','Telugu','Kannada','Malayalam','Sourashtra','Others')  
                self.p7.grid(row=5,column=6,sticky='W',padx=40,pady=5,ipadx=16,ipady=5)
                self.p7.current()
                
                self.l17=ttk.Label(self,text="Name Of State",font=("Times 15 italic"),background="darkorchid")
                self.l17.grid(row=6,column=4,sticky='W',pady=3,padx=40)
                self.ll17=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll17.grid(row=6,column=5,sticky='W',pady=3,padx=40)
                self.t8 = tk.StringVar(self)
                self.p8 =ttk.Combobox(self, width=25, textvariable=self.t8)
                self.p8['values']=('Andhra Pradesh','Arunachal Pradesh','Assam','Bihar','Chhattisgarh','Goa','Gujarat','Haryana','Himachal Pradesh','Jammu and Kashmir','Jharkhand','Karnataka','Kerala','Madya Pradesh','Maharashtra','Manipur','Meghalaya','Mizoram','Nagaland','Orissa','Punjab','Rajasthan','Sikkim','Tamil Nadu','Telagana','Tripura','Uttaranchal','Uttar Pradesh','West Bengal','Andaman and Nicobar','Chandigarh','Dadar and Nagar Haveli','Daman and Diu','Delhi','Lakshadeep','Pondicherry')  
                self.p8.grid(row=6,column=6,sticky='W',padx=40,pady=5,ipadx=16,ipady=5)
                self.p8.current()
                
                self.l18=ttk.Label(self,text="Present address ",font=("Times 15 italic"),background="darkorchid")
                self.l18.grid(row=7,column=4,sticky='W',pady=3,padx=40)
                self.ll18=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll18.grid(row=7,column=5,sticky='W',pady=3,padx=40)
                self.e18=ttk.Entry(self,font=("Times 15 italic"))
                self.e18.grid(row=7,column=6,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l19=ttk.Label(self,text="Taluk ",font=("Times 15 italic"),background="darkorchid")
                self.l19.grid(row=8,column=4,sticky='W',pady=3,padx=40)
                self.ll19=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll19.grid(row=8,column=5,sticky='W',pady=3,padx=40)
                self.e19=ttk.Entry(self,font=("Times 15 italic"))
                self.e19.grid(row=8,column=6,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l20=ttk.Label(self,text="City",font=("Times 15 italic"),background="darkorchid")
                self.l20.grid(row=9,column=4,sticky='W',pady=3,padx=40)
                self.ll20=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll20.grid(row=9,column=5,sticky='W',pady=3,padx=40)
                self.e20=ttk.Entry(self,font=("Times 15 italic"))
                self.e20.grid(row=9,column=6,sticky='W',padx=40,pady=3,ipady=3)
                
                self.l21=ttk.Label(self,text="District",font=("Times 15 italic"),background="darkorchid")
                self.l21.grid(row=10,column=4,sticky='W',pady=3,padx=40)
                self.ll21=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.ll21.grid(row=10,column=5,sticky='W',pady=3,padx=40)
                self.t9 = tk.StringVar(self)
                self.p9 =ttk.Combobox(self, width=25, textvariable=self.t9)
                self.p9['values']=('Ariyalur','Chengalpet','Chennai','Coimbatore','Cuddalore','Dharmapuri','Dindigul','Erode','Kallakurichi','Kancheepuram','Karur','Krishnagiri','Madurai','Nagapattinam','Nilgiris','Kanyakumari','Namakkal','Perambalur','Pudukottai','Ramanathapuram','Ranipet','Salem','Sivagangai','Tenkasi','Thanjavur','Theni','Thiruvallur','Thiruvarur','Tuticorin','Trichirappalli','Thirunelveli','Tirupattur','Tiruppur','Thiruvannamalai','Vellore','Viluppuram','Virudhunagar')  
                self.p9.grid(row=10,column=6,sticky='W',padx=40,pady=5,ipadx=16,ipady=5)
                self.p9.current()
            
                self.la22=ttk.Label(self,text="Mobile Number",font=("Times 15 italic"),background="darkorchid")
                self.la22.grid(row=11,column=4,sticky='W',pady=3,padx=40)
                self.lla22=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.lla22.grid(row=11,column=5,sticky='W',pady=3,padx=40)
                self.ea22=ttk.Entry(self,font=("Times 15 italic"))
                self.ea22.grid(row=11,column=6,sticky='W',padx=40,pady=3,ipady=3)
                
                self.la23=ttk.Label(self,text="Aadhar Number",font=("Times 15 italic"),background="darkorchid")
                self.la23.grid(row=12,column=4,sticky='W',pady=3,padx=40)
                self.lla23=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.lla23.grid(row=12,column=5,sticky='W',pady=3,padx=40)
                self.ea23=ttk.Entry(self,font=("Times 15 italic"))
                self.ea23.grid(row=12,column=6,sticky='W',padx=40,pady=3,ipady=3)
                
                 
                self.la24=ttk.Label(self,text="T.C.No",font=("Times 15 italic"),background="darkorchid")
                self.la24.grid(row=13,column=4,sticky='W',pady=3,padx=40)
                self.lla24=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.lla24.grid(row=13,column=5,sticky='W',pady=3,padx=40)
                self.ea24=ttk.Entry(self,font=("Times 15 italic"))
                self.ea24.grid(row=13,column=6,sticky='W',padx=40,pady=3,ipady=3)
                 
                self.la25=ttk.Label(self,text="Issued ON",font=("Times 15 italic"),background="darkorchid")
                self.la25.grid(row=12,column=0,sticky='W',pady=3,padx=40)
                self.lla25=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.lla25.grid(row=12,column=1,sticky='W',pady=3,padx=40)
                self.ea25=ttk.Entry(self,font=("Times 15 italic"))
                self.ea25.grid(row=12,column=2,sticky='W',padx=40,pady=3,ipady=3)
                
                self.la26=ttk.Label(self,text="Year Of PassOut",font=("Times 15 italic"),background="darkorchid")
                self.la26.grid(row=13,column=0,sticky='W',pady=3,padx=40)
                self.lla26=ttk.Label(self,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                self.lla26.grid(row=13,column=1,sticky='W',pady=3,padx=40)
                self.ea26=ttk.Entry(self,font=("Times 15 italic"))
                self.ea26.grid(row=13,column=2,sticky='W',padx=40,pady=3,ipady=3)
                self.b1=ttk.Button(self,text="REGISTER",style = 'W.TButton',command=self.validatered)
                self.b1.grid(row=14,column=6,sticky='W',padx=40,pady=3,ipady=3)
                
                
                
                self.but=ttk.Button(self,style = 'W.TButton',text="BACK",command=lambda: master.switch_frame(pageOne))
                self.but.grid(row=14,column=4,sticky='W',pady=3,padx=40)
          
          
              def validatered(self):
                self.name=self.e1.get()
                self.regno=self.e2.get()
                self.rollno=self.e3.get()
                self.father=self.e4.get()
                self.nation=self.t1.get()
                self.religion=self.t2.get()
                self.caste=self.e7.get()
                self.community=self.t3.get()
                self.sex=self.t4.get()
                self.dob=self.e10.get()
                self.course=self.t5.get()
                self.branch=self.t6.get()
                self.admiton=self.e13.get()
                self.receiptno=self.e14.get()
                self.receiptdate=self.e15.get()
                self.mothertongue=self.t7.get()
                self.state=self.t8.get()
                self.address=self.e18.get()
                self.taluk=self.e19.get()
                self.city=self.e20.get()  
                self.district=self.t9.get()
                self.cellno=self.ea22.get()
                self.aadharno=self.ea23.get()
                self.tcno=self.ea24.get()
                self.issuedon=self.ea25.get()
                self.yopo=self.ea26.get()
                 
                regex1 = '^\d{4}\d{4}\d{4}$'
               
                regex3 = '(0/91)?[7-9][0-9]{9}'
                
                
                #day,month,year = self.dob.split('-')
                #d,m,y=self.receiptdate.split('-')
                #d1,m1,y1=self.issuedon.split('-')
                try:
                    if(re.search(regex1,self.aadharno)):  
                         if(re.search(regex3,self.cellno)):  
                            datetime.datetime.strptime(self.dob, '%Y-%m-%d')
                            datetime.datetime.strptime(self.receiptdate, '%Y-%m-%d')
                            datetime.datetime.strptime(self.issuedon, '%Y-%m-%d')
                            if(self.name!="" and self.regno !="" and self.rollno !="" and self.father!="" and self.nation !="" and self.religion!="" and self.caste!="" and self.community!="" and self.sex!="" and self.dob!="" and self.course!="" and  self.branch !="" and self.admiton!="" and self.receiptno!="" and self.receiptdate!="" and self.mothertongue!="" and self.state!="" and self.address!="" and self.taluk!="" and self.city!="" and self.district!="" and self.cellno!="" and self.aadharno!="" and self.tcno!="" and self.issuedon!="" and self.yopo!=""):
                                self.calle()
                            else:
                             messagebox.showinfo("Insert Error", "Fill all the fields")
                                
                         else:  
                             messagebox.showinfo("Insert Error", "Not valid phone no")
                    else:  
                        messagebox.showinfo("Insert Error", "Not valid aadhar no") 
                    
                    
                    

                except ValueError :
                    messagebox.showinfo("Insert Error", "Not valid date no")

                
                
               
                    
                
                    
              def calle(self):
                
                self.name=self.e1.get()
                self.regno=self.e2.get()
                self.rollno=self.e3.get()
                self.father=self.e4.get()
                self.nation=self.t1.get()
                self.religion=self.t2.get()
                self.caste=self.e7.get()
                self.community=self.t3.get()
                self.sex=self.t4.get()
                self.dob=self.e10.get()
                self.course=self.t5.get()
                self.branch=self.t6.get()
                self.admiton=self.e13.get()
                self.receiptno=self.e14.get()
                self.receiptdate=self.e15.get()
                self.mothertongue=self.t7.get()
                self.state=self.t8.get()
                self.address=self.e18.get()
                self.taluk=self.e19.get()
                self.city=self.e20.get()  
                self.district=self.t9.get()
                self.cellno=self.ea22.get()
                self.aadharno=self.ea23.get()
                self.tcno=self.ea24.get()
                self.issuedon=self.ea25.get()
                self.yopo=self.ea26.get()
                print(self.community)
                #INSERT INTO `studentdetails`(`name`, `reg_number`, `roll_number`, `father_name`, `nationality`, `religion`, `caste`, `community`, `sex`, `dateofbirth`, `course`, `branch`, `admittedon`, `receiptno`, `receiptdate`, `mothertongue`, `state`, `present_address`, `city`, `district`, `cell_number`, `aadhar_number`, `tcno`, `issuedon`) 
                
                try:
                    sql1="INSERT INTO `studentdetails` (`name`, `reg_number`, `roll_number`, `father_name`, `nationality`, `religion`, `caste`, `community`, `sex`, `dateofbirth`, `course`, `branch`, `admittedon`, `receiptno`, `receiptdate`, `mothertongue`, `state`, `present_address`, `taluk`, `city`, `district`, `cell_number`, `aadhar_number`, `tcno`, `issuedon`, `Year_Of_Passout`)VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    val1=(self.name,self.regno,self.rollno,self.father,self.nation,self.religion,self.community,self.caste,self.sex,self.dob,self.course,self.branch,self.admiton,self.receiptno,self.receiptdate,self.mothertongue,self.state,self.address,self.taluk,self.city,self.district,self.cellno, self.aadharno,self.tcno,self.issuedon,self.yopo)
                    db.execute(sql1, val1)
                    db_cur.commit()
                    self.master.switch_frame(pageOne) 
                except mysql.connector.IntegrityError:
                       messagebox.showinfo("Registration Error", "Duplicate entry register number")

           
       
class search(tk.Frame):
                  def __init__(self, master):
                        tk.Frame.__init__(self,master)
                        self.regis=""
                        self.religion=""
                        self.caste=""
                        self.gen=""#self.eta3.get()
                        self.bran=""#self.eta4.get()
                        self.talu=""#self.eta5.get()
                        self.dist=""
                        self.configure(background='darkorchid')
                        self.la=ttk.Label(self,text="search",font=("Times 30 italic"),background='darkorchid')
                        self.la.grid(row=0,column=1,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                        self.la1=ttk.Label(self,text="Register",font=("Times 20 italic"),background='darkorchid')
                        self.la1.grid(row=1,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                        self.en =ttk.Entry(self,font=("Times 15 italic"))
                        self.en.grid(row=1,column=1,sticky='W',padx=20,pady=5,ipadx=40,ipady=5)

                        self.la1=ttk.Label(self,text="Religion",font=("Times 20 italic"),background='darkorchid')
                        self.la1.grid(row=2,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                        self.tkv1 = tk.StringVar(self)
                        self.popupMenu1 =ttk.Combobox(self, width=25, textvariable=self.tkv1,font=("Times 15 italic"))
                        self.popupMenu1['values']=('Hindu','Muslim','Christian','Others')
                        self.popupMenu1.grid(row=2,column=1,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                        self.popupMenu1.current()
                        self.la2=ttk.Label(self,text="Caste",font=("Times 20 italic"),background='darkorchid')
                        self.la2.grid(row=3,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                        self.tkv2 = tk.StringVar(self)
                        self.popupMenu2 =ttk.Combobox(self, width=25, textvariable=self.tkv2,font=("Times 15 italic"))
                        self.popupMenu2['values']=('OC','BC','BCM','MBC','SC','ST','Others')  
                        self.popupMenu2.grid(row=3,column=1,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                        self.popupMenu2.current()
                        self.la3=ttk.Label(self,text="Gender",font=("Times 20 italic"),background='darkorchid')
                        self.la3.grid(row=4,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                        self.tkv3 = tk.StringVar(self)
                        self.popupMenu3 =ttk.Combobox(self, width=25, textvariable=self.tkv3,font=("Times 15 italic"))
                        self.popupMenu3['values']=('Male','Female','Transgender')
                        self.popupMenu3.grid(row=4,column=1,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                        self.popupMenu3.current()
                        #self.la5=ttk.Label(self,text="Programme",font=("Times 20 italic"),background='darkorchid')
                       # self.la5.grid(row=5,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                       # self.tkv4 = tk.StringVar(self)
                        #self.popupMenu4=ttk.Combobox(self, width=25, textvariable=self.tkv4,font=("Times 15 italic"))
                        #self.popupMenu4['values']=('BE','ME','BE-Parttime')
                        #self.popupMenu4.grid(row=5,column=1,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                        #self.popupMenu4.current()
                        self.la4=ttk.Label(self,text="Branch Of Study",font=("Times 20 italic"),background='darkorchid')
                        self.la4.grid(row=6,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                        self.tkv5 = tk.StringVar(self)
                        self.popupMenu5=ttk.Combobox(self, width=25, textvariable=self.tkv5,font=("Times 15 italic"))
                        self.popupMenu5['values']=('Civil Engineering','Mechanical Engineering','Electrical and Electronics Engineering','Electronics and Communication Engineering','Computer Science and Engineering')
                        self.popupMenu5.grid(row=6,column=1,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                        self.popupMenu5.current()
                        self.la6=ttk.Label(self,text="Taluk",font=("Times 25 italic"),background='darkorchid')
                        self.la6.grid(row=7,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                        self.eta5=ttk.Entry(self,font=("Times 15 italic"))
                        self.eta5.grid(row=7,column=1,sticky='W',padx=20,pady=5,ipadx=37,ipady=5)
                        self.la7=ttk.Label(self,text="District",font=("Times 25 italic"),background='darkorchid')
                        self.la7.grid(row=5,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                        self.tkvar = tk.StringVar(self)
                        self.popupMenu =ttk.Combobox(self, width=27, textvariable=self.tkvar,font=("Times 15 italic"))
                        self.popupMenu['values']=('Ariyalur','Chengalpet','Chennai','Coimbatore','Cuddalore','Dharmapuri','Dindigul','Erode','Kallakurichi','Kancheepuram','Karur','Krishnagiri','Madurai','Nagapattinam','Nilgiris','Kanyakumari','Namakkal','Perambalur','Pudukottai','Ramanathapuram','Ranipet','Salem','Sivagangai','Tenkasi','Thanjavur','Theni','Thiruvallur','Thiruvarur','Tuticorin','Trichirappalli','Thirunelveli','Tirupattur','Tiruppur','Thiruvannamalai','Vellore','Viluppuram','Virudhunagar')
                        self.popupMenu.grid(row=5,column=1,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                        self.popupMenu.current()
                        
                        self.la8=ttk.Label(self,text="Year Of Passout",font=("Times 25 italic"),background='darkorchid')
                        self.la8.grid(row=8,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                        self.en1 =ttk.Entry(self,font=("Times 15 italic"))
                        self.en1.grid(row=8,column=1,sticky='W',padx=20,pady=5,ipadx=45,ipady=5)
                        self.l11=ttk.Label(self,text="Course Of Study",font=("Times 25 italic"),background="darkorchid")
                        self.l11.grid(row=9,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                        self.t5 = tk.StringVar(self)
                        self.p5 =ttk.Combobox(self, width=27, textvariable=self.t5,font=("Times 15 italic"))
                        self.p5['values']=('BE','ME','BE Parttime')  
                        self.p5.grid(row=9,column=1,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                        self.p5.current()
                        #self.eta6=ttk.Entry(self,font=("Times 10 italic"))
                        #self.eta6.grid(row=7,column=1,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                        self.buty=ttk.Button(self,text="search",style = 'W.TButton',command=self.say_hello)
                        self.buty.grid(row=10,column=1,sticky='W',padx=30,pady=10)
                        self.butxx=ttk.Button(self,style = 'W.TButton',text="Back",command=lambda: master.switch_frame(pageOne))
                        self.butxx.grid(row=10,column=0,sticky='W',padx=30,pady=10)
                 
                 
                  #def _login_btn_clickked(self):
                       # sex = self.religion #gets the value stored in gender and assigns it to sex
                       # print(sex)
                
                 
                      
                  def say_hello(self):
                      # reli,cast,seex,bos,tal,dis
                      self.regis=self.en.get()
                      self.yopo=self.en1.get()
                      self.reli=self.tkv1.get()
                      print(self.reli)
                      self.cast=self.tkv2.get()
                      print(self.cast)#self.eta2.get()
                      self.seex=self.tkv3.get()
                      print(self.seex)#self.eta3.get()
                      self.bos=self.tkv5.get()
                      print(self.bos)#self.eta4.get()
                      self.tal=self.eta5.get()
                      self.dis=self.tkvar.get()#self.eta6.get()
                      print(self.dis)
                      self.course=self.t5.get()
                      print(self.course)
                      wb = Workbook()
                      #global ff
                      #print(self.reli)
                      # 1 combo
                      if(self.course=="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' " %(self.reli))
                          #ff=db.fetchone()
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religion"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' " %(self.course))
                          #ff=db.fetchone()
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "course"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                      elif(self.regis!=""):
                          db.execute("select *from studentdetails where reg_number='%s' " %(self.regis))
                          #ff=db.fetchone()
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "register"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' " %(self.yopo))
                          #ff=db.fetchone()
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "YearOfPassOut"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          print("entered")
                          db.execute("select *from studentdetails where caste ='%s'" %(self.cast))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "caste"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex ='%s'" %(self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "gender"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where branch ='%s'" %(self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "Branch of study"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where taluk ='%s'" %(self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "Taluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district ='%s'" %(self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "District"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          #2combo
                      elif(self.course!="" and self.reli!=" " and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and course='%s'" %(self.reli,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli==" " and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste ='%s'" %(self.course,self.cast))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecaste"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli==" " and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course ='%s' and sex ='%s'" %(self.course,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli==" " and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course ='%s' and branch='%s'" %(self.course,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursebranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli==" " and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and taluk='%s'" %(self.course,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursetaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli==" " and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and district='%s'" %(self.course,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursedistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli==" " and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and Year_Of_Passout='%s'" %(self.course,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "courseyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

#2 combo la 7 completed
                          
                      elif(self.course=="" and self.reli!=" " and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and caste ='%s'" %(self.reli,self.cast))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncaste"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                      elif(self.course=="" and self.reli!="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and sex='%s'" %(self.reli,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiongender"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and branch='%s'" %(self.reli,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and taluk='%s'" %(self.reli,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and branch='%s'" %(self.reli,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion ='%s' and Year_Of_Passout='%s'" %(self.reli,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                    
                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s'" %(self.cast,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and branch='%s'" %(self.cast,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castebranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s'" %(self.cast,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and district='%s'" %(self.cast,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castedistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and Year_Of_Passout='%s'" %(self.cast,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "casteyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and branch='%s'" %(self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and taluk='%s'" %(self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s'" %(self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and Year_Of_Passout='%s'" %(self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where branch='%s' and taluk='%s'" %(self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where branch='%s' and district='%s'" %(self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where branch='%s' and Year_Of_Passout='%s'" %(self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)    
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where taluk='%s' and district='%s'" %(self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "talukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where taluk='%s' and Year_Of_Passout='%s'" %(self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "talukyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and district='%s'" %(self.yopo,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      
                    # 3 combo

                      elif(self.course!="" and self.reli!="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and course='%s'" %(self.reli,self.cast,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastecourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and course='%s' and sex='%s'" %(self.reli,self.course,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncoursesex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and course='%s' and branch='%s'" %(self.reli,self.course,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncoursebranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and course='%s' and taluk='%s'" %(self.reli,self.course,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncoursetaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and course='%s' and district='%s'" %(self.reli,self.course,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncoursedistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and course='%s' and Year_Of_Passout='%s'" %(self.reli,self.course,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncourseyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast !="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s'" %(self.course,self.cast,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and branch='%s'" %(self.course,self.cast,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastebranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and taluk='%s'" %(self.course,self.cast,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastetaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and district='%s'" %(self.course,self.cast,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastedistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecasteyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s'" %(self.course,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s'" %(self.course,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and district='%s'" %(self.course,self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                        

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.course,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)   
                          
                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and branch='%s' and taluk='%s'" %(self.course,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                         
                        
                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and branch='%s' and district='%s'" %(self.course,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursebranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.course,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursebranchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and taluk='%s' and district='%s'" %(self.course,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursetalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursetalukyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursedistrictyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)


# 2 combo 21 completed


  
                      elif(self.course=="" and self.reli!="" and self.cast !="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s'" %(self.reli,self.cast,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast !="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s'" %(self.reli,self.cast,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s'" %(self.reli,self.cast,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s'" %(self.reli,self.cast,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncasteyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s'" %(self.reli,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and taluk='%s'" %(self.reli,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and district='%s'" %(self.reli,self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and taluk='%s' and branch='%s'" %(self.reli,self.tal,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontalukbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and branch='%s'" %(self.reli,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and branch='%s'" %(self.reli,self.yopo,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearfpassouttbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)    
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and taluk='%s' and district='%s'" %(self.reli,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.reli,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontalukyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and Year_Of_Passout='%s'" %(self.reli,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and branch='%s'" %(self.cast,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and taluk='%s'" %(self.cast,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and district='%s'" %(self.cast,self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.cast,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s' and branch='%s'" %(self.cast,self.tal,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetalukbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and district='%s' and branch='%s'" %(self.cast,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castedistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                     
                        
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and Year_Of_Passout='%s' and branch='%s'" %(self.cast,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "casteyopobranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      
                    
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s' and district='%s'" %(self.cast,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.cast,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetalukyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and Year_Of_Passout='%s' and district='%s'" %(self.cast,self.yopo,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "casteyearofpassoutdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where taluk='%s' and sex='%s' and branch='%s'" %(self.tal,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "taluksexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district='%s' and sex='%s' and branch='%s'" %(self.dis,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtsexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and sex='%s' and branch='%s'" %(self.yopo,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutsexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)    
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district='%s' and sex='%s' and taluk='%s'" %(self.dist,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtsextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and sex='%s' and taluk='%s'" %(self.yopo,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "YearOfPassoutsextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      
                        # seex and dis and yopo
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.seex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrictYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                        
                        
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district='%s' and taluk='%s' and branch='%s'" %(self.dis,self.tal,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districttalukbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and taluk='%s' and branch='%s'" %(self.yopo,self.tal,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "YearOfPassouttalukbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # bos and dis and yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and district='%s' and branch='%s'" %(self.yopo,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "YearOfPassoutdistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                    
                          
                          # tal and dis and yopo
            
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and district='%s' and taluk='%s'" %(self.yopo,self.dis,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "YearOfPassoutdistricttaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)    

                    # 4 combo


                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s' and course='%s'" %(self.reli,self.cast,self.seex,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesexcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and course='%s'" %(self.reli,self.cast,self.bos,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s' and course='%s'" %(self.reli,self.cast,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s' and course='%s'" %(self.reli,self.cast,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and Year_Of_Passout='%s' and course='%s'" %(self.reli,self.cast,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncasteyearofpassoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and course='%s'" %(self.reli,self.seex,self.bos,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexbranchcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and taluk='%s' and course='%s'" %(self.reli,self.seex,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsextalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and district='%s' and course='%s'" %(self.reli,self.seex,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and Year_Of_Passout='%s' and course='%s'" %(self.reli,self.seex,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexyearofpassoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and branch='%s' and taluk='%s' and course='%s'" %(self.reli,self.bos,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionbranchtalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and branch='%s' and district='%s' and course='%s'" %(self.reli,self.bos,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionbranchdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and branch='%s' and Year_Of_Passout='%s' and course='%s'" %(self.reli,self.bos,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionbranchyearofpassoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and taluk='%s' and district='%s' and course='%s'" %(self.reli,self.tal,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontalukdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and taluk='%s' and Year_Of_Passout='%s' and course='%s'" %(self.reli,self.tal,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontalukYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and Year_Of_Passout='%s' and course='%s'" %(self.reli,self.dis,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and branch='%s' and course='%s'" %(self.cast,self.seex,self.bos,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexbranchcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and taluk='%s' and course='%s'" %(self.cast,self.seex,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesextalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and district='%s' and course='%s'" %(self.cast,self.seex,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and Year_Of_Passout='%s' and course='%s'" %(self.cast,self.seex,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and branch='%s' and taluk='%s' and course='%s'" %(self.cast,self.bos,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castebranchtalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and branch='%s' and district='%s' and course='%s'" %(self.cast,self.bos,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castebranchdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and branch='%s' and Year_Of_Passout='%s' and course='%s'" %(self.cast,self.bos,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castebranchYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s' and district='%s' and course='%s'" %(self.cast,self.tal,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetalukdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s' and Year_Of_Passout='%s' and course='%s'" %(self.cast,self.tal,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetalukYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and district='%s' and Year_Of_Passout='%s' and course='%s'" %(self.cast,self.dis,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castedistrictYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and branch='%s' and taluk='%s' and course='%s'" %(self.seex,self.bos,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexbranchtalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)         
                          
                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and branch='%s' and district='%s' and course='%s'" %(self.seex,self.bos,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexbranchdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and branch='%s' and Year_Of_Passout='%s' and course='%s'" %(self.seex,self.bos,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexbranchYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and taluk='%s' and district='%s' and course='%s'" %(self.seex,self.tal,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sextalukdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and taluk='%s' and Year_Of_Passout='%s' and course='%s'" %(self.seex,self.tal,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sextalukYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s' and Year_Of_Passout='%s' and course='%s'" %(self.seex,self.dis,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrictYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where branch='%s' and district='%s' and taluk='%s' and course='%s'" %(self.bos,self.dis,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchdistricttalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where branch='%s' and Year_Of_Passout='%s' and taluk='%s' and course='%s'" %(self.bos,self.yopo,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchYear_Of_Passouttalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where branch='%s' and Year_Of_Passout='%s' and district='%s' and course='%s'" %(self.bos,self.yopo,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchYear_Of_Passoutdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where taluk='%s' and Year_Of_Passout='%s' and district='%s' and course='%s'" %(self.tal,self.yopo,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "talukYear_Of_Passoutdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)



# 4 combo 35 combinations completed




                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s' and branch='%s'" %(self.reli,self.cast,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s' and taluk='%s'" %(self.reli,self.cast,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s' and district='%s'" %(self.reli,self.cast,self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s'" %(self.reli,self.cast,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and district='%s'" %(self.reli,self.cast,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s' and taluk='%s'" %(self.reli,self.cast,self.dis,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistricttaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      #reli,cast,tal,yopo    
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s'and Year_Of_Passout='%s'" %(self.reli,self.cast,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetalukYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                    #reli,cast,dis,yopo
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistrictYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                   
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and taluk='%s'" %(self.reli,self.seex,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and district='%s'" %(self.reli,self.seex,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexbranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.seex,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexbranchYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and district='%s' and taluk='%s'" %(self.reli,self.seex,self.dis,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexdistricttaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      #reli,seex,tal,yopo
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.reli,self.seex,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsextalukYearOfPassout"
                          wb.save(workbook_name + ".xlsx")                     
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      #reli,seex,dis,yopo
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and district='%s' and  Year_Of_Passout='%s'" %(self.reli,self.seex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexdistrictYearOfPassout"
                          wb.save(workbook_name + ".xlsx")                          
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and branch='%s' and taluk='%s'" %(self.reli,self.dis,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and branch='%s' and taluk='%s'" %(self.reli,self.yopo,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyopobranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.dis,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictbranchyopo"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      #reli,tal,dis,yopo
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.reli,self.dis,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistricttalukYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                    # 4 combo continutation
                    
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and branch='%s' and taluk='%s'" %(self.seex,self.cast,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcastebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and branch='%s' and district='%s'" %(self.seex,self.cast,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcastebranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.seex,self.cast,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcastebranchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and district='%s' and taluk='%s'" %(self.seex,self.cast,self.dis,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcaste districtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      # cast,seex,tal,yopo
                      
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo !=""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and Year_Of_Passout='%s' and taluk='%s'" %(self.seex,self.cast,self.yopo,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcasteYearOfPassouttaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                      
                      
                      # cast,seex,dis,yopo

                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis !="" and self.yopo !=""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and Year_Of_Passout='%s' and district='%s'" %(self.seex,self.cast,self.yopo,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcasteYearOfPassoutdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)   


                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district='%s' and caste='%s' and branch='%s' and taluk='%s'" %(self.dis,self.cast,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtcastebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and caste='%s' and branch='%s' and taluk='%s'" %(self.yopo,self.cast,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtcastebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          

                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal=="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where district='%s' and caste='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.dis,self.cast,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtcastebranchYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # cast,tal,dis,yopo
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal!="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where district='%s' and caste='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.dis,self.cast,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtcastetalukYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s' and branch='%s' and taluk='%s'" %(self.seex,self.dis,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrictbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and Year_Of_Passout='%s' and branch='%s' and taluk='%s'" %(self.seex,self.yopo,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexYear_Of_Passoutbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.seex,self.dis,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrictbranchYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                     
                        # seex,tal,dis,yopo
                        
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.seex,self.dis,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistricttalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                        
                        
                        # bos,tal,dis,yopo
                        
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where branch='%s' and district='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.bos,self.dis,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchdistricttalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)   
                          
                          
                         # 5 combo
                         
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranch"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                         
                         
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and taluk='%s'" %(self.course,self.reli,self.cast,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesextaluk"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)   
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and district='%s'" %(self.course,self.reli,self.cast,self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
 
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                        
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and taluk='%s'" %(self.course,self.reli,self.cast,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and district='%s'" %(self.course,self.reli,self.cast,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.cast,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastetalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastetalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastedistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and taluk='%s'" %(self.course,self.reli,self.seex,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and district='%s'" %(self.course,self.reli,self.seex,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.seex,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsextalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsextalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)  
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionbranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
 
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionbranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)  
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionbranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligiontalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s'" %(self.course,self.cast,self.seex,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                         
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and district='%s'" %(self.course,self.cast,self.seex,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and taluk='%s' and district='%s'" %(self.course,self.cast,self.seex,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesextalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesextalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.cast,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastebranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastebranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastebranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastetalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.seex,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.seex,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.seex,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.seex,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesextalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursebranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                         
                         
# 5 combo 35 combinations completed                         
                         
                         
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s'" %(self.reli,self.cast,self.bos,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtaluksex"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and district='%s' and sex='%s'" %(self.reli,self.cast,self.bos,self.dis,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchdistrictsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # reli,cast,seex,bos,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and Year_Of_Passout='%s' and sex='%s'" %(self.reli,self.cast,self.bos,self.yopo,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchyearofpassoutsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s' and taluk='%s' and sex='%s'" %(self.reli,self.cast,self.dis,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistricttaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and Year_Of_Passout='%s' and taluk='%s' and sex='%s'" %(self.reli,self.cast,self.yopo,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncasteYearOfPassouttaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s' and Year_Of_Passout='%s' and sex='%s'" %(self.reli,self.cast,self.dis,self.yopo,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistrictYearOfPassoutsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.reli,self.cast,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # reli,cast,bos,tal,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtalukyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # reli,cast,bos,dis,yopo

                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchdistrictyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                          # reli,cast,tal,dis,yopo

                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetalukdistrictyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and branch='%s' and taluk='%s' and sex='%s'" %(self.reli,self.dis,self.bos,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictbranchtaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # reli,seex,bos,tal,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and branch='%s' and taluk='%s' and sex='%s'" %(self.reli,self.yopo,self.bos,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearofpassoutbranchtaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # reli,seex,bos,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and branch='%s' and district='%s' and sex='%s'" %(self.reli,self.yopo,self.bos,self.dis,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearofpassoutbranchdistrictsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                           
                          
                          # reli,seex,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and taluk='%s' and district='%s' and sex='%s'" %(self.reli,self.yopo,self.tal,self.dis,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearofpassouttalukdistrictsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                            
                          
                          # reli,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and taluk='%s' and district='%s' and branch='%s'" %(self.reli,self.yopo,self.tal,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearofpassouttalukdistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                           
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s'" %(self.dis,self.cast,self.bos,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtcastebranchtaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)    
                          
                          
                          # cast,seex,bos,tal,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s'" %(self.yopo,self.cast,self.bos,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutcastebranchtaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # cast,seex,bos,dis,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and caste='%s' and branch='%s' and district='%s' and sex='%s'" %(self.yopo,self.cast,self.bos,self.dis,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutcastebranchdistrictsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # cast,seex,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and caste='%s' and taluk='%s' and district='%s' and sex='%s'" %(self.yopo,self.cast,self.tal,self.dis,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutcastetalukdistrictsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                           
                          
                          # cast,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and caste='%s' and taluk='%s' and district='%s' and branch='%s'" %(self.yopo,self.cast,self.tal,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutcastetalukdistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # seex,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and sex='%s' and taluk='%s' and district='%s' and branch='%s'" %(self.yopo,self.sex,self.tal,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutsextalukdistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)            
                          
                          
                          
                          # 6 combo
                          
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and district='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
 
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.cast,self.seex,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesextalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesextalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.cast,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastetalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.seex,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsextalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionfbranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.cast,self.seex,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesextalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastebranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.seex,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          
                          
# 6 combo 21 combinations completed                          
                          
                                                    
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s' and district='%s'" %(self.reli,self.cast,self.bos,self.tal,self.sex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtaluksexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                          # reli,cast,seex,bos,tal,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.tal,self.sex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtaluksexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                         
                         
                          # reli,cast,seex,bos,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and district='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.dis,self.sex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchdistrictsexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                            
                          
                          # reli,cast,seex,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s' and district='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.tal,self.dis,self.sex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetalukdistrictsexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                           
                          
                          # reli,cast,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s' and district='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.tal,self.dis,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetalukdistrictbranchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # reli,seex,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and taluk='%s' and district='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.sex,self.tal,self.dis,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsextalukdistrictbranchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # cast,seex,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and taluk='%s' and district='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.cast,self.sex,self.tal,self.dis,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesextalukdistrictbranchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                           
                          
                          
                          
                          # 7 combo



                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                    
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesextalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          
# 7 combo 7 combinations completed                          
                          
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.tal,self.sex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtaluksexdistrictyopo"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)


                    # 8 combo


                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.bos,self.tal,self.sex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchtaluksexdistrictyopo"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                  
# 8 combo 1 combinations completed  
                  
                      


if __name__ == "__main__":
    firstgui=MyFirstGui()
    firstgui.mainloop()

