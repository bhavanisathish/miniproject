# -*- coding: utf-8 -*-
"""
Created on Fri Jun  5 10:37:24 2020

@author: dell
"""

from tkcalendar import DateEntry
import re 
from itertools import cycle
import datetime
import webbrowser
from openpyxl import Workbook
from tkinter import messagebox
import os
import glob
from fpdf import FPDF
from tkinter import *
import tkinter as tk
import tkinter.ttk as ttk 
from ttkthemes import ThemedStyle
#import cv2
import mysql.connector
db_cur=mysql.connector.connect(host='localhost',user='root',passwd='',database='studentform')
db=db_cur.cursor()
class MyFirstGui(tk.Tk):
       def __init__(self):
                tk.Tk.__init__(self)
                self._frame = None
                
                self.switch_frame(StartPage)
                self.state("zoomed")
       
       def switch_frame(self, frame_class):
                new_frame = frame_class(self)
                if self._frame is not None:
                    self._frame.destroy()
                style = ThemedStyle(self)
                style.set_theme("plastik")
                
               
                style.configure('W.TButton', font =
                   ('calibri', 20, 'bold')) 
                style.configure('W1search.TButton', font =
                   ('calibri', 20, 'bold')) 
                style.configure('Wbb.TButton', font =
                   ('calibri', 10, 'bold'))  
                style.configure('B.TButton', font =
                   ('calibri', 5, 'bold')) 
                style.configure('L.TButton', font =
                   ('calibri', 20, 'bold')) 
                style.configure('Wild.RButton', font =
                   ('calibri', 20, 'bold'),background="burlywood1") 
                
                style.configure('Wild.TRadiobutton', font=('calibri',15,'bold'),   # First argument is the name of style. Needs to end with: .TRadiobutton
                 background='skyblue3',foreground='black') 
                 
                self._frame = new_frame
                self._frame.pack()
                self.state("zoomed")
                self.title("Destop application")
                #style = ttk.Style()
                #style.theme_use('')
                style.configure("black.Horizontal.TProgressbar", background='black')

                self.configure(background='midnightblue')
               
        
         
class StartPage(tk.Frame): 
           
      def __init__(self, master):
                    tk.Frame.__init__(self,master)
                    self.configure(background='midnightblue')
                    self.Frame1 = tk.Frame(self.master)
                    self.Frame1.pack(side="top",  pady=10,padx=10,expand=True )
                    self.Frame1.configure(background='midnightblue')
                    photo=tk.PhotoImage(file="ss.png")  
                    label1221=tk.Label(self.Frame1,text="Alagappa Chettiar Government College of Engineering and Technology,Karaikudi",background= "midnightblue",foreground="white",font =
                   ('calibri', 30, 'bold'))
                    label1221.grid(row=0,rowspan=1,columnspan=40)
                    label12221=tk.Label(self.Frame1,text="(An autonomous government institution permanently affilitated to Anna University)",background= "midnightblue",foreground="white",font =
                   ('calibri', 20, 'bold'))
                    label12221.grid(row=1,rowspan=1,columnspan=40)
                    
                    #photo = photo.zoom(1)
                    label =tk.Label(self,image = photo)#,width=1680,height=1080)
                    label.image = photo # keep a reference!
                    label.grid(row=2,column=0,columnspan=5,rowspan=20)
                    self.Frame2 = tk.Frame(self.master)
                    
                    self.Frame2.pack(side="right",  pady=80,padx=10, expand=True )
                    self.Frame2.configure(background='midnightblue')
                    wer=tk.Label(self.Frame2,text='VISION',background= "midnightblue",foreground="white",font = ('calibri', 22, 'bold'))
                    wer.grid(row=2,column=0,sticky='W')
                    le2=tk.Label(self.Frame2,text='Our  Commitment  as  a  Centre   of  ',background= "midnightblue",foreground="white", font = ('calibri', 18, 'bold'))
                    le2.grid(row=3,column=0,sticky='W')#olumnspan=40,rowspan=4)
                    let2=tk.Label(self.Frame2,text='Engineering  Educations   to  impart ',background= "midnightblue",foreground="white",font = ('calibri', 18, 'bold'))
                    let2.grid(row=4,column=0,sticky='W')
                    le12=tk.Label(self.Frame2,text='Technical   Knowledge  par excellence, ',background= "midnightblue",foreground="white",font = ('calibri', 18, 'bold'))
                    le12.grid(row=5,column=0,sticky='W')#olumnspan=40,rowspan=4)
                    lere=tk.Label(self.Frame2,text='motivate the learners   in  Research,    ',background= "midnightblue",foreground="white",font = ('calibri', 18, 'bold'))
                    lere.grid(row=6,column=0,sticky='W')
                    lee=tk.Label(self.Frame2,text='evolve  result  –  oriented,  innovative   ',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
                    lee.grid(row=7,column=0,sticky='W')
                    lee1r=tk.Label(self.Frame2,text='techniques in   Engineering, provide ',background= "midnightblue",foreground="white",font = ('calibri', 18, 'bold'))
                    lee1r.grid(row=8,column=0,sticky='W')
                    leew1=tk.Label(self.Frame2,text='necessary career guidance,and  train  ',background= "midnightblue",foreground="white",font = ('calibri', 18, 'bold'))
                    leew1.grid(row=9,column=0,sticky='W')
                    lee2=tk.Label(self.Frame2,text='our learners in leadership',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
                    lee2.grid(row=10,column=0,sticky='W')
                    lee2=tk.Label(self.Frame2,text='qualities so as to achieve better',background= "midnightblue",foreground="white",font = ('calibri', 18, 'bold'))
                    lee2.grid(row=11,column=0,sticky='W')
                    lee3=tk.Label(self.Frame2,text='productivity and prosperity',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
                    lee3.grid(row=12,column=0,sticky='W')
                    lee3=tk.Label(self.Frame2,text='for our country.',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
                    lee3.grid(row=13,column=0,sticky='W')
                    #self.state("zoomed")
                    self.l1=ttk.Label(self,text="Login here",font=("Times 40 italic"),background="seashell3")
                    self.l1.grid(row=3,column=1,columnspan=2,sticky='W',padx=10,pady=50)
                    
                    self.l2=ttk.Label(self,text="Username",font=("Times 30 italic"),background="seashell3")
                    self.l2.grid(row=5,column=0,padx=10,pady=50)
                    self.password=ttk.Label(self,text="Password",font=("Times 30 italic"),background="seashell3")
                    self.password.grid(row=7,column=0,padx=10,pady=50)
                    self.e1=ttk.Entry(self,font=("Times 18 italic"))
                    self.e1.grid(row=5,column=2,padx=5,pady=5,ipady=5)
                    self.e2=ttk.Entry(self,font=("Times 18 italic"))
                    self.e2.grid(row=7,column=2,padx=5,pady=5,ipady=5)
                    self.b1=ttk.Button(self,text="Submit",style = 'W.TButton',command=self.passd)
                    self.b1.grid(row=9,column=2,padx=30,pady=50)
                    self.b12=ttk.Button(self,text="Register",style = 'W.TButton',command=self.reg)
                    self.b12.grid(row=9,column=1,padx=30,pady=50)
                    self.b2=ttk.Button(self,text="Reset",style = 'W.TButton',command=self.reset)
                    self.b2.grid(row=9,column=0,padx=30,pady=50)
                    
                    #self.button=tk.Button(self,text="->",width=10,command=lambda: master.switch_frame(pageOne))
                    #self.button.grid(row=4,column=0,sticky='W')
                    
            #self.ll1=tk.Button(self, text = 'Click Me !', image = photo1)
      def reset(self):
          self.e1.delete(first=0,last=300)
          self.e2.delete(first=0,last=300)
      def reg(self):
          self.Frame2.destroy()
          self.Frame1.destroy()
          self.master.switch_frame(registration)
      def passd(self):
            global username,passwords
            username = self.e1.get()
            self.username = self.e1.get()
            self.passwords = self.e2.get()
            db.execute("select *from adminlogin where name='%s' and password='%s'"%(self.username,self.passwords))
            l=db.fetchone()
            if(l):
                self.Frame2.destroy()
                self.Frame1.destroy()
                print("check")
                #self.switch_frame(pageOne)
                self.master.switch_frame(pageOne)
                db_cur.commit()
            else:
                   messagebox.showinfo("Login Error", "Invalid Username Or Passord")

                
     # def make_label(parent, img):
      #    label11 = ttk.Label(parent, image=img)
       #   label11.pack()


class registration(tk.Frame):
    def __init__(self,master):
        tk.Frame.__init__(self,master)
        self.Frame10 = tk.Frame(self.master)
        self.Frame10.pack(side="top",  pady=10,padx=10,expand=True )
        self.Frame10.configure(background='midnightblue')
        photo=tk.PhotoImage(file="ss.png")  
        label1221=tk.Label(self.Frame10,text="Alagappa Chettiar Government College of Engineering and Technology,Karaikudi",background= "midnightblue",foreground="white",font =
                   ('calibri', 30, 'bold'))
        label1221.grid(row=0,rowspan=1,columnspan=40)
        label12221=tk.Label(self.Frame10,text="(An autonomous government institution permanently affilitated to Anna University)",background= "midnightblue",foreground="white",font =
                   ('calibri', 20, 'bold'))
        label12221.grid(row=1,rowspan=1,columnspan=40)
        
        #self.configure(background='powderblue')
        self.configure(background='midnightblue')
        photo=tk.PhotoImage(file="ss.png") 
                    
                    #photo = photo.zoom(1)
        label =tk.Label(self,image = photo)#,width=1680,height=1080)
        label.image = photo # keep a reference!
        label.grid(row=0,column=0,columnspan=5,rowspan=20)
        self.Frame3 = tk.Frame(self.master)
                    
        self.Frame3.pack(side="right",  pady=50,padx=10, expand=True )
        self.Frame3.configure(background='midnightblue')
        la=tk.Label(self.Frame3,text='MISSION',background= "midnightblue",foreground="white",font = ('calibri',30, 'bold'))
        la.grid(row=0,column=0,sticky='W')
        la1=tk.Label(self.Frame3,text='Constantly updating the departmental resources,',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
        la1.grid(row=1,column=0,sticky='W')
        la2=tk.Label(self.Frame3,text=' facility and other infrastructure ',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
        la2.grid(row=2,column=0,sticky='W')
        la3=tk.Label(self.Frame3,text=' by acquiring state of art equipment.',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
        la3.grid(row=3,column=0,sticky='W')
        la4=tk.Label(self.Frame3,text='',background= "midnightblue",foreground="white")
        la4.grid(row=4,column=0,sticky='W')
        la5=tk.Label(self.Frame3,text='Imparting constant in–service training',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
        la5.grid(row=5,column=0,sticky='W')
        la6=tk.Label(self.Frame3,text='to the faculty and supporting staff.',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
        la6.grid(row=6,column=0,sticky='W')
        la6=tk.Label(self.Frame3,text='',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
        la6.grid(row=7,column=0,sticky='W')
        la6=tk.Label(self.Frame3,text='Inculcating the feeling of oneness responsibility ',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
        la6.grid(row=8,column=0,sticky='W')
        la6=tk.Label(self.Frame3,text='and service to community in the minds',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
        la6.grid(row=9,column=0,sticky='W')
        la6=tk.Label(self.Frame3,text='of students to serve the society better.',background= "midnightblue",foreground="white",font = ('calibri',18, 'bold'))
        la6.grid(row=10,column=0,sticky='W')
        self.lb1=ttk.Label(self,text="Register here",font=("Times 40 bold"),background="seashell3")
        self.lb1.grid(row=0,column=0,padx=10,pady=30)
        self.lb2=ttk.Label(self,text="Name",font=("Times 18 italic"),background="seashell3")
        self.lb2.grid(row=1,column=0,padx=10,pady=15)
        self.ent1=ttk.Entry(self,font=("Times 18 italic"))
        self.ent1.grid(row=1,column=1,padx=5,pady=5,ipady=5)
        self.lb3=ttk.Label(self,text="Password",font=("Times 18 italic"),background="seashell3")
        self.lb3.grid(row=2,column=0,padx=10,pady=15)
        self.ent2=ttk.Entry(self,font=("Times 18 italic"))
        self.ent2.grid(row=2,column=1,padx=5,pady=5,ipady=5)
        self.lb4=ttk.Label(self,text="Department",font=("Times 18 italic"),background="seashell3")
        self.lb4.grid(row=3,column=0,padx=10,pady=15)
        self.ent3=ttk.Entry(self,font=("Times 18 italic"))
        self.ent3.grid(row=3,column=1,padx=5,pady=5,ipady=5)
        self.lb5=ttk.Label(self,text="Posting",font=("Times 18 italic"),background="seashell3")
        self.lb5.grid(row=4,column=0,padx=10,pady=15)
        self.ent4=ttk.Entry(self,font=("Times 18 italic"))
        self.ent4.grid(row=4,column=1,padx=5,pady=5,ipady=5)
        self.lb6=ttk.Label(self,text="Admin Name",font=("Times 18 italic"),background="seashell3")
        self.lb6.grid(row=5,column=0,padx=10,pady=15)
        self.ent5=ttk.Entry(self,font=("Times 18 italic"))
        self.ent5.grid(row=5,column=1,padx=5,pady=5,ipady=5)
        self.lb7=ttk.Label(self,text="Admin Password",font=("Times 18 italic"),background="seashell3")
        self.lb7.grid(row=6,column=0,padx=10,pady=15)
        self.ent6=ttk.Entry(self,font=("Times 18 italic"))
        self.ent6.grid(row=6,column=1,padx=5,pady=5,ipady=5)
        self.bt1=ttk.Button(self,text="Register",command=self.regg,style = 'W.TButton')
        self.bt1.grid(row=7,column=2,sticky='S',padx=10,pady=15)
        self.btt2=ttk.Button(self,text="Back",style = 'W.TButton',command=self.back)
        self.btt2.grid(row=7,column=0,padx=10,pady=15)
        self.bt2=ttk.Button(self,text="Reset",style = 'W.TButton',command=self.rese)
        self.bt2.grid(row=7,column=1,padx=10,pady=15)
    def rese(self):  
        self.ent1.delete(first=0,last=300)
        self.ent2.delete(first=0,last=300)
        self.ent3.delete(first=0,last=300)
        self.ent4.delete(first=0,last=300)
        self.ent5.delete(first=0,last=300)
        self.ent6.delete(first=0,last=300)
    def back(self):
        self.Frame3.destroy()
        self.Frame10.destroy()
        self.master.switch_frame(StartPage)
        
    def regg(self):
        self.namee=self.ent1.get()
        self.passsword=self.ent2.get()
        self.depart=self.ent3.get()
        self.post=self.ent4.get()
        self.adminu=self.ent5.get()
        self.adminpass=self.ent6.get()
        db.execute("select *from adminlogin where name='%s' and password='%s'"%(self.adminu,self.adminpass))
        l=db.fetchone()
        if(l):
            sql1="INSERT INTO `adminlogin`(`name`, `password`, `department`, `posting`) VALUES(%s,%s,%s,%s)"
            val1=(self.namee,self.passsword,self.depart,self.post)
            db.execute(sql1, val1)
            db_cur.commit()
            self.Frame3.destroy()
            self.Frame10.destroy()
            self.master.switch_frame(StartPage)
        else:
            messagebox.showinfo("Registration Error", "You Are Not a Valid Admin.. Please Get Verification From The Admin")
                 


       
class pageOne(tk.Frame):
      def __init__(self,master):
             tk.Frame.__init__(self,master)
             #self.Frame1 = tk.Frame(self.master)
             #self.Frame1.pack(side="top",  pady=10,padx=10,expand=True )
             #self.Frame1.configure(background='cadetBlue1')
             #photo=tk.PhotoImage(file="ss.png") 
             adminname='logged in as '+username
             label1221=tk.Label(self,text="Alagappa Chettiar Government College of Engineering and Technology,Karaikudi",background= "midnightblue",foreground="white",font =
                   ('calibri', 30, 'bold'))
             label1221.grid(row=0)
             label12221=tk.Label(self,text="(An autonomous government institution permanently affilitated to Anna University)",background= "midnightblue",foreground="white",font =
                   ('calibri', 20, 'bold'))
             label12221.grid(row=1)
             self.configure(background='midnightblue')
             self.Frame4 = tk.Frame(self.master,bg="yellow",bd=3,relief=tk.RAISED)
                    
             self.Frame4.pack()
            
             self.menubar = tk.Menu(self.Frame4,tearoff=0,bd=2)
             self.filemenu = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black')
             self.filemenu.add_command(label="Individual Transfer Certificate",command=self.transsing,font=("Times 18"))
             self.filemenu.add_command(label="Batch Transfer Certificate",command=self.transgrpo,font=("Times 18"))
             self.filemenu.add_command(label="Update TC",command=self.extratc,font=("Times 18"))
             self.menubar.add_cascade(label="Transfer Certificate Generation", menu=self.filemenu,font = "Times 38")
             self.edit = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
             self.edit.add_command(label="Edit",command=self.edt ,font=("Times 18"))
             
             self.menubar.add_cascade(label='Edit Details',menu=self.edit,font = "Times 38")
             self.insert = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
             self.insert.add_command(label="Insert",command=self.ins,font=("Times 18"))
         
             self.menubar.add_cascade(label='Insert Details',menu=self.insert,font = "Times 38")
             self.search = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
             self.search.add_command(label="Search",command=self.ser,font=("Times 18"))
         
             self.menubar.add_cascade(label='Search Details',menu=self.search,font = "Times 38")
             self.menubar.add_command(label="Logout",command=self.logg,font=("Times 18"))
             self.menubar.add_command(label=adminname,font=("Times 18"))
         
             #self.menubar.config("Verdana", 14)
             self.master.config(menu=self.menubar)
             #print(self.menubar.winfo_height())
             #self.top =Toplevel(menu=self.menubar, width=500, relief=tk.RAISED,borderwidth=2)
             self.delay = 3000
             image_files = [
            'ss.png',
            'ss.png',
            'ss.png',
            'ss.png',
            'ss.png'
            ]
            # allows repeat cycling through the pictures
            # store as (img_object, img_name) tuple
             self.pictures = cycle((tk.PhotoImage(file=image), image)
                                  for image in image_files)
             self.picture_display = tk.Label(self,width=600,height=400)
             self.picture_display.grid(row=3,column=0,sticky='W')
             #self.p = cycle((tk.PhotoImage(file=image), image)for image in image_files)
             self.p_d=tk.Label(self,width=600,height=400)             
             self.p_d.grid(row=3,column=0,sticky='E')
             self.label=tk.Label(self,text='',background= "midnightblue",foreground="white")
             self.label.grid(row=2,column=1)
             self.lee=tk.Label(self,background= "midnightblue")
             self.lee.grid(row=4)
             self.lee2=tk.Label(self,text='HISTORY OF ACGCET',background= "midnightblue",foreground="white",font=("Times 15 italic"))
             self.lee2.grid(row=5,sticky='W')
             self.lw=tk.Label(self,font=("Times 11 "),text='Dr.RM.Alagappa Chettiar, a man of rare wisdom, and forethought, founded  Alagappa Chettiar Educational Trust  with the sole aim of developing the backward area of Karaikudi into a centre for higher education and provided ',background= "midnightblue",foreground="white")
             self.lw.grid(row=6,column=0,sticky='W')
             self.lw1=tk.Label(self,text='necessary funds for the establishment of education institutions. On the Occasion of laying of the foundation stone of the central Electro Chemical Research Institue',background= "midnightblue",foreground="white",font=("Times 11 "))
             self.lw1.grid(row=7,column=0,sticky='W')
             self.lw2=tk.Label(self,text='by pandit jawaharlal Nehru on 25th july 1948. Dr.Alagappa Chettiar, in his Welcome address, said “It is my hope to start here an Engineering College immediately. A College with Dr.Alagappa Chettiar and the University of Madras ',background= "midnightblue",foreground="white",font=("Times 11 "))
             self.lw2.grid(row=8,column=0,sticky='W')
             self.lw3=tk.Label(self,text='willing start functioning with Civil Engineering by academic year 1949.” In 1952, Dr. Alagappa Chettiar’s dream came true. Alagappa Chettiar College of Engg & Tech., Started functioning from 21st July 1952 with three ',background= "midnightblue",foreground="white",font=("Times 11 "))
             self.lw3.grid(row=9,column=0,sticky='W')
             self.lw4=tk.Label(self,text='faculties-Civil,Mechanical, Electrical & Electronics Engineering.The Foundation tablet for the main building of the college was laid by Dr.Rajendra Prasad, the then President of India on 19 th February 1953.',background= "midnightblue",foreground="white",font=("Times 11 "))
             self.lw4.grid(row=10,column=0,sticky='W')
             img_object, img_name = next(self.pictures)
             self.picture_display.config(image=img_object)
             self.p_d.config(image=img_object)
             # shows the image filename, but could be expanded
             # to show an associated description of the image
             #self.title(img_name)
             self.after(self.delay, self.show_slides)
      
      def transsing(self):
          self.Frame4.destroy()
          self.master.switch_frame(transfersingle)
      def transgrpo(self):
          self.Frame4.destroy()
          self.master.switch_frame(transfergroup)
      def ins(self):
          self.Frame4.destroy()
          self.master.switch_frame(newreg)
      def edt(self):
          self.Frame4.destroy() 
          self.master.switch_frame(editspecify)
      def ser(self):
          self.Frame4.destroy()
          self.master.switch_frame(search)
      def extratc(self):
              self.Frame4.destroy()
              self.master.switch_frame(updatetc)
      def logg(self):
          self.Frame4.destroy()
          self.master.switch_frame(StartPage)
      def show_slides(self):
        '''cycle through the images and show them'''
        # next works with Python26 or higher
        img_object, img_name = next(self.pictures)
        self.picture_display.config(image=img_object)
        self.p_d.config(image=img_object)
        # shows the image filename, but could be expanded
        # to show an associated description of the image
        #self.title(img_name)
        self.after(self.delay, self.show_slides)



             #self.configure(menu=menubar)

             #self.configure(background='powderblue')
             
             

        
class updatetc(tk.Frame):

            def __init__(self, master):
                tk.Frame.__init__(self,master)
                self.configure(background='burlywood1')
                
                self.Frame14 = tk.Frame(self.master,bg="yellow",bd=3,relief=tk.RAISED)
                    
                self.Frame14.pack()
                adminname='logged in as '+username
                self.menubar = tk.Menu(self.Frame14,tearoff=0,bd=2)
                self.filemenu = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black')
                self.filemenu.add_command(label="Individual Transfer Certificate",command=self.transsing,font=("Times 18 italic"))
                self.filemenu.add_command(label="Batch Transfer Certificate",command=self.transgrpo,font=("Times 18 italic"))
                self.filemenu.add_command(label="Update TC",command=self.extratc,font=("Times 18 italic"))
                self.menubar.add_cascade(label="Transfer Certificate Generation", menu=self.filemenu,font = "Times 38")
                self.edit = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                self.edit.add_command(label="Edit",command=self.edt ,font=("Times 18 italic"))
                 
                self.menubar.add_cascade(label='Edit Details',menu=self.edit,font = "Times 38")
                self.insert = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                self.insert.add_command(label="Insert",command=self.ins,font=("Times 18 italic"))
                    
                self.menubar.add_cascade(label='Insert Details',menu=self.insert,font = "Times 38")
                self.search = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                self.search.add_command(label="Search",command=self.ser,font=("Times 18 italic"))
                
                self.menubar.add_cascade(label='Search Details',menu=self.search,font = "Times 38")
                self.menubar.add_command(label="Logout",command=self.logg,font=("Times 18 italic"))
                self.menubar.add_command(label=adminname,font=("Times 18 italic"))
                     
                         #self.menubar.config("Verdana", 14)
                self.master.config(menu=self.menubar)
                self.Frame11 = tk.Frame(self.master)
                self.Frame11.pack(side="top",  pady=1,padx=1,expand=True )
                self.Frame11.configure(background='midnightblue')
                label1221=tk.Label(self.Frame11,text="Alagappa chettiar Government College of Engineering and Technology,Karaikudi",background= "midnightblue",foreground="white",font =
                                         ('Times', 20, 'bold'))
                label1221.grid(row=0,rowspan=1,column=1,columnspan=40)
                label12221=tk.Label(self.Frame11,text="(An autonomous government institution permanently affilitated to Anna University)",background= "midnightblue",foreground="white",font =
                                          ('Times', 12, 'bold'))
                label12221.grid(row=1,rowspan=1,column=1,columnspan=40)
                self.left_frame = tk.Frame(self, width=200, height=800, bg='skyblue3')
                self.left_frame.grid(row=1, column=0, padx=10, pady=5)
                self.right_frame = tk.Frame(self, width=650, height=400, bg='skyblue3')
                self.right_frame.grid(row=1, column=1, padx=10, pady=5)
 
# Create frames and labels in left_frame
                #tk.Label(self.left_frame, text="Original Image").grid(row=0, column=0, padx=5, pady=5)
                image = tk.PhotoImage(file="tc.png")
                original_image = image.subsample(1) # resize image using subsample
                label =tk.Label(self.left_frame,image = original_image)#,width=1680,height=1080)
                label.image = original_image # keep a reference!
                label.grid(row=1, column=0, padx=5, pady=5)
                
                
                self.l=ttk.Label(self,text="Update Transfer Certificate Details",font=("Times 20 bold"),background="burlywood1",foreground="white")
                self.l.grid(row=0,column=1)
                
                self.l1=ttk.Label(self.right_frame,text="Register Number",font=("Times 15 bold"),background="skyblue3",foreground="black")
                self.l1.grid(row=1,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.ent1=ttk.Entry(self.right_frame,font=("Times 15"))
                self.ent1.grid(row=1,column=1,padx=5,pady=1,ipady=5,sticky='W')
                self.l2=ttk.Label(self.right_frame,text="Community",font=("Times 15 bold"),background="skyblue3")
                
                self.l2.grid(row=2,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.t3 = tk.StringVar(self)
                
                self.e2 =ttk.Combobox(self.right_frame, width=18, textvariable=self.t3,font=("Times 15 italic"))
                self.e2['values']=('OC','BC','BCM','MBC','SC','ST','Others')  
                self.e2.grid(row=2,column=1,padx=5,pady=1,ipady=5,sticky='W')
                
                
                self.vieww=tk.IntVar()
                self.viewe=tk.IntVar()
                self.viiew=tk.IntVar()
        
                
                self.l1=ttk.Label(self.right_frame,text="Whether the student has paid all the fees due to the college?",font=("Times 15 bold"),background='skyblue3')
                self.l1.grid(row=5,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.r10=ttk.Radiobutton(self.right_frame, text="Yes",variable=self.vieww, value=1,style="Wild.TRadiobutton",command=self.feey)
                self.r10.grid(row=5,column=1,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                self.r11=ttk.Radiobutton(self.right_frame, text="No",variable=self.vieww, value=2,style = 'Wild.TRadiobutton',command=self.feex)
                self.r11.grid(row=5,column=2,sticky='W',padx=1,pady=10,ipadx=1,ipady=10)
                
                self.l2=ttk.Label(self.right_frame,text="Whether the student was in the receipt of any scholarship",font=("Times 15 bold"),background='skyblue3')
                self.l2.grid(row=6,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.r20=ttk.Radiobutton(self.right_frame, text="Yes",variable=self.viewe, value=1,style = 'Wild.TRadiobutton',command=self.schy)
                self.r20.grid(row=6,column=1,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                self.r21=ttk.Radiobutton(self.right_frame, text="No",variable=self.viewe, value=2,style = 'Wild.TRadiobutton',command=self.schx)
                self.r21.grid(row=6,column=2,sticky='W',padx=1,pady=10,ipadx=1,ipady=10)
                
                self.l3=ttk.Label(self.right_frame,text="Whether the student has undergone Medical inspection during the year",font=("Times 15 bold"),background='skyblue3')
                self.l3.grid(row=7,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.r30=ttk.Radiobutton(self.right_frame, text="Yes",variable=self.viiew, value=1,style = 'Wild.TRadiobutton',command=self.medy)
                self.r30.grid(row=7,column=1,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                self.r31=ttk.Radiobutton(self.right_frame, text="No",variable=self.viiew, value=2,style = 'Wild.TRadiobutton',command=self.medx)
                self.r31.grid(row=7,column=2,sticky='W',padx=1,pady=10,ipadx=1,ipady=10)
                                                        
                self.l4=ttk.Label(self.right_frame,text="Reason for leaving the college",font=("Times 15 bold"),background="skyblue3")
                self.l4.grid(row=8,column=0,padx=10,pady=30,sticky='W')
                self.e40=ttk.Entry(self.right_frame,font=("Times 15"))
                self.e40.grid(row=8,column=1,padx=5,pady=5,ipady=5,sticky='W')
                
                self.l4=ttk.Label(self.right_frame,text="Date on which application for Transfer Certificate was ",font=("Times 15 bold"),background="skyblue3")
                self.l4.grid(row=9,column=0,padx=10,pady=10,sticky='W')                
                self.l5=ttk.Label(self.right_frame,text="made by the student or on his/her behalf by parent/guardian",font=("Times 15 bold"),background="skyblue3")
                self.l5.grid(row=10,column=0,padx=10,pady=1,sticky='W')
                self.e41=DateEntry(self.right_frame, date_pattern='dd-mm-yyyy', font=("Times 15"), borderwidth=2)
                           
                self.e41.grid(row=10,column=1,padx=5,pady=1,ipady=5,sticky='W')
                
                self.but=ttk.Button(self.right_frame,text="Submit",style = 'W.TButton',command=self.updatz)
                self.but.grid(row=11,column=1,padx=15,pady=5,ipady=5,sticky='W')
                
            def feey(self):
                self.feeey="yes"
            def feex(self):
                self.feeey="no"
            def schy(self):
                self.schhy="yes"
            def schx(self):
                self.schhy="no"
            def medy(self):
                self.meedy="yes"
            def medx(self):
                self.meedy="no"
            def transsing(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(transfersingle)
            def transgrpo(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.right_frame.destroy()
                  self.left_frame.destroy()
                  self.master.switch_frame(transfergroup)
            def extratc(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.right_frame.destroy()
                  self.left_frame.destroy()
                  self.master.switch_frame(updatetc)    
            def ins(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.right_frame.destroy()
                  self.left_frame.destroy()
                  self.master.switch_frame(newreg)
            def edt(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.right_frame.destroy()
                  self.left_frame.destroy()
                  self.master.switch_frame(editspecify)
            def ser(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.right_frame.destroy()
                  self.left_frame.destroy()
                  self.master.switch_frame(search)
            def logg(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.right_frame.destroy()
                  self.left_frame.destroy()
                  self.master.switch_frame(StartPage)      
        
            def updatz(self):
                self.comm=self.t3.get()
                #print()
                self.reg=self.ent1.get()
                print(self.reg)
                
                self.reasn=self.e40 .get()
                self.donw=self.e41.get()
                if(self.comm !=' ' and self.reg == '' and self.feeey !='' and self.schhy!='' and  self.meedy!='' and self.reasn!='' and self.donw!=''):
                    db.execute("select *from studentdetails where caste='%s'  and flag='0'"%(self.comm))
                    my=db.fetchall()
                    #j=1
                    for o in my:
                        reg_no = o[1]
                        if(datetime.datetime.strptime(self.donw, '%d-%m-%Y') and self.feeey!="" and reg_no !="" and self.schhy!="" and self.meedy!="" and self.schhy!="" ):
                             sql0="INSERT INTO `addtcinfo`(`reg_no`, `student_bill`, `scholarship`, `medicalinspection`, `reasonforleaving`, `addofaplication`) VALUES(%s,%s,%s,%s,%s,%s)"
                             val0=(reg_no,self.feeey,self.schhy,self.meedy,self.reasn,self.donw)
                             db.execute(sql0,val0)
                             db_cur.commit() 
                             messagebox.showinfo("TransferCertificate", 'Datas related to Transfer Cetificate were Updated ') 
                             
                             self.Frame14.destroy()
                             self.Frame11.destroy()
                             self.right_frame.destroy()
                             self.left_frame.destroy() 
                             self.master.switch_frame(pageOne) 
                        else:
                            messagebox.showinfo("Transfer Error", "Fill all the fields and Check above details are True") 
                elif(self.comm !=' ' and self.reg != '' and self.feeey !='' and self.schhy!='' and  self.meedy!='' and self.reasn!='' and self.donw!=''):
                    print(self.reg)
                    db.execute("select *from studentdetails where reg_number='%s'  and flag='0'"%(self.reg))
                    my=db.fetchone()
                    regg_no=my[1]
                    if(datetime.datetime.strptime(self.donw, '%Y-%m-%d') and self.feeey!="" and regg_no !="" and self.schhy!="" and self.meedy!="" and self.schhy!="" ):
                             sql0="INSERT INTO `addtcinfo`(`reg_no`, `student_bill`, `scholarship`, `medicalinspection`, `reasonforleaving`, `addofaplication`) VALUES(%s,%s,%s,%s,%s,%s)"
                             val0=(regg_no,self.feeey,self.schhy,self.meedy,self.reasn,self.donw)
                             db.execute(sql0,val0)
                             db_cur.commit()
                             messagebox.showinfo("TransferCertificate", 'Datas related to transferCetificate were Updated ') 
                             
                             self.Frame14.destroy()
                             self.Frame11.destroy()
                             self.right_frame.destroy()
                             self.left_frame.destroy()           
                             self.master.switch_frame(pageOne) 
                    else:
                            messagebox.showinfo("Transfer Error", "Fill all the fields and Check whether above details are True") 
          
              
class transfergroup(tk.Frame):
        def __init__(self, master):
                tk.Frame.__init__(self,master)
                self.configure(background='midnightblue')
                self.view=tk.IntVar()
                self.Frame14 = tk.Frame(self.master,bg="yellow",bd=3,relief=tk.RAISED)
                    
                self.Frame14.pack()
                adminname='logged in as '+username
                self.menubar = tk.Menu(self.Frame14,tearoff=0,bd=2)
                self.filemenu = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black')
                self.filemenu.add_command(label="Individual Transfer Certificate",command=self.transsing,font=("Times 18"))
                self.filemenu.add_command(label="Batch Transfer Certificate",command=self.transgrpo,font=("Times 18"))
                self.filemenu.add_command(label="Update TC",command=self.extratc,font=("Times 18"))
                self.menubar.add_cascade(label="Transfer Certificate Generation", menu=self.filemenu,font = "Times 38")
                self.edit = tk.Menu(self.menubar, tearoff=0,font=("Times 18"),background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black')
                self.edit.add_command(label="Edit",command=self.edt ,font=("Times 18"))
                 
                self.menubar.add_cascade(label='Edit Details',menu=self.edit,font = "Times 38")
                self.insert = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                self.insert.add_command(label="Insert",command=self.ins,font=("Times 18"))
                    
                self.menubar.add_cascade(label='Insert Details',menu=self.insert,font = "Times 38")
                self.search = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                self.search.add_command(label="Search",command=self.ser,font=("Times 18"))
                
                self.menubar.add_cascade(label='Search Details',menu=self.search,font = "Times 38")
                self.menubar.add_command(label="Logout",command=self.logg,font=("Times 18"))
                self.menubar.add_command(label=adminname,font=("Times 18"))
                     
                         #self.menubar.config("Verdana", 14)
                self.master.config(menu=self.menubar)
                self.Frame11 = tk.Frame(self.master)
                self.Frame11.pack(side="top",  pady=1,padx=1,expand=True )
                self.Frame11.configure(background='midnightblue')
                label1221=tk.Label(self.Frame11,text="Alagappa chettiar Government College of Engineering and Technology,Karaikudi",background= "midnightblue",foreground="white",font =
                                         ('Times', 26, 'bold'))
                label1221.grid(row=0,rowspan=1,column=1)
                label12221=tk.Label(self.Frame11,text="(An autonomous government institution permanently affilitated to Anna University)",background= "midnightblue",foreground="white",font =
                                          ('Times', 20, 'bold'))
                label12221.grid(row=1,rowspan=1,column=1)
                self.l=ttk.Label(self.Frame11,text="Batch Transfer Certificate Details",font=("Times 20 bold"),background="midnightblue", foreground="white")
                self.l.grid(row=2,column=1)
                
                self.left_frame = tk.Frame(self, width=200, height=400, bg='skyblue3')
                self.left_frame.grid(row=0, column=0, padx=10, pady=5)
                self.right_frame = tk.Frame(self, width=650, height=400, bg='skyblue3')
                self.right_frame.grid(row=0, column=1, padx=10, pady=5)
                
# Create frames and labels in left_frame
                #tk.Label(self.left_frame, text="Original Image").grid(row=0, column=0, padx=5, pady=5)
                image = tk.PhotoImage(file="tc.png")
                original_image = image.subsample(1) # resize image using subsample
                label =tk.Label(self.left_frame,image = original_image)#,width=1680,height=1080)
                label.image = original_image # keep a reference!
                label.grid(row=1, column=0, padx=5, pady=5)
                
                self.l=ttk.Label(self.right_frame,text="Enter the Details",font=("Times 22 bold"),background="skyblue3")
                self.l.grid(row=0,padx=10,pady=20)
                
                
                self.l12=ttk.Label(self.right_frame,text="Branch Of Study",font=("Times 18 bold"),background="skyblue3")
                self.l12.grid(row=1,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                #self.ll12=ttk.Label(self.right_frame,text="  :  ",font=("Times 15 italic"),background="darkorchid")
                #self.ll12.grid(row=1,column=1,sticky='W',pady=3,padx=40)
                self.t6 = tk.StringVar(self)
                self.p6 =ttk.Combobox(self.right_frame, width=19, textvariable=self.t6,font=("Times 18 bold"))
                self.p6['values']=('Civil Engineering','Mechanical Engineering','Electrical and Electronics Engineering','Electronics and Communication Engineering','Computer Science and Engineering')  
                self.p6.grid(row=1,column=1,padx=5,pady=5,ipady=5,sticky='W')
                self.p6.current()
                
                
                              
                self.le2=ttk.Label(self.right_frame,text="Leaving Date",font=("Times 18 bold"),background="skyblue3")
                self.le2.grid(row=2,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.ee2=DateEntry(self.right_frame,  date_pattern='dd-mm-yyyy', font=("Times 18 bold"), borderwidth=2)
            
                self.ee2.grid(row=2,column=1,padx=5,pady=5,ipady=5,sticky='W')
                self.le3=ttk.Label(self.right_frame,text="Year Of PassOut",font=("Times 18 bold"),background="skyblue3")
                self.le3.grid(row=3,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.ee3=ttk.Entry(self.right_frame,font=("Times 18 italic"))
                self.ee3.grid(row=3,column=1,padx=5,pady=5,ipady=5,sticky='W')
                #print(entr1.get())
                
                #print(entr1.get())
                
                self.but=ttk.Button(self.right_frame,text="SUBMIT",style = 'W.TButton',command=self.tranzgrp)
                self.but.grid(row=4,column=1,padx=5,pady=5,sticky='E')
              
                #self.button16=ttk.Button(self,text="BACK",style = 'W.TButton',command=lambda: master.switch_frame(transferspecify))
                #self.button16.grid(row=11,column=0,sticky='W')
        
        def transsing(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(transfersingle)
        def transgrpo(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(transfergroup)
        def ins(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(newreg)
        def edt(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(editspecify)
        def ser(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(search)
        def logg(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(StartPage)      
        
        
        def extratc(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(updatetc)
            
        def tranzgrp(self):
            global sentt,getit,yopo,ppname
            sentt=self.p6.get()
            yopo = self.ee3.get()
            getit=self.ee2.get()
           
            db.execute("select *from studentdetails where branch='%s' and Year_Of_Passout ='%s' and flag='0' "%(sentt,yopo))
            myresult=db.fetchall()
            pdf=FPDF()
            #reg=
            i=1
            msges = 'Transfer certificate has been generated for the batch'+yopo
            messagebox.showinfo("TransferCertificate", msges) 
            for f in myresult:
                reg_no=f[1]
                db.execute("update studentdetails set flag='1' where reg_number='%s'"%(reg_no))
                                 
                db_cur.commit() 
                print(i)
                i=i+1
                db.execute("SELECT * FROM `addtcinfo` WHERE `reg_no`= '%s' "%(f[1]))
                mys=db.fetchone()
                if(mys):
                  
                  pdf.add_page()
                  pdf.set_font("Arial",size=12)
                  name="alagappa.jpg"
                  pdf.image(name,w=200,h=20)
                    
                  pdf.cell(100,10,txt="                                                              TRANSFER CERTIFICATE")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     1.  Name of the student   ")
                  pdf.cell(100,10,txt=":              "+str(f[0]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     2.  Name of the parent/Guardian                               ")
                  pdf.cell(100,10,txt=":              "+str(f[3]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     3.  Nationality Religion and community                      ")
                  pdf.cell(100,10,txt=":              "+str(f[4])+' '+str(f[5])+' '+str(f[6]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     4.  Sex                                                            ")
                  pdf.cell(100,10,txt=":              "+str(f[8]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     5.  Date of Birth (in figure and words)                        ")
                  #l2=datetime.datetime.strptime(str(f[9]), "%Y-%m-%d").strftime("%d-%m-%Y")
                  pdf.cell(100,10,txt=":              "+str(f[9]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="         as entered in the admission register                        ")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     6.  Course of Study                                             ")
                  pdf.cell(100,10,txt=":              "+str(f[10])+'('+str(f[11])+')')
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     7.  Date of Admission to this college                          ")
                  #l3=datetime.datetime.strptime(str(f[12]), "%Y-%m-%d").strftime("%d-%m-%Y")
                  pdf.cell(100,10,txt=":              "+str(f[12]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     8.  a) Whether the Student has paid all the                   ")
                  pdf.cell(100,10,txt=":              "+str(mys[1]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="            Fees due to the college ?")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     8.  b) Whether the Student was in receipt of                  ")
                  pdf.cell(100,10,txt=":              "+str(mys[2]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="            any scholarship")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     9.  Whether the Student has undergone                          ")
                    
                  pdf.cell(100,10,txt=":              "+str(mys[3]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="          Medical inspection during the year")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     10.  Reasons for leaving the College                       ")
                    
                  pdf.cell(100,10,txt=":              "+str(mys[4]))
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     11.  Date of Leaving                                      ")
                  l6=datetime.datetime.strptime(str(getit), "%d-%m-%Y").strftime("%d-%m-%Y")
                  pdf.cell(100,10,txt=":              "+l6 )
                  
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     12.  Date on which application for Transfer  ")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="           Certificate was made by the Student or On                    ")
                  #l4=datetime.datetime.strptime(str(mys[5]), "%Y-%m-%d").strftime("%d-%m-%Y")
                  pdf.cell(100,10,txt=":              "+str(mys[5]) )
                  pdf.ln(10)
                  pdf.cell(100,10,txt="           his/her behalf by Parent/Guardian")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     13.  Date of the Transfer Certificate                         ")
                  #l5=datetime.datetime.strptime(str(f[24]), "%Y-%m-%d").strftime("%d-%m-%Y")
                  pdf.cell(100,10,txt=":              "+str(f[24]) )
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     14.  Medium                                                      ")
                  pdf.cell(100,10,txt=":              English")
                  pdf.ln(10)
                  pdf.cell(100,10,txt="     Seal")
                  pdf.cell(100,10,txt="              PRINCIPAL/VICE-PRINCIPAL")
                  pdf.ln(10)
            r=sentt+".pdf"
            pdf.output(r)
            if os.path.exists(r):
                     os.startfile(r)
                     self.Frame14.destroy()
                     self.Frame11.destroy()
                     self.right_frame.destroy()
                     self.left_frame.destroy()
                     self.master.switch_frame(pageOne)
            
class transfersingle(tk.Frame):
          
          def __init__(self, master):
                tk.Frame.__init__(self,master)
                self.configure(background='midnightblue')
                self.Frame14 = tk.Frame(self.master,bg="yellow",bd=3,relief=tk.RAISED)
                    
                self.Frame14.pack()
                adminname='logged in as '+username
                self.menubar = tk.Menu(self.Frame14,tearoff=0,bd=2)
                self.filemenu = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black')
                self.filemenu.add_command(label="Individual Transfer Certificate",command=self.transsing,font=("Times 18 italic"))
                self.filemenu.add_command(label="Batch Transfer Certificate",command=self.transgrpo,font=("Times 18 italic"))
                self.filemenu.add_command(label="Update TC",command=self.extratc,font=("Times 18 italic"))
                self.menubar.add_cascade(label="Transfer Certificate Generation", menu=self.filemenu,font = "Times 38")
                self.edit = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                self.edit.add_command(label="Edit",command=self.edt ,font=("Times 18 italic"))
                 
                self.menubar.add_cascade(label='Edit Details',menu=self.edit,font = "Times 38")
                self.insert = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                self.insert.add_command(label="Insert",command=self.ins,font=("Times 18 italic"))
                    
                self.menubar.add_cascade(label='Insert Details',menu=self.insert,font = "Times 38")
                self.search = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                self.search.add_command(label="Search",command=self.ser,font=("Times 18 italic"))
                
                self.menubar.add_cascade(label='Search Details',menu=self.search,font = "Times 38")
                self.menubar.add_command(label="logout",command=self.logg,font=("Times 18 italic"))
                self.menubar.add_command(label=adminname,font=("Times 18 italic"))
                     
                         #self.menubar.config("Verdana", 14)
                self.master.config(menu=self.menubar)
                self.Frame11 = tk.Frame(self.master)
                self.Frame11.pack(side="top",  pady=1,padx=1,expand=True )
                self.Frame11.configure(background='midnightblue')
                label1221=tk.Label(self.Frame11,text="Alagappa chettiar Government College of Engineering and Technology,Karaikudi",background= "midnightblue",foreground="white",font =
                                         ('Times', 20, 'bold'))
                label1221.grid(row=0,rowspan=1,column=1,columnspan=40)
                label12221=tk.Label(self.Frame11,text="(An autonomous government institution permanently affilitated to Anna University)",background= "midnightblue",foreground="white",font =
                                          ('Times', 12, 'bold'))
                label12221.grid(row=1,rowspan=1,column=1,columnspan=40)
                self.left_frame = tk.Frame(self, width=200, height=800, bg='skyblue3')
                self.left_frame.grid(row=1, column=0, padx=10, pady=5)
                self.right_frame = tk.Frame(self, width=650, height=400, bg='skyblue3')
                self.right_frame.grid(row=1, column=1, padx=10, pady=5)
 
# Create frames and labels in left_frame
                #tk.Label(self.left_frame, text="Original Image").grid(row=0, column=0, padx=5, pady=5)
                image = tk.PhotoImage(file="tc.png")
                original_image = image.subsample(1) # resize image using subsample
                label =tk.Label(self.left_frame,image = original_image)#,width=1680,height=1080)
                label.image = original_image # keep a reference!
                label.grid(row=1, column=0, padx=5, pady=5)
                
                
                self.l=ttk.Label(self,text="Individual Transfer Certificate Details",font=("Times 20 bold"),background= 'midnightblue',foreground="white")
                self.l.grid(row=0,column=1)
                
                self.l1=ttk.Label(self.right_frame,text="Register Number",font=("Times 15 bold"),background= "skyblue3",foreground="black")
                self.l1.grid(row=1,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.entr1=ttk.Entry(self.right_frame,font=("Times 15 bold"))
                self.entr1.grid(row=1,column=1,padx=5,pady=1,ipady=5,sticky='W')
                self.vieww=tk.IntVar()
                self.viewe=tk.IntVar()
                self.viiew=tk.IntVar()
                
                          
                self.le2=ttk.Label(self.right_frame,text="Leaving Date",font=("Times 15 bold"),background="skyblue3")
                self.le2.grid(row=2,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.e2=DateEntry(self.right_frame,  date_pattern='dd-mm-yyyy', font=("Times 15 bold"), borderwidth=2)
            
                self.e2.grid(row=2,column=1,padx=5,pady=1,ipady=5,sticky='W')
                
                
                self.l1=ttk.Label(self.right_frame,text="Whether the student has paid all the fees due to the college?",font=("Times 15 bold"),background= "skyblue3",foreground="black")
                self.l1.grid(row=5,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.r10=ttk.Radiobutton(self.right_frame, text="Yes",variable=self.vieww, value=1,style = 'Wild.TRadiobutton',command=self.feey)
                self.r10.grid(row=5,column=1,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                self.r11=ttk.Radiobutton(self.right_frame, text="No",variable=self.vieww, value=2,style = 'Wild.TRadiobutton',command=self.feex)
                self.r11.grid(row=5,column=2,sticky='W',padx=1,pady=10,ipadx=1,ipady=10)
                
                self.l2=ttk.Label(self.right_frame,text="Whether the student was in the receipt of any scholarship",font=("Times 15 bold"),background= "skyblue3",foreground="black")
                self.l2.grid(row=6,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.r20=ttk.Radiobutton(self.right_frame, text="Yes",variable=self.viewe, value=1,style = 'Wild.TRadiobutton',command=self.schy)
                self.r20.grid(row=6,column=1,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                self.r21=ttk.Radiobutton(self.right_frame, text="No",variable=self.viewe, value=2,style = 'Wild.TRadiobutton',command=self.schx)
                self.r21.grid(row=6,column=2,sticky='W',padx=1,pady=10,ipadx=1,ipady=10)
                
                self.l3=ttk.Label(self.right_frame,text="Whether the student has undergone Medical inspection during the year",font=("Times 15 bold"),background= "skyblue3",foreground="black")
                self.l3.grid(row=7,column=0,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                self.r30=ttk.Radiobutton(self.right_frame, text="Yes",variable=self.viiew, value=1,style = 'Wild.TRadiobutton',command=self.medy)
                self.r30.grid(row=7,column=1,sticky='W',padx=20,pady=10,ipadx=20,ipady=10)
                self.r31=ttk.Radiobutton(self.right_frame, text="No",variable=self.viiew, value=2,style = 'Wild.TRadiobutton',command=self.medx)
                self.r31.grid(row=7,column=2,sticky='W',padx=1,pady=10,ipadx=1,ipady=10)
                                                        
                self.l4=ttk.Label(self.right_frame,text="Reason for leaving the college",font=("Times 15 bold"),background= "skyblue3",foreground="black")
                self.l4.grid(row=8,column=0,padx=10,pady=30,sticky='W')
                self.e40=ttk.Entry(self.right_frame,font=("Times 15 bold"))
                self.e40.grid(row=8,column=1,padx=5,pady=5,ipady=5,sticky='W')
                
                self.l4=ttk.Label(self.right_frame,text="Date on which application for Transfer Certificate was ",font=("Times 15 bold"),background= "skyblue3",foreground="black")
                self.l4.grid(row=9,column=0,padx=10,pady=10,sticky='W')                
                self.l5=ttk.Label(self.right_frame,text="made by the student or on his/her behalf by parent/guardian",font=("Times 15 bold"),background= "skyblue3",foreground="black")
                self.l5.grid(row=10,column=0,padx=10,pady=1,sticky='W')
                self.e41=DateEntry(self.right_frame,  date_pattern='dd-mm-yyyy', font=("Times 15 bold"), borderwidth=2)
            
                self.e41.grid(row=10,column=1,padx=5,pady=1,ipady=5,sticky='W')
                #print(entr1.get())
                self.but=ttk.Button(self.right_frame,text="Submit",style = 'W.TButton',command=self.tranz)
                self.but.grid(row=11,column=1,padx=15,pady=5,ipady=5,sticky='W')
              
                #self.button16=ttk.Button(self,text="BACK",style = 'W.TButton',command=lambda: master.switch_frame(transferspecify))
                #self.button16.grid(row=11,column=0,sticky='W')
          def feey(self):
            self.feeey="yes"
          def feex(self):
                self.feeey="no"
          def schy(self):
                self.schhy="yes"
          def schx(self):
                self.schhy="no"
          def medy(self):
                self.meedy="yes"
          def medx(self):
                self.meedy="no"
          
          def transsing(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(transfersingle)
          def transgrpo(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(transfergroup)
          def extratc(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(updatetc)    
          def ins(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(newreg)
          def edt(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(editspecify)
          def ser(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(search)
          def logg(self):
              self.Frame14.destroy()
              self.Frame11.destroy()
              self.right_frame.destroy()
              self.left_frame.destroy()
              self.master.switch_frame(StartPage)      
        
          def tranz(self):
                self.reasn=self.e40.get()
                self.donw=self.e41.get()
                global setentvar,sete2var
                setentvar=self.entr1.get()
                #self.regnoo=self.e412.get()
                #self.mbso=self.e42.get()
                print(self.feeey)
                print(self.schhy)
                print(self.meedy)
                print(self.reasn)
                print(self.donw)
                #print(self.mbso)
                
                if(datetime.datetime.strptime(self.donw, '%d-%m-%Y') and self.feeey!="" and setentvar!="" and self.schhy!="" and self.meedy!="" and self.schhy!="" ):
                         sql0="INSERT INTO `addtcinfo`(`reg_no`, `student_bill`, `scholarship`, `medicalinspection`, `reasonforleaving`, `addofaplication`) VALUES(%s,%s,%s,%s,%s,%s)"
                         val0=(setentvar,self.feeey,self.schhy,self.meedy,self.reasn,self.donw)
                         db.execute(sql0,val0)
                         db_cur.commit() 
                         db.execute("update studentdetails set flag='1' where reg_number='%s'"%(setentvar))
                         db_cur.commit()             
                         #self.master.switch_frame(pageOne) 
                         
                         sete2var=self.e2.get()
                         #MyFirstGui().switch_frame(transferCertificate)
                         db.execute("select *from studentdetails where reg_number='%s'"%(setentvar))
                         f=db.fetchone()
                         db.execute("SELECT * FROM `addtcinfo` WHERE `reg_no`= '%s' "%(setentvar))
                         mys=db.fetchone()
                         if(f and mys):
                              reg=setentvar
                              msge = 'A transfercertificate has been generated for the register number'+reg
                              messagebox.showinfo("TransferCertificate", msge) 
                              pdf=FPDF()
                              pdf.add_page()
                              pdf.set_font("Arial",size=12)
                              name="alagappa.jpg"
                              pdf.image(name,w=200,h=20)
                                
                              pdf.cell(100,10,txt="                                                              TRANSFER CERTIFICATE")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     1.  Name of the student   ")
                              pdf.cell(100,10,txt=":              "+str(f[0]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     2.  Name of the parent/Guardian                               ")
                              pdf.cell(100,10,txt=":              "+str(f[3]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     3.  Nationality Religion and community                      ")
                              pdf.cell(100,10,txt=":              "+str(f[4])+' '+str(f[5])+' '+str(f[6]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     4.  Sex                                                            ")
                              pdf.cell(100,10,txt=":              "+str(f[8]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     5.  Date of Birth (in figure and words)                        ")
                              #l2=datetime.datetime.strptime(str(f[9]), "%Y-%m-%d").strftime("%d-%m-%Y")
                              pdf.cell(100,10,txt=":              "+str(f[9]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="         as entered in the admission register                        ")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     6.  Course of Study                                             ")
                              pdf.cell(100,10,txt=":              "+str(f[10])+'('+str(f[11])+')')
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     7.  Date of Admission to this college                          ")
                              #l3=datetime.datetime.strptime(str(f[12]), "%Y-%m-%d").strftime("%d-%m-%Y")
                              pdf.cell(100,10,txt=":              "+str(f[12]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     8.  a) Whether the Student has paid all the                   ")
                              pdf.cell(100,10,txt=":              "+str(mys[1]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="            Fees due to the college ?")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     8.  b) Whether the Student was in receipt of                  ")
                              pdf.cell(100,10,txt=":              "+str(mys[2]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="            any scholarship")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     9.  Whether the Student has undergone                          ")
                                
                              pdf.cell(100,10,txt=":              "+str(mys[3]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="          Medical inspection during the year")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     10.  Reasons for leaving the College                       ")
                                
                              pdf.cell(100,10,txt=":              "+str(mys[4]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     11.  Date of Leaving                                      ")
                              l6=datetime.datetime.strptime(str(sete2var), "%d-%m-%Y").strftime("%d-%m-%Y")
                              pdf.cell(100,10,txt=":              "+l6 )
                              
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     12.  Date on which application for Transfer  ")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="           Certificate was made by the Student or On                    ")
                              #l4=datetime.datetime.strptime(str(mys[5]), "%Y-%m-%d").strftime("%d-%m-%Y")
                              pdf.cell(100,10,txt=":              "+str(mys[5]) )
                              pdf.ln(10)
                              pdf.cell(100,10,txt="           his/her behalf by Parent/Guardian")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     13.  Date of the Transfer Certificate                         ")
                              #l5=datetime.datetime.strptime(str(f[24]), "%Y-%m-%d").strftime("%d-%m-%Y")
                              pdf.cell(100,10,txt=":              "+str(f[24]))
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     14.  Medium                                                      ")
                              pdf.cell(100,10,txt=":              English")
                              pdf.ln(10)
                              pdf.cell(100,10,txt="     Seal")
                              pdf.cell(100,10,txt="              PRINCIPAL/VICE-PRINCIPAL")
                              pdf.ln(10)
                              r=reg+".pdf"
                              pdf.output(r)
                              if os.path.exists(r):
                                           os.startfile(r)
                                           self.Frame14.destroy()
                                           self.Frame11.destroy()
                                           self.right_frame.destroy()
                                           self.left_frame.destroy()
                                           self.master.switch_frame(pageOne)
                              else:
                                  messagebox.showinfo("TC Error", "Enter A Valid Register Number")
                         else:
                             messagebox.showinfo("Transfer Error", "Fill all the fields and Check above details are True") 


      


class editspecify(tk.Frame):
        def __init__(self, master):
             tk.Frame.__init__(self,master)
             self.configure(background='skyblue3')
             self.Frame14 = tk.Frame(self.master,bg="yellow",bd=3,relief=tk.RAISED)
                    
             self.Frame14.pack()
             adminname='logged in as '+username
             self.menubar = tk.Menu(self.Frame14,tearoff=0,bd=2)
             self.filemenu = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black')
             self.filemenu.add_command(label="Individual Transfer Certificate",command=self.transsing,font=("Times 18"))
             self.filemenu.add_command(label="Batch Transfer Certificate",command=self.transgrpo,font=("Times 18"))
             self.filemenu.add_command(label="Update TC",command=self.extratc,font=("Times 18"))
             self.menubar.add_cascade(label="Transfer Certificate Generation", menu=self.filemenu,font = "Times 38")
             self.edit = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
             self.edit.add_command(label="Edit",command=self.edt ,font=("Times 18"))
             
             self.menubar.add_cascade(label='Edit Details',menu=self.edit,font = "Times 38")
             self.insert = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
             self.insert.add_command(label="Insert",command=self.ins,font=("Times 18"))
                
             self.menubar.add_cascade(label='Insert Details',menu=self.insert,font = "Times 38")
             self.search = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
             self.search.add_command(label="Search",command=self.ser,font=("Times 18"))
             
             self.menubar.add_cascade(label='Search Details',menu=self.search,font = "Times 38")
             self.menubar.add_command(label="Logout",command=self.logg,font=("Times 18"))
             self.menubar.add_command(label=adminname,font=("Times 18"))
                 
                     #self.menubar.config("Verdana", 14)
             self.master.config(menu=self.menubar)
             lablle1=ttk.Label(self,text="EDIT SPECIFICATION",font=("Times 30 bold"),background="skyblue3")
             lablle1.grid(row=0,column=0,pady=5)
             lp=ttk.Label(self,text="Register number",font=("Times 25 bold"),background="skyblue3")
             lp.grid(row=1,column=0,sticky='W',padx=35,pady=10,ipady=3)
             self.ent=ttk.Entry(self,font=("Times 18"))
             self.ent.grid(row=1,column=1,sticky='W',padx=5,pady=10,ipady=5)
             buton1=ttk.Button(self,text="submit",style = 'W.TButton',command=self.editz)
             buton1.grid(row=3,column=1,sticky='W',pady=50)
             
             #button18=ttk.Button(self,text="BACK",style = 'W.TButton',command=lambda: master.switch_frame(pageOne))
             #button18.grid(row=3,column=0,sticky='W',pady=50,padx=35)
        def transsing(self):
              self.Frame14.destroy()
              self.master.switch_frame(transfersingle)
        def transgrpo(self):
              self.Frame14.destroy()
              self.master.switch_frame(transfergroup)
        def extratc(self):
          self.Frame14.destroy()
          self.master.switch_frame(updatetc) 
        def ins(self):
              self.Frame14.destroy()
              self.master.switch_frame(newreg)
        def edt(self):
              self.Frame14.destroy() 
              self.master.switch_frame(editspecify)
        def ser(self):
              self.Frame14.destroy()
              self.master.switch_frame(search)
        def logg(self):
              self.Frame14.destroy()
              self.master.switch_frame(StartPage)
        def editz(self):
          global setvar
          setvar=self.ent.get()
          
          db.execute("select *from studentdetails where reg_number='%s'"%(setvar))
          e=db.fetchone()
          if(e):
              self.master.switch_frame(edit)
          else:
              messagebox.showinfo("Edit Error", "Enter A Valid Register Number")
          
                                                     
class edit(tk.Frame):
              def __init__(self, master):
                    tk.Frame.__init__(self, master)
                    self.Frame14 = tk.Frame(self.master,bg="yellow",bd=3,relief=tk.RAISED)
                    
                    self.Frame14.pack()
                    adminname='logged in as '+username
                    self.menubar = tk.Menu(self.Frame14,tearoff=0,bd=2)
                    self.filemenu = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black')
                    self.filemenu.add_command(label="Individual Transfer Certificate",command=self.transsing,font=("Times 18"))
                    self.filemenu.add_command(label="Batch Transfer Certificate",command=self.transgrpo,font=("Times 18"))
                    self.filemenu.add_command(label="Update TC",command=self.extratc,font=("Times 18"))
                    self.menubar.add_cascade(label="Transfer Certificate Generation", menu=self.filemenu,font = "Times 38")
                    self.edit = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                    self.edit.add_command(label="Edit",command=self.edt ,font=("Times 18"))
                    
                    self.menubar.add_cascade(label='Edit Details',menu=self.edit,font = "Times 38")
                    self.insert = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                    self.insert.add_command(label="Insert",command=self.ins,font=("Times 18"))
                
                    self.menubar.add_cascade(label='Insert Details',menu=self.insert,font = "Times 38")
                    self.search = tk.Menu(self.menubar, tearoff=0,background= "midnightblue",foreground="white",activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                    self.search.add_command(label="Search",command=self.ser,font=("Times 18"))
                
                    self.menubar.add_cascade(label='Search Details',menu=self.search,font = "Times 38")
                    self.menubar.add_command(label="Logout",command=self.logg,font=("Times 18"))
                    self.menubar.add_command(label=adminname,font=("Times 18"))
                 
                     #self.menubar.config("Verdana", 14)
                    self.master.config(menu=self.menubar)
                    self.Frame11 = tk.Frame(self.master)
                    self.Frame11.pack(side="top",  pady=1,padx=1,expand=True )
                    self.Frame11.configure(background='midnightblue')
                    label1221=tk.Label(self.Frame11,text="Alagappa chettiar Government College of Engineering and Technology,Karaikudi",background= "midnightblue",foreground="white",font =
                                         ('Times', 15, 'bold'))
                    label1221.grid(row=0,rowspan=1,column=1,columnspan=40)
                    label12221=tk.Label(self.Frame11,text="(An autonomous government institution permanently affilitated to Anna University)",background= "midnightblue",foreground="white",font =
                                          ('Times', 8, 'bold'))
                    label12221.grid(row=1,rowspan=1,column=1,columnspan=40)
                    #self._rootwindow.bind('<Configure>', self.onResize)
                    self.configure(background='midnightblue')
                    db.execute("select *from studentdetails where reg_number='%s'"%(setvar))
                    e=db.fetchone()
                    if(e):
                        self.Frame1 = tk.Frame(self.master)
                        self.Frame1.pack(side="top",padx=30,expand=True )
                        self.Frame1.configure(background='midnightblue')                        
                        
                        self.labe=ttk.Label(self.Frame11,text="EDIT",font=("Times 15 bold"),background='midnightblue',foreground="white")
                        self.labe.grid(row=0,column=0,sticky='W',pady=5,padx=35)
                        
                        self.Frame2 = tk.Frame(self.master)
                        self.Frame2.pack(side="left",ipady=40,expand=True )
                        self.Frame2.configure(background='midnightblue')                        
                        
                        self.Frame4 = tk.Frame(self.Frame2)
                        self.Frame4.pack(side="top",expand=True,ipadx=15,ipady=0 )
                        self.Frame4.configure(background='skyblue3')
                        
                        self.l1=ttk.Label(self.Frame4,text="Name Of The Candidate",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l1.grid(row=1,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll1=ttk.Label(self.Frame4,text=":",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll1.grid(row=1,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e1=ttk.Entry(self.Frame4,font=("Times 14"),width=23)
                        self.e1.grid(row=1,column=2,sticky='W',padx=20,pady=5,ipady=3)
                        self.e1.insert(0,e[0])
                        
                        self.l4=ttk.Label(self.Frame4,text="Father's Name",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l4.grid(row=2,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll4=ttk.Label(self.Frame4,text=":",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll4.grid(row=2,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e4=ttk.Entry(self.Frame4,font=("Times 14"),width=23)
                        self.e4.grid(row=2,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.e4.insert(0,e[3])
                        
                        self.l9=ttk.Label(self.Frame4,text="Sex",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l9.grid(row=3,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll9=ttk.Label(self.Frame4,text=":",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll9.grid(row=3,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e9 = tk.StringVar(self)
                        self.p4 =ttk.Combobox(self.Frame4, width=21, textvariable=self.e9,font=("Times 14"))
                        self.p4['values']=(e[8],'Male','Female','Others')  
                        self.p4.grid(row=3,column=2,sticky='W',padx=20,pady=5,ipady=5)
                        self.p4.current(0)
                           
                        self.l6=ttk.Label(self.Frame4,text="Religion",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l6.grid(row=5,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll6=ttk.Label(self.Frame4,text=":",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll6.grid(row=5,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e6 = tk.StringVar(self)
                        self.p2 =ttk.Combobox(self.Frame4, width=21, textvariable=self.e6,font=("Times 14"))
                        self.p2['values']=(e[5],'Hindu','Musim','Christian','Others')
                        self.p2.grid(row=5,column=2,sticky='W',padx=20,pady=5,ipady=5)
                        self.p2.current(0)
                        
                        self.l7=ttk.Label(self.Frame4,text="Caste",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l7.grid(row=7,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll7=ttk.Label(self.Frame4,text=":",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll7.grid(row=7,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.tt3 = tk.StringVar(self)
                        self.pp3 =ttk.Combobox(self.Frame4, width=21, textvariable=self.tt3,font=("Times 12"))
                        
                        self.pp3.grid(row=7,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                        self.pp3['values']=(e[7])
                        self.pp3.current(0)
                        self.l8=ttk.Label(self.Frame4,text="Community",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l8.grid(row=6,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll8=ttk.Label(self.Frame4,text=":",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll8.grid(row=6,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e8 = tk.StringVar(self)
                        self.p3 =ttk.Combobox(self.Frame4, width=21, textvariable=self.e8,font=("Times 14"))
                        self.p3['values']=(e[6],'OC','BC','BCM','MBC','SC','ST','Others')  
                        self.p3.grid(row=6,column=2,sticky='W',padx=20,pady=5,ipady=5)
                        self.p3.current(0)
                        
                        def valcom(event):  
                          self.aadharno=self.e8.get()
                          
                          if(self.aadharno == 'BC'):
                              print('ok')
                              self.pp3['values']=('OC','BCM','MBC','SC','ST','Others') 
                          
                        self.p3.bind('<FocusOut>', valcom)
                        
                                              
                
                        
                        self.l14=ttk.Label(self.Frame4,text="Mother Tongue",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l14.grid(row=8,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll14=ttk.Label(self.Frame4,text=":",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll14.grid(row=8,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e16 = tk.StringVar(self)
                        self.p7 =ttk.Combobox(self.Frame4, width=21, textvariable=self.e16,font=("Times 14"))
                        self.p7['values']=(e[15],'Tamil','Hindi','Telugu','Kannada','Malayalam','Sourashtra','Others')  
                        self.p7.grid(row=8,column=2,sticky='W',padx=20,pady=5,ipady=5)
                        self.p7.current(0)
                        
                        self.Frame5 = tk.Frame(self.Frame2)
                        self.Frame5.pack(side="bottom",expand=True,ipadx=12,ipady=1)
                        self.Frame5.configure(background='skyblue3')
                        
                        self.l5=ttk.Label(self.Frame5,text="Nationality",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l5.grid(row=1,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll5=ttk.Label(self.Frame5,text="            :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll5.grid(row=1,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e5 = tk.StringVar(self)
                        self.pp1 =ttk.Combobox(self.Frame5, width=21, textvariable=self.e5,font=("Times 14"))
                        self.pp1['values']=(e[4],'Indian','Others')
                        self.pp1.grid(row=1,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                        self.pp1.current(0)
                        
                        self.l14=ttk.Label(self.Frame5,text="Present address  ",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l14.grid(row=2,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll14=ttk.Label(self.Frame5,text="            :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll14.grid(row=2,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e18=ttk.Entry(self.Frame5,font=("Times 14"),width=23)
                        self.e18.grid(row=2,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.e18.insert(0,e[17])

                        self.l19=ttk.Label(self.Frame5,text="Taluk ",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l19.grid(row=3,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll19=ttk.Label(self.Frame5,text="            :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll19.grid(row=3,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e19=ttk.Entry(self.Frame5,font=("Times 14"),width=23)
                        self.e19.grid(row=3,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.e19.insert(0,e[18])

                        self.l20=ttk.Label(self.Frame5,text="City",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l20.grid(row=4,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll20=ttk.Label(self.Frame5,text="            :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll20.grid(row=4,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e20=ttk.Entry(self.Frame5,font=("Times 14"),width=23)
                        self.e20.grid(row=4,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.e20.insert(0,e[19])

                        self.l21=ttk.Label(self.Frame5,text="District",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l21.grid(row=5,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll21=ttk.Label(self.Frame5,text="            :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll21.grid(row=5,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e21= tk.StringVar(self)
                        self.p9 =ttk.Combobox(self.Frame5, width=21, textvariable=self.e21,font=("Times 14"))
                        self.p9['values']=(e[20],'Ariyalur','Chengalpet','Chennai','Coimbatore','Cuddalore','Dharmapuri','Dindigul','Erode','Kallakurichi','Kancheepuram','Karur','Krishnagiri','Madurai','Nagapattinam','Nilgiris','Kanyakumari','Namakkal','Perambalur','Pudukottai','Ramanathapuram','Ranipet','Salem','Sivagangai','Tenkasi','Thanjavur','Theni','Thiruvallur','Thiruvarur','Tuticorin','Trichirappalli','Thirunelveli','Tirupattur','Tiruppur','Thiruvannamalai','Vellore','Viluppuram','Virudhunagar')  
                        self.p9.grid(row=5,column=2,sticky='W',padx=20,pady=5,ipady=5)
                        self.p9.current(0)
                        
                        self.l17=ttk.Label(self.Frame5,text="Name Of the State",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l17.grid(row=6,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll17=ttk.Label(self.Frame5,text="            :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll17.grid(row=6,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e17 = tk.StringVar(self)
                        self.p8 =ttk.Combobox(self.Frame5, width=21, textvariable=self.e17,font=("Times 14"))
                        self.p8['values']=(e[16],'Andhra Pradesh','Arunachal Pradesh','Assam','Bihar','Chhattisgarh','Goa','Gujarat','Haryana','Himachal Pradesh','Jammu and Kashmir','Jharkhand','Karnataka','Kerala','Madya Pradesh','Maharashtra','Manipur','Meghalaya','Mizoram','Nagaland','Orissa','Punjab','Rajasthan','Sikkim','Tamil Nadu','Telagana','Tripura','Uttaranchal','Uttar Pradesh','West Bengal','Andaman and Nicobar','Chandigarh','Dadar and Nagar Haveli','Daman and Diu','Delhi','Lakshadeep','Pondicherry')  
                        self.p8.grid(row=6,column=2,sticky='W',padx=20,pady=5,ipady=5)
                        self.p8.current(0)
                        
                        self.Frame6 = tk.Frame(self.Frame2)
                        self.Frame6.pack(side="bottom",expand=True,ipadx=15)
                        self.Frame6.configure(background='midnightblue')
                        
                        self.l100=ttk.Label(self.Frame6,text="",font=("Times 7 italic"),background='midnightblue')
                        self.l100.grid(row=1,column=0,sticky='W')
                        
                        self.Frame3 = tk.Frame(self.master)
                        self.Frame3.pack(side="right",ipady=30,expand=True )
                        self.Frame3.configure(background='skyblue3')                        
                        
                        self.Frame7 = tk.Frame(self.Frame3)
                        self.Frame7.pack(side="top",expand=True,ipady=1)
                        self.Frame7.configure(background='skyblue3')
                        
                        self.l10=ttk.Label(self.Frame7,text="Date Of Birth",font=("Times 14 bold"),background='skyblue3', foreground="white")
                        self.l10.grid(row=1,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll10=ttk.Label(self.Frame7,text="          :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll10.grid(row=1,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e10=DateEntry(self.Frame7,width=23,date_pattern='dd-mm-yyyy',font=("Times 14"), borderwidth=2)
            
                        self.e10.grid(row=1,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        
                        self.e10.insert(0,e[9])
                        
                        self.la23=ttk.Label(self.Frame7,text="Aadhar Number",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.la23.grid(row=2,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.lla23=ttk.Label(self.Frame7,text="          :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.lla23.grid(row=2,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.ea23=ttk.Entry(self.Frame7,font=("Times 14"),width=23)
                        self.ea23.grid(row=2,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.ea23.insert(0,e[22])
                        def valaad(event):  
                              regex1 = '^\d{4}\d{4}\d{4}$'
                              self.aadharno=self.ea23.get()
                              
                              if(re.search(regex1,self.aadharno)):
                                  print('ok')
                              else:
                                     messagebox.showinfo("Insert Error", "Enter a valid aadhar number")
                        self.ea23.bind('<FocusOut>', valaad)                                           
                
                        
                        
                        self.la22=ttk.Label(self.Frame7,text="Mobile Number",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.la22.grid(row=3,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.lla22=ttk.Label(self.Frame7,text="          :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.lla22.grid(row=3,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.ea22=ttk.Entry(self.Frame7,font=("Times 14"),width=23)
                        self.ea22.grid(row=3,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.ea22.insert(0,e[21])
                        
                        def valphno(event):  
                          regex3 = '(0/91)?[7-9][0-9]{9}'
                          self.cellno=self.ea22.get()
                          
                          if(re.search(regex3,self.cellno)):
                              print('ok')
                          else:
                                 messagebox.showinfo("Insert Error", "Enter a valid phone number")
                        self.ea22.bind('<FocusOut>', valphno)
                        
                        self.l2=ttk.Label(self.Frame7,text="Register Number",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l2.grid(row=4,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll2=ttk.Label(self.Frame7,text="          :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll2.grid(row=4,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e2=ttk.Entry(self.Frame7,font=("Times 14"),width=23)
                        self.e2.grid(row=4,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.e2.insert(0,e[1])
                        
                        self.l3=ttk.Label(self.Frame7,text="Roll Number",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l3.grid(row=5,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll3=ttk.Label(self.Frame7,text="          :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll3.grid(row=5,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e3=ttk.Entry(self.Frame7,font=("Times 14"),width=23)
                        self.e3.grid(row=5,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.e3.insert(0,e[2])
                        
                        self.l11=ttk.Label(self.Frame7,text="Course Of Study",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l11.grid(row=6,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll11=ttk.Label(self.Frame7,text="          :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll11.grid(row=6,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e11 = tk.StringVar(self)
                        self.p5 =ttk.Combobox(self.Frame7, width=21, textvariable=self.e11,font=("Times 14"))
                        self.p5['values']=(e[10],'BE','ME','BE Parttime')  
                        self.p5.grid(row=6,column=2,sticky='W',padx=20,pady=5,ipady=5)
                        self.p5.current(0)
                        
                        self.l12=ttk.Label(self.Frame7,text="Branch Of Study",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l12.grid(row=7,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll12=ttk.Label(self.Frame7,text="          :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll12.grid(row=7,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e12 = tk.StringVar(self)
                        self.p6 =ttk.Combobox(self.Frame7,width=21, textvariable=self.e12,font=("Times 14"))
                        self.p6['values']=(e[11],'Civil Engineering','Mechanical Engineering','Electrical and Electronics Engineering','Electronics and Communication Engineering','Computer science and Engineering')  
                        self.p6.grid(row=7,column=2,sticky='W',padx=20,pady=5,ipady=5)
                        self.p6.current(0)
                        
                        self.Frame8 = tk.Frame(self.Frame3)
                        self.Frame8.pack(side="top",expand=True,ipady=1)
                        self.Frame8.configure(background='skyblue3')
                        
                        self.l13=ttk.Label(self.Frame8,text="Admitted On ",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l13.grid(row=1,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll13=ttk.Label(self.Frame8,text="           :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll13.grid(row=1,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e13=ttk.Entry(self.Frame8,font=("Times 14"),width=23)
                        self.e13.grid(row=1,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        
                        self.e13.insert(0,e[12])
                        
                        self.l14=ttk.Label(self.Frame8,text="Receipt Number",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l14.grid(row=2,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll14=ttk.Label(self.Frame8,text="           :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll14.grid(row=2,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e14=ttk.Entry(self.Frame8,font=("Times 14"),width=23)
                        self.e14.grid(row=2,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.e14.insert(0,e[13])

                        self.l15=ttk.Label(self.Frame8,text="Receipt Date",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.l15.grid(row=3,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.ll15=ttk.Label(self.Frame8,text="           :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.ll15.grid(row=3,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.e15=DateEntry(self.Frame8,width=23,date_pattern='dd-mm-yyyy', font=("Times 14"), borderwidth=2)
            
                        self.e15.grid(row=3,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.e15.insert(0,e[14])

                        self.la25=ttk.Label(self.Frame8,text="Issued ON",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.la25.grid(row=4,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.lla25=ttk.Label(self.Frame8,text="           :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.lla25.grid(row=4,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.ea25=DateEntry(self.Frame8,width=23,date_pattern='dd-mm-yyyy', font=("Times 14"), borderwidth=2)
            
                        self.ea25.grid(row=4,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        
                        self.ea25.insert(0,e[24])
                         
                        self.la24=ttk.Label(self.Frame8,text="T.C.No",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.la24.grid(row=5,column=0,sticky='W',pady=5,padx=20,ipady=5)
                        self.lla24=ttk.Label(self.Frame8,text="           :",font=("Times 14 bold"),background='skyblue3',foreground="white")
                        self.lla24.grid(row=5,column=1,sticky='W',pady=5,padx=20,ipady=5)
                        self.ea24=ttk.Entry(self.Frame8,font=("Times 14"),width=23)
                        self.ea24.grid(row=5,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        self.ea24.insert(0,e[23]) 
                        
                        
                        self.Frame9 = tk.Frame(self.Frame3)
                        self.Frame9.pack(side="bottom",expand=True,ipadx=10)
                        self.Frame9.configure(background='skyblue3')
                                                
                        self.llla24=ttk.Button(self.Frame9,style = 'W.TButton',text="Submit",command=self.valme)
                        self.llla24.grid(row=1,column=2,sticky='W',padx=35,pady=3,ipady=1)
                        self.button19=ttk.Button(self.Frame9,text="Back",style = 'W.TButton',command=self.edt)
                        self.button19.grid(row=1,column=0,sticky='W',pady=5,padx=35)
                        
                        print(self.e1.get())
              def transsing(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(transfersingle)
              def transgrpo(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(transfergroup)
              def ins(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(newreg)
              def edt(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(editspecify)
              def extratc(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(updatetc)    
              def ser(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(search)
              def logg(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(StartPage)          
              def valme(self):
                regex1 = '^\d{4}\d{4}\d{4}$'
               
                regex3 = '(0/91)?[7-9][0-9]{9}'
                #print(self.e13.get())
                
                #day,month,year = self.dob.split('-')
                #d,m,y=self.receiptdate.split('-')
                #d1,m1,y1=self.issuedon.split('-')
                try:
                    if(re.search(regex1,self.ea23.get())):  
                         if(re.search(regex3,self.ea22.get())):  
                            datetime.datetime.strptime(self.e10.get(), '%d-%m-%Y')
                            datetime.datetime.strptime(self.e13.get(), '%d-%m-%Y')
                            datetime.datetime.strptime(self.e15.get(), '%d-%m-%Y')
                            if(self.e1.get()!="" and self.e3.get()!="" and self.e4.get()!="" and self.e5.get()!="" and self.e6.get()!="" and self.tt3.get()!="" and self.e8.get() !="" and self.e9.get(),self.e10.get(),self.e11.get(),self.e12.get(),self.e13.get(),self.e14.get(),self.e15.get()!="" and self.e16.get()!="" and self.e17.get() !="" and self.e18.get()!="" and self.e19.get()!="" and self.e20.get()!="" and self.e21.get()!="" and self.ea22.get()!="" and self.ea23.get()!="" ):
                                self.callme()
                                
                            else:
                             messagebox.showinfo("Insert Error", "Fill all the fields")
                                
                         else:  
                             messagebox.showinfo("Insert Error", "Not valid phone no")
                    else:  
                        messagebox.showinfo("Insert Error", "Not valid aadhar no") 
                    
                    
                    

                except ValueError :
                    messagebox.showinfo("Insert Error", "Not valid date no")

                
                  
              def callme(self):
                       
                        print(self.e1.get())
                        print(self.e1.get())
                        print(self.e1.get())
                        self.name=self.e1.get()
                        self.regno=self.e2.get()
                        self.rollno=self.e3.get()
                        self.father=self.e4.get()
                        self.nation=self.e5.get()
                        self.religion=self.e5.get()
                        self.caste=self.tt3.get()
                        self.community=self.e8.get()
                        self.sex=self.e9.get()
                        self.dob=self.e10.get()
                        print(str(self.e10.get()))
                        print(self.dob)
                        self.course=self.e11.get()
                        self.branch=self.e12.get()
                        self.admiton=self.e13.get()
                        self.receiptno=self.e14.get()
                        self.receiptdate=self.e15.get()
                        self.mothertongue=self.e16.get()
                        self.state=self.e17.get()
                        self.address=self.e18.get()
                        self.taluk=self.e19.get()
                        self.city=self.e20.get()  
                        self.district=self.e21.get()
                        self.cellno=self.ea22.get()
                        self.aadharno=self.ea23.get()
                        self.tcno=self.ea24.get()
                        self.issuedon=self.ea25.get()
                        #self.yopo=self.ea26.get()
                        
                        sql="UPDATE `studentdetails` SET `name`= %s,`father_name`=%s,`nationality`=%s,`religion`=%s,`caste`=%s,`community`=%s,`sex`=%s,`dateofbirth`=%s,`course`=%s,`branch`=%s,`admittedon`=%s,`receiptno`=%s,`receiptdate`=%s,`mothertongue`=%s,`state`=%s,`present_address`=%s,taluk=%s,`city`=%s,`district`=%s,`cell_number`=%s,`aadhar_number`=%s,`tcno`=%s,`issuedon`=%s WHERE reg_number=%s"
                        val=(self.e1.get(),self.e4.get(),self.e5.get(),self.e6.get(),self.e8.get(),self.tt3.get(),self.e9.get(),self.e10.get(),self.e11.get(),self.e12.get(),self.e13.get(),self.e14.get(),self.e15.get(),self.e16.get(),self.e17.get(),self.e18.get(),self.e19.get(),self.e20.get(),self.e21.get(),self.ea22.get(),self.ea23.get(),self.ea24.get(),self.ea25.get(),setvar)
                        db.execute(sql, val)
                        db_cur.commit()
                        messagebox.showinfo("Edit ", "All the datas are updated ")
                        
                        self.Frame14.destroy()
                        self.Frame11.destroy()
                        self.Frame1.destroy()
                        self.Frame2.destroy()
                        self.Frame3.destroy()
                        self.Frame4.destroy()
                        self.Frame5.destroy()
                        self.Frame6.destroy()
                        self.Frame7.destroy()
                        self.Frame8.destroy()
                        self.master.switch_frame(pageOne)
                        
                    
                
           
class newreg(tk.Frame):
              def __init__(self, master):
                tk.Frame.__init__(self,master)
                self.configure(background='midnightblue')
                #db.execute("select *from studentdetails where reg_number='%s'"%(setvar))
                #e=db.fetchone()
                adminname='logged in as '+username
                self.Frame14 = tk.Frame(self.master,bg="yellow",bd=3,relief=tk.RAISED)
                    
                self.Frame14.pack()
                
                self.menubar = tk.Menu(self.Frame14,tearoff=0,bd=2)
                self.filemenu = tk.Menu(self.menubar, tearoff=0,background='midnightblue', foreground='white',activebackground='cyan', activeforeground='black')
                self.filemenu.add_command(label="Individual Transfer Certificate",command=self.transsing,font=("Times 18"))
                self.filemenu.add_command(label="Batch Transfer Certificate",command=self.transgrpo,font=("Times 18"))
                self.filemenu.add_command(label="Update TC",command=self.extratc,font=("Times 18"))
                self.menubar.add_cascade(label="Transfer Certificate Generation", menu=self.filemenu,font = "Times 38")
                self.edit = tk.Menu(self.menubar, tearoff=0,background='midnightblue', foreground='white',activebackground='cyan', activeforeground='black',font=("Times 18"))
                self.edit.add_command(label="Edit",command=self.edt ,font=("Times 18"))
                 
                self.menubar.add_cascade(label='Edit Details',menu=self.edit,font = "Times 38")
                self.insert = tk.Menu(self.menubar, tearoff=0,background='midnightblue', foreground='white',activebackground='cyan', activeforeground='black',font=("Times 18"))
                self.insert.add_command(label="Insert",command=self.ins,font=("Times 18"))
             
                self.menubar.add_cascade(label='Insert Details',menu=self.insert,font = "Times 38")
                self.search = tk.Menu(self.menubar, tearoff=0,background='midnightblue', foreground='white',activebackground='cyan', activeforeground='black',font=("Times 18"))
                self.search.add_command(label="Search",command=self.ser,font=("Times 18"))
             
                self.menubar.add_cascade(label='Search Details',menu=self.search,font = "Times 38")
                self.menubar.add_command(label="Logout",command=self.logg,font=("Times 18"))
                self.menubar.add_command(label=adminname,font=("Times 18"))
             
                 #self.menubar.config("Verdana", 14)
                self.master.config(menu=self.menubar)
                
                self.Frame11 = tk.Frame(self.master)
                self.Frame11.pack(side="top",  pady=1,padx=1,expand=True )
                self.Frame11.configure(background='midnightblue')
                label1221=tk.Label(self.Frame11,text="Alagappa chettiar Government College of Engineering and Technology,Karaikudi",background='midnightblue', foreground='white',font =
                                         ('Times', 15, 'bold'))
                label1221.grid(row=0,column=1,rowspan=1,columnspan=40)
                label12221=tk.Label(self.Frame11,text="(An autonomous government institution permanently affilitated to anna university)",background='midnightblue', foreground='white',font =
                                          ('Times', 8, 'bold'))
                label12221.grid(row=1,rowspan=1,column=1,columnspan=40)
                self.Frame1 = tk.Frame(self.master)
                self.Frame1.pack(side="top",padx=30,expand=True )
                self.Frame1.configure(background='midnightblue')                        
                        
                self.labe=ttk.Label(self.Frame11,text="INSERT RECORD",font=("Times 15 bold"),background='midnightblue',foreground="white")
                self.labe.grid(row=0,column=0,sticky='W',pady=5,padx=35)
                        
                self.Frame2 = tk.Frame(self.master)
                self.Frame2.pack(side="left",ipady=35,expand=True )
                self.Frame2.configure(background='midnightblue')                        
                        
                self.Frame4 = tk.Frame(self.Frame2)
                self.Frame4.pack(side="top",expand=True ,ipadx=20)
                self.Frame4.configure(background='skyblue3')
                        
                self.l1=ttk.Label(self.Frame4,text="Name Of The Candidate",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l1.grid(row=1,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll1=ttk.Label(self.Frame4,text=":",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll1.grid(row=1,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e1=ttk.Entry(self.Frame4,font=("Times 12"),width=24)
                self.e1.grid(row=1,column=2,sticky='W',padx=20,pady=5,ipady=3)
                
                self.l4=ttk.Label(self.Frame4,text="Father's Name",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l4.grid(row=2,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll4=ttk.Label(self.Frame4,text=":",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll4.grid(row=2,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e4=ttk.Entry(self.Frame4,font=("Times 12"),width=24)
                self.e4.grid(row=2,column=2,sticky='W',padx=20,pady=3,ipady=3)
                
                self.l9=ttk.Label(self.Frame4,text="Sex",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l9.grid(row=3,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll9=ttk.Label(self.Frame4,text=":",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll9.grid(row=3,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.t4 = tk.StringVar(self)
                self.p4 =ttk.Combobox(self.Frame4, width=21, textvariable=self.t4,font=("Times 12"))
                self.p4['values']=('Male','Female','Others')  
                self.p4.grid(row=3,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                self.p4.current()
                
                self.l8=ttk.Label(self.Frame4,text="Community",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l8.grid(row=6,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll8=ttk.Label(self.Frame4,text=":",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll8.grid(row=6,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.t3 = tk.StringVar(self)
                self.p3 =ttk.Combobox(self.Frame4, width=21, textvariable=self.t3,font=("Times 12"))
                self.p3['values']=('OC','BC','BCM','MBC','SC','ST','Others')  
                self.p3.grid(row=6,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                self.p3.current()
                def valcom(event):  
                  self.aadharno=self.t3.get()
                  
                  if(self.aadharno == 'BC'):
                      print('ok')
                      self.pp3['values']=('OC','BCM','MBC','SC','ST','Others') 
                  
                self.p3.bind('<FocusOut>', valcom)
                
                self.l7=ttk.Label(self.Frame4,text="Caste",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l7.grid(row=7,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll7=ttk.Label(self.Frame4,text=":",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll7.grid(row=7,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.tt3 = tk.StringVar(self)
                self.pp3 =ttk.Combobox(self.Frame4, width=21, textvariable=self.tt3,font=("Times 12"))
                
                self.pp3.grid(row=7,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                self.pp3.current()
                
                #self.e7=ttk.Entry(self.Frame4,font=("Times 12"),width=24)
                #self.e7.grid(row=7,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        
                
                self.l6=ttk.Label(self.Frame4,text="Religion",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l6.grid(row=5,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll6=ttk.Label(self.Frame4,text=":",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll6.grid(row=5,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.t2 = tk.StringVar(self)
                self.p2 =ttk.Combobox(self.Frame4, width=21, textvariable=self.t2,font=("Times 12"))
                self.p2['values']=('Hindu','Musim','Christian','Others')
                self.p2.grid(row=5,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                self.p2.current()
                
                
                
                self.l14=ttk.Label(self.Frame4,text="Mother Tongue",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l14.grid(row=8,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll14=ttk.Label(self.Frame4,text=":",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll14.grid(row=8,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.t7 = tk.StringVar(self)
                self.p7 =ttk.Combobox(self.Frame4, width=21, textvariable=self.t7,font=("Times 12"))
                self.p7['values']=('Tamil','Hindi','Telugu','Kannada','Malayalam','Sourashtra','Others')  
                self.p7.grid(row=8,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                self.p7.current()
                        
                self.Frame5 = tk.Frame(self.Frame2)
                self.Frame5.pack(side="top",expand=True,ipadx=20)
                self.Frame5.configure(background='skyblue3')
                
                self.l5=ttk.Label(self.Frame5,text="Nationality",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l5.grid(row=1,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll5=ttk.Label(self.Frame5,text="            :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll5.grid(row=1,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.t1 = tk.StringVar(self)
                self.pp1 =ttk.Combobox(self.Frame5, width=21, textvariable=self.t1,font=("Times 12"))
                self.pp1['values']=('Indian','Others')
                self.pp1.grid(row=1,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                self.pp1.current()
                
                self.l14=ttk.Label(self.Frame5,text="Present address  ",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l14.grid(row=2,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll14=ttk.Label(self.Frame5,text="            :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll14.grid(row=2,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e18=ttk.Entry(self.Frame5,font=("Times 12"),width=24)
                self.e18.grid(row=2,column=2,sticky='W',padx=20,pady=3,ipady=3)
                
                self.l19=ttk.Label(self.Frame5,text="Taluk ",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l19.grid(row=3,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll19=ttk.Label(self.Frame5,text="            :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll19.grid(row=3,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e19=ttk.Entry(self.Frame5,font=("Times 12"),width=24)
                self.e19.grid(row=3,column=2,sticky='W',padx=20,pady=3,ipady=3)
                              

                self.l20=ttk.Label(self.Frame5,text="City",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l20.grid(row=4,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll20=ttk.Label(self.Frame5,text="            :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll20.grid(row=4,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e20=ttk.Entry(self.Frame5,font=("Times 12"),width=24)
                self.e20.grid(row=4,column=2,sticky='W',padx=20,pady=3,ipady=3)            
            
               
                self.l21=ttk.Label(self.Frame5,text="District",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l21.grid(row=5,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll21=ttk.Label(self.Frame5,text="            :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll21.grid(row=5,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.t9= tk.StringVar(self)
                self.p9 =ttk.Combobox(self.Frame5, width=21, textvariable=self.t9,font=("Times 12"))
                self.p9['values']=('Ariyalur','Chengalpet','Chennai','Coimbatore','Cuddalore','Dharmapuri','Dindigul','Erode','Kallakurichi','Kancheepuram','Karur','Krishnagiri','Madurai','Nagapattinam','Nilgiris','Kanyakumari','Namakkal','Perambalur','Pudukottai','Ramanathapuram','Ranipet','Salem','Sivagangai','Tenkasi','Thanjavur','Theni','Thiruvallur','Thiruvarur','Tuticorin','Trichirappalli','Thirunelveli','Tirupattur','Tiruppur','Thiruvannamalai','Vellore','Viluppuram','Virudhunagar')  
                self.p9.grid(row=5,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5) 
                
                self.l17=ttk.Label(self.Frame5,text="Name Of State",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l17.grid(row=6,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll17=ttk.Label(self.Frame5,text="            :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll17.grid(row=6,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.t8 = tk.StringVar(self)
                self.p8 =ttk.Combobox(self.Frame5, width=21, textvariable=self.t8,font=("Times 12"))
                self.p8['values']=('Andhra Pradesh','Arunachal Pradesh','Assam','Bihar','Chhattisgarh','Goa','Gujarat','Haryana','Himachal Pradesh','Jammu and Kashmir','Jharkhand','Karnataka','Kerala','Madya Pradesh','Maharashtra','Manipur','Meghalaya','Mizoram','Nagaland','Orissa','Punjab','Rajasthan','Sikkim','Tamil Nadu','Telagana','Tripura','Uttaranchal','Uttar Pradesh','West Bengal','Andaman and Nicobar','Chandigarh','Dadar and Nagar Haveli','Daman and Diu','Delhi','Lakshadeep','Pondicherry')  
                self.p8.grid(row=6,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                self.p8.current()
                      
                             
                self.Frame6 = tk.Frame(self.Frame2)
                self.Frame6.pack(side="bottom",expand=True,ipadx=20)
                self.Frame6.configure(background='midnightblue')
                        
                #self.l100=ttk.Label(self.Frame6,text="",font=("Times 7 italic"),background='midnightblue')
                #self.l100.grid(row=1,column=0,sticky='W')
                        
                self.Frame3 = tk.Frame(self.master)
                self.Frame3.pack(side="right",expand=True,ipadx=30,ipady=30 )
                self.Frame3.configure(background='midnightblue')                        
                        
                self.Frame7 = tk.Frame(self.Frame3)
                self.Frame7.pack(side="top",expand=True)
                self.Frame7.configure(background='skyblue3')
                
                self.l10=ttk.Label(self.Frame7,text="Date Of Birth",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l10.grid(row=1,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll10=ttk.Label(self.Frame7,text="          :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll10.grid(row=1,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e10=DateEntry(self.Frame7,width=23,date_pattern='dd-mm-yyyy', font=("Times 12"), borderwidth=2)
            
                self.e10.grid(row=1,column=2,sticky='W',padx=20,pady=3,ipady=3)
                                
                        
                self.la23=ttk.Label(self.Frame7,text="Aadhar Number",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.la23.grid(row=2,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.lla23=ttk.Label(self.Frame7,text="          :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.lla23.grid(row=2,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.ea23=ttk.Entry(self.Frame7,font=("Times 12"),width=24)
                self.ea23.grid(row=2,column=2,sticky='W',padx=20,pady=3,ipady=3)
                
                def valaad(event):  
                  regex1 = '^\d{4}\d{4}\d{4}$'
                  self.aadharno=self.ea23.get()
                  
                  if(re.search(regex1,self.aadharno)):
                      print('ok')
                  else:
                         messagebox.showinfo("Insert Error", "Enter a valid aadhar number")
                self.ea23.bind('<FocusOut>', valaad)
                                                
                self.la22=ttk.Label(self.Frame7,text="Cell Number",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.la22.grid(row=3,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.lla22=ttk.Label(self.Frame7,text="          :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.lla22.grid(row=3,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.ea22=ttk.Entry(self.Frame7,font=("Times 12"),width=24)
                self.ea22.grid(row=3,column=2,sticky='W',padx=20,pady=3,ipady=3)
                def valphno(event):  
                  regex3 = '(0/91)?[7-9][0-9]{9}'
                  self.cellno=self.ea22.get()
                  
                  if(re.search(regex3,self.cellno)):
                      print('ok')
                  else:
                         messagebox.showinfo("Insert Error", "Enter a valid phone number")
                self.ea22.bind('<FocusOut>', valphno)
                
                
                self.l2=ttk.Label(self.Frame7,text="Register Number",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l2.grid(row=4,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll2=ttk.Label(self.Frame7,text="          :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll2.grid(row=4,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e2=ttk.Entry(self.Frame7,font=("Times 12"),width=24)
                self.e2.grid(row=4,column=2,sticky='W',padx=20,pady=3,ipady=3)
                                        
                self.l3=ttk.Label(self.Frame7,text="Roll Number",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l3.grid(row=5,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll3=ttk.Label(self.Frame7,text="          :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll3.grid(row=5,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e3=ttk.Entry(self.Frame7,font=("Times 12"),width=24)
                self.e3.grid(row=5,column=2,sticky='W',padx=20,pady=3,ipady=3)
                
                self.l11=ttk.Label(self.Frame7,text="Course Of Study",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l11.grid(row=6,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll11=ttk.Label(self.Frame7,text="          :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll11.grid(row=6,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.t5 = tk.StringVar(self)
                self.p5 =ttk.Combobox(self.Frame7, width=21, textvariable=self.t5,font=("Times 12"))
                self.p5['values']=('BE','ME','BE Parttime')  
                self.p5.grid(row=6,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                self.p5.current()
                        
                self.l12=ttk.Label(self.Frame7,text="Branch Of Study",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l12.grid(row=7,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll12=ttk.Label(self.Frame7,text="          :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll12.grid(row=7,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.t6 = tk.StringVar(self)
                self.p6 =ttk.Combobox(self.Frame7,width=21, textvariable=self.t6,font=("Times 12"))
                self.p6['values']=('Civil Engineering','Mechanical Engineering','Electrical and Electronics Engineering','Electronics and Communication Engineering','Computer science and Engineering')  
                self.p6.grid(row=7,column=2,sticky='W',padx=20,pady=5,ipadx=2,ipady=5)
                self.p6.current()
                
                self.la26=ttk.Label(self.Frame7,text="Year Of Completion",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.la26.grid(row=8,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.lla26=ttk.Label(self.Frame7,text="          :",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.lla26.grid(row=8,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.ea26=ttk.Entry(self.Frame7,font=("Times 12"),width=24)
                self.ea26.grid(row=8,column=2,sticky='W',padx=20,pady=3,ipady=3)
                        
                self.Frame8 = tk.Frame(self.Frame3)
                self.Frame8.pack(side="top",expand=True)
                self.Frame8.configure(background='skyblue3')
                        
                self.l13=ttk.Label(self.Frame8,text=" Admitted On ",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l13.grid(row=1,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll13=ttk.Label(self.Frame8,text="                : ",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll13.grid(row=1,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e13=DateEntry(self.Frame8,width=24,date_pattern='dd-mm-yyyy', font=("Times 12"), borderwidth=2)
            
                self.e13.grid(row=1,column=2,sticky='W',padx=20,pady=3,ipady=3)
                                        
                self.l14=ttk.Label(self.Frame8,text="Receipt Number",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l14.grid(row=2,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll14=ttk.Label(self.Frame8,text="                : ",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll14.grid(row=2,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e14=ttk.Entry(self.Frame8,font=("Times 12"),width=24)
                self.e14.grid(row=2,column=2,sticky='W',padx=20,pady=3,ipady=3)
                
                self.l15=ttk.Label(self.Frame8,text="Receipt Date",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.l15.grid(row=3,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.ll15=ttk.Label(self.Frame8,text="                : ",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.ll15.grid(row=3,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.e15=DateEntry(self.Frame8,width=24,date_pattern='dd-mm-yyyy', font=("Times 12"), borderwidth=2)
            
                self.e15.grid(row=3,column=2,sticky='W',padx=20,pady=3,ipady=3)

                self.la25=ttk.Label(self.Frame8,text="Issued ON",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.la25.grid(row=4,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.lla25=ttk.Label(self.Frame8,text="                : ",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.lla25.grid(row=4,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.ea25=DateEntry(self.Frame8,width=24,date_pattern='dd-mm-yyyy', font=("Times 12"), borderwidth=2)
            
                self.ea25.grid(row=4,column=2,sticky='W',padx=20,pady=3,ipady=3)
                                         
                self.la24=ttk.Label(self.Frame8,text="T.C.No",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.la24.grid(row=5,column=0,sticky='W',pady=5,padx=20,ipady=5)
                self.lla24=ttk.Label(self.Frame8,text="                : ",font=("Times 12 bold"),background='skyblue3',foreground="white")
                self.lla24.grid(row=5,column=1,sticky='W',pady=5,padx=20,ipady=5)
                self.ea24=ttk.Entry(self.Frame8,font=("Times 12"),width=24)
                self.ea24.grid(row=5,column=2,sticky='W',padx=20,pady=3,ipady=3)
                
               
                
                               
                self.Frame9 = tk.Frame(self.Frame3)
                self.Frame9.pack(side="bottom",expand=True,ipadx=10)
                self.Frame9.configure(background='midnightblue')
                                                           
                self.b1=ttk.Button(self.Frame9,text="REGISTER",style = 'W.TButton',command=self.validatered)
                self.b1.grid(row=1,column=2,sticky='W',padx=5,pady=3,ipady=1)
                
                
                
                #self.but=ttk.Button(self,style = 'W.TButton',text="BACK",command=lambda: master.switch_frame(pageOne))
                #self.but.grid(row=14,column=4,sticky='W',pady=3,padx=40)
             
              def transsing(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(transfersingle)
              def transgrpo(self):
                  self.Frame14.destroy()
                  
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(transfergroup)
              def ins(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(newreg)
              def extratc(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(updatetc)
              def edt(self):
                  self.Frame14.destroy() 
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(editspecify)
              def ser(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(search)
              def logg(self):
                  self.Frame14.destroy()
                  self.Frame11.destroy()
                  self.Frame1.destroy()
                  self.Frame2.destroy()
                  self.Frame3.destroy()
                  self.Frame4.destroy()
                  self.Frame5.destroy()
                  self.Frame6.destroy()
                  self.Frame7.destroy()
                  self.Frame8.destroy()
                  self.master.switch_frame(StartPage)  
              def validatered(self):
                self.name=self.e1.get()
                self.regno=self.e2.get()
                self.rollno=self.e3.get()
                self.father=self.e4.get()
                self.nation=self.t1.get()
                self.religion=self.t2.get()
                self.caste=self.tt3.get()
                self.community=self.t3.get()
                self.sex=self.t4.get()
                self.dob=self.e10.get()
                self.course=self.t5.get()
                self.branch=self.t6.get()
                self.admiton=self.e13.get()
                self.receiptno=self.e14.get()
                self.receiptdate=self.e15.get()
                self.mothertongue=self.t7.get()
                self.state=self.t8.get()
                self.address=self.e18.get()
                self.taluk=self.e19.get()
                self.city=self.e20.get()  
                self.district=self.t9.get()
                self.cellno=self.ea22.get()
                self.aadharno=self.ea23.get()
                self.tcno=self.ea24.get()
                self.issuedon=self.ea25.get()
                self.yopo=self.ea26.get()
                 
                regex1 = '^\d{4}\d{4}\d{4}$'
               
                regex3 = '(0/91)?[7-9][0-9]{9}'
                
                
                #day,month,year = self.dob.split('-')
                #d,m,y=self.receiptdate.split('-')
                #d1,m1,y1=self.issuedon.split('-')
                try:
                    if(re.search(regex1,self.aadharno)):  
                         if(re.search(regex3,self.cellno)):  
                            datetime.datetime.strptime(self.dob, '%d-%m-%Y')
                            datetime.datetime.strptime(self.receiptdate, '%d-%m-%Y') 
                            datetime.datetime.strptime(self.issuedon, '%d-%m-%Y')
                            datetime.datetime.strptime(self.e13.get(), '%d-%m-%Y')
                            if(self.name!="" and self.regno !="" and self.rollno !="" and self.father!="" and self.nation !="" and self.religion!="" and self.caste!="" and self.community!="" and self.sex!="" and self.dob!="" and self.course!="" and  self.branch !="" and self.admiton!="" and self.receiptno!="" and self.receiptdate!="" and self.mothertongue!="" and self.state!="" and self.address!="" and self.taluk!="" and self.city!="" and self.district!="" and self.cellno!="" and self.aadharno!="" and self.tcno!="" and self.issuedon!="" and self.yopo!=""):
                                self.calle()
                                
                            else:
                             messagebox.showinfo("Insert Error", "Fill all the fields")
                                
                         else:  
                             messagebox.showinfo("Insert Error", "Not valid phone no")
                    else:  
                        messagebox.showinfo("Insert Error", "Not valid aadhar no") 
                    
                    
                    

                except ValueError :
                    messagebox.showinfo("Insert Error", "Not valid date no")

                
                
               
                    
                
                    
              def calle(self):
                
                self.name=self.e1.get()
                self.regno=self.e2.get()
                self.rollno=self.e3.get()
                self.father=self.e4.get()
                self.nation=self.t1.get()
                self.religion=self.t2.get()
                self.caste=self.tt3.get()
                self.community=self.t3.get()
                self.sex=self.t4.get()
                self.dob=self.e10.get()
                self.course=self.t5.get()
                self.branch=self.t6.get()
                self.admiton=self.e13.get()
                self.receiptno=self.e14.get()
                self.receiptdate=self.e15.get()
                self.mothertongue=self.t7.get()
                self.state=self.t8.get()
                self.address=self.e18.get()
                self.taluk=self.e19.get()
                self.city=self.e20.get()  
                self.district=self.t9.get()
                self.cellno=self.ea22.get()
                self.aadharno=self.ea23.get()
                self.tcno=self.ea24.get()
                self.issuedon=self.ea25.get()
                self.yopo=self.ea26.get()
                print(self.community)
                #INSERT INTO `studentdetails`(`name`, `reg_number`, `roll_number`, `father_name`, `nationality`, `religion`, `caste`, `community`, `sex`, `dateofbirth`, `course`, `branch`, `admittedon`, `receiptno`, `receiptdate`, `mothertongue`, `state`, `present_address`, `city`, `district`, `cell_number`, `aadhar_number`, `tcno`, `issuedon`) 
                
                try:
                    
                    sql1="INSERT INTO `studentdetails` (`name`, `reg_number`, `roll_number`, `father_name`, `nationality`, `religion`, `caste`, `community`, `sex`, `dateofbirth`, `course`, `branch`, `admittedon`, `receiptno`, `receiptdate`, `mothertongue`, `state`, `present_address`, `taluk`, `city`, `district`, `cell_number`, `aadhar_number`, `tcno`, `issuedon`, `Year_Of_Passout`,`flag`)VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    val1=(self.name,self.regno,self.rollno,self.father,self.nation,self.religion,self.community,self.caste,self.sex,self.dob,self.course,self.branch,self.admiton,self.receiptno,self.receiptdate,self.mothertongue,self.state,self.address,self.taluk,self.city,self.district,self.cellno, self.aadharno,self.tcno,self.issuedon,self.yopo,'0')
                    db.execute(sql1, val1)
                    db_cur.commit()
                    msg = 'A new record '+self.regno + ' is inserted.'
                    messagebox.showinfo("Insert", msg)
                    self.Frame14.destroy()
                    self.Frame11.destroy()
                    self.Frame1.destroy()
                    self.Frame2.destroy()
                    self.Frame3.destroy()
                    self.Frame4.destroy()
                    self.Frame5.destroy()
                    self.Frame6.destroy()
                    self.Frame7.destroy()
                    self.Frame8.destroy()
                    self.master.switch_frame(pageOne) 
                except mysql.connector.IntegrityError:
                       messagebox.showinfo("Registration Error", "Duplicate entry register number")

           
       
class search(tk.Frame):
                  def __init__(self, master):
                      tk.Frame.__init__(self,master)
                      self.configure(background='midnightblue')
                      adminname='logged in as '+username
                      self.Frame14 = tk.Frame(self.master,bg="yellow",bd=3,relief=tk.RAISED)
                    
                      self.Frame14.pack()
                        
                      self.menubar = tk.Menu(self.Frame14,tearoff=0,bd=2)
                      self.filemenu = tk.Menu(self.menubar, tearoff=0,background='midnightblue', foreground='white',activebackground='cyan', activeforeground='black')
                      self.filemenu.add_command(label="Individual Transfer Certificate",command=self.transsing,font=("Times 18 italic"))
                      self.filemenu.add_command(label="Batch Transfer Certificate",command=self.transgrpo,font=("Times 18 italic"))
                      self.filemenu.add_command(label="Update TC",command=self.extratc,font=("Times 18 italic"))
                      self.menubar.add_cascade(label="Transfer Certificate Generation", menu=self.filemenu,font = "Times 38")
                      self.edit = tk.Menu(self.menubar, tearoff=0,background='midnightblue', foreground='white',activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                      self.edit.add_command(label="Edit",command=self.edt ,font=("Times 18 italic"))
                         
                      self.menubar.add_cascade(label='Edit Details',menu=self.edit,font = "Times 38")
                      self.insert = tk.Menu(self.menubar, tearoff=0,background='midnightblue', foreground='white',activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                      self.insert.add_command(label="Insert",command=self.ins,font=("Times 18 italic"))
                     
                      self.menubar.add_cascade(label='Insert Details',menu=self.insert,font = "Times 38")
                      self.search = tk.Menu(self.menubar, tearoff=0,background='midnightblue', foreground='white',activebackground='cyan', activeforeground='black',font=("Times 18 italic"))
                      self.search.add_command(label="Search",command=self.ser,font=("Times 18 italic"))
                     
                      self.menubar.add_cascade(label='Search Details',menu=self.search,font = "Times 38")
                      self.menubar.add_command(label="Logout",command=self.logg,font=("Times 18 italic"))
                      self.menubar.add_command(label=adminname,font=("Times 18 italic"))
                     
                         #self.menubar.config("Verdana", 14)
                      self.master.config(menu=self.menubar)
                      self.Frame1 = tk.Frame(self.master)
                      self.Frame1.pack(side="top",  pady=1,padx=1,expand=True )
                      self.Frame1.configure(background='midnightblue')
                      photo=tk.PhotoImage(file="ss1.png")  
                      label1221=tk.Label(self.Frame1,text="Alagappa chettiar Government College of Engineering and Technology,Karaikudi",background="midnightblue",foreground="white",font =
                                         ('calibri', 25, 'bold'))
                      label1221.grid(row=0,rowspan=1,columnspan=40)
                      label12221=tk.Label(self.Frame1,text="(An autonomous government institution permanently affilitated to anna university)",background="midnightblue",foreground="white",font =
                                          ('calibri', 10, 'bold'))
                      label12221.grid(row=1,rowspan=1,columnspan=40)
                      label =tk.Label(self,image = photo,width=300,height=300)#,width=1680,height=1080)
                      label.image = photo # keep a reference!
                      label.grid(row=0,column=0,columnspan=4,rowspan=20,sticky='W')
                      mis=tk.Label(self,text='Search Here',font = ('calibri',30, 'bold'))
                      mis.grid(row=9,column=0,sticky='W')
                      self.regis=""
                      self.religion=""
                      self.caste=""
                      self.gen=""#self.eta3.get()
                      self.bran=""#self.eta4.get()
                      self.talu=""#self.eta5.get()
                      self.dist=""
                      
                      self.configure(background='burlywood1')
                      self.la=ttk.Label(self,text="search",font=("Times 25 bold"),background='burlywood1')
                      self.la.grid(row=0,column=5,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                      self.la1=ttk.Label(self,text="Register Number",font=("Times 20 bold"),background='burlywood1')
                      self.la1.grid(row=1,column=4,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                      self.en =ttk.Entry(self,font=("Times 12"))
                      self.en.grid(row=1,column=5,sticky='W',padx=20,pady=5,ipadx=40,ipady=5)

                      self.la1=ttk.Label(self,text="Religion",font=("Times 20 bold"),background='burlywood1')
                      self.la1.grid(row=2,column=4,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                      self.tkv1 = tk.StringVar(self)
                      self.popupMenu1 =ttk.Combobox(self, width=18, textvariable=self.tkv1,font=("Times 15"))
                      self.popupMenu1['values']=('Hindu','Muslim','Christian','Others')
                      self.popupMenu1.grid(row=2,column=5,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                      self.popupMenu1.current()
                      self.la2=ttk.Label(self,text="Community",font=("Times 20 bold"),background='burlywood1')
                      self.la2.grid(row=3,column=4,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                      self.tkv2 = tk.StringVar(self)
                      self.popupMenu2 =ttk.Combobox(self, width=18, textvariable=self.tkv2,font=("Times 15 bold"))
                      self.popupMenu2['values']=('OC','BC','BCM','MBC','SC','ST','Others')  
                      self.popupMenu2.grid(row=3,column=5,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                      self.popupMenu2.current()
                      self.la3=ttk.Label(self,text="Gender",font=("Times 20 bold"),background='burlywood1')
                      self.la3.grid(row=4,column=4,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                      self.tkv3 = tk.StringVar(self)
                      self.popupMenu3 =ttk.Combobox(self, width=18, textvariable=self.tkv3,font=("Times 15"))
                      self.popupMenu3['values']=('Male','Female','Transgender')
                      self.popupMenu3.grid(row=4,column=5,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                      self.popupMenu3.current()
                        #self.la5=ttk.Label(self,text="Programme",font=("Times 20 italic"),background='papayawhip')
                       # self.la5.grid(row=5,column=0,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                       # self.tkv4 = tk.StringVar(self)
                        #self.popupMenu4=ttk.Combobox(self, width=25, textvariable=self.tkv4,font=("Times 15 italic"))
                        #self.popupMenu4['values']=('BE','ME','BE-Parttime')
                        #self.popupMenu4.grid(row=5,column=1,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                        #self.popupMenu4.current()
                      self.la4=ttk.Label(self,text="Branch Of Study",font=("Times 20 bold"),background='burlywood1')
                      self.la4.grid(row=6,column=4,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                      self.tkv5 = tk.StringVar(self)
                      self.popupMenu5=ttk.Combobox(self, width=18, textvariable=self.tkv5,font=("Times 15"))
                      self.popupMenu5['values']=('Civil Engineering','Mechanical Engineering','Electrical and Electronics Engineering','Electronics and Communication Engineering','Computer Science and Engineering')
                      self.popupMenu5.grid(row=6,column=5,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                      self.popupMenu5.current()
                      self.la6=ttk.Label(self,text="Taluk",font=("Times 20 bold"),background='burlywood1')
                      self.la6.grid(row=7,column=4,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                      self.eta5=ttk.Entry(self,font=("Times 12"))
                      self.eta5.grid(row=7,column=5,sticky='W',padx=20,pady=5,ipadx=40,ipady=5)
                      self.la7=ttk.Label(self,text="District",font=("Times 20 bold"),background='burlywood1')
                      self.la7.grid(row=5,column=4,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                      self.tkvar = tk.StringVar(self)
                      self.popupMenu =ttk.Combobox(self, width=18, textvariable=self.tkvar,font=("Times 15"))
                      self.popupMenu['values']=('Ariyalur','Chengalpet','Chennai','Coimbatore','Cuddalore','Dharmapuri','Dindigul','Erode','Kallakurichi','Kancheepuram','Karur','Krishnagiri','Madurai','Nagapattinam','Nilgiris','Kanyakumari','Namakkal','Perambalur','Pudukottai','Ramanathapuram','Ranipet','Salem','Sivagangai','Tenkasi','Thanjavur','Theni','Thiruvallur','Thiruvarur','Tuticorin','Trichirappalli','Thirunelveli','Tirupattur','Tiruppur','Thiruvannamalai','Vellore','Viluppuram','Virudhunagar')
                      self.popupMenu.grid(row=5,column=5,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                      self.popupMenu.current()
                        
                      self.la8=ttk.Label(self,text="Year Of Completion",font=("Times 20 bold"),background='burlywood1')
                      self.la8.grid(row=8,column=4,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                      self.en1 =ttk.Entry(self,font=("Times 12"))
                      self.en1.grid(row=8,column=5,sticky='W',padx=20,pady=5,ipadx=40,ipady=5)
                      self.l11=ttk.Label(self,text="Course Of Study",font=("Times 20 bold"),background="burlywood1")
                      self.l11.grid(row=9,column=4,sticky='W',padx=10,pady=5,ipadx=20,ipady=5)
                      self.t5 = tk.StringVar(self)
                      self.p5 =ttk.Combobox(self, width=18, textvariable=self.t5,font=("Times 15"))
                      self.p5['values']=('BE','ME','BE Part time')  
                      self.p5.grid(row=9,column=5,sticky='W',padx=20,pady=5,ipadx=20,ipady=5)
                      self.p5.current()
                        #self.eta6=ttk.Entry(self,font=("Times 10 italic"))
                        #self.eta6.grid(row=7,column=1,sticky='W',padx=10,pady=10,ipadx=20,ipady=10)
                      self.buty=ttk.Button(self,text="search",style = 'W1search.TButton',command=self.say_hello)
                      self.buty.grid(row=10,column=5,sticky='W',padx=30,pady=10)
                 
                 
                  #def _login_btn_clickked(self):
                       # sex = self.religion #gets the value stored in gender and assigns it to sex
                       # print(sex)
                
                 
                  def transsing(self):
                      self.Frame14.destroy()
                      self.Frame1.destroy()
                      self.master.switch_frame(transfersingle)
                  def transgrpo(self):
                      self.Frame14.destroy()
                      self.Frame1.destroy()
                      self.master.switch_frame(transfergroup)
                  def ins(self):
                      self.Frame14.destroy()
                      self.Frame1.destroy()
                      self.master.switch_frame(newreg)
                  def edt(self):
                      self.Frame14.destroy()
                      self.Frame1.destroy()
                      self.master.switch_frame(editspecify)
                  def extratc(self):
                      self.Frame14.destroy()
                      self.Frame1.destroy()
                      self.master.switch_frame(updatetc)
                  def ser(self):
                      self.Frame14.destroy()
                      self.Frame1.destroy()
                      self.master.switch_frame(search)
                  def logg(self):
                      self.Frame14.destroy()
                      self.Frame1.destroy()
                      self.master.switch_frame(StartPage)    
                  def say_hello(self):
                      
                      # reli,cast,seex,bos,tal,dis
                      self.regis=self.en.get()
                      self.yopo=self.en1.get()
                      self.reli=self.tkv1.get()
                      print(self.reli)
                      self.cast=self.tkv2.get()
                      print(self.cast)#self.eta2.get()
                      self.seex=self.tkv3.get()
                      print(self.seex)#self.eta3.get()
                      self.bos=self.tkv5.get()
                      print(self.bos)#self.eta4.get()
                      self.tal=""#self.eta5.get()
                      self.dis=self.tkvar.get()#self.eta6.get()
                      print(self.dis)
                      self.course=self.t5.get()
                      wb = Workbook()
                      #global ff
                      #print(self.reli)
                      # 1 combo
                      if(self.course=="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' " %(self.reli))
                          #ff=db.fetchone()
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religion"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' " %(self.course))
                          #ff=db.fetchone()
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "course"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                      elif(self.regis!=""):
                          db.execute("select *from studentdetails where reg_number='%s' " %(self.regis))
                          #ff=db.fetchone()
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "register"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' " %(self.yopo))
                          #ff=db.fetchone()
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "YearOfPassOut"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          print("entered")
                          db.execute("select *from studentdetails where caste ='%s'" %(self.cast))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "caste"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex ='%s'" %(self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "gender"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where branch ='%s'" %(self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "Branch of study"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where taluk ='%s'" %(self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "Taluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district ='%s'" %(self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "District"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          #2combo
                      elif(self.course!="" and self.reli!=" " and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and course='%s'" %(self.reli,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli==" " and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste ='%s'" %(self.course,self.cast))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecaste"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli==" " and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course ='%s' and sex ='%s'" %(self.course,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli==" " and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course ='%s' and branch='%s'" %(self.course,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursebranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli==" " and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and taluk='%s'" %(self.course,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursetaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli==" " and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and district='%s'" %(self.course,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursedistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli==" " and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and Year_Of_Passout='%s'" %(self.course,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "courseyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

#2 combo la 7 completed
                          
                      elif(self.course=="" and self.reli!=" " and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and caste ='%s'" %(self.reli,self.cast))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncaste"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                      elif(self.course=="" and self.reli!="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and sex='%s'" %(self.reli,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiongender"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and branch='%s'" %(self.reli,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and taluk='%s'" %(self.reli,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion ='%s' and branch='%s'" %(self.reli,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion ='%s' and Year_Of_Passout='%s'" %(self.reli,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                    
                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s'" %(self.cast,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and branch='%s'" %(self.cast,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castebranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s'" %(self.cast,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and district='%s'" %(self.cast,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castedistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and Year_Of_Passout='%s'" %(self.cast,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "casteyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and branch='%s'" %(self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and taluk='%s'" %(self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s'" %(self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and Year_Of_Passout='%s'" %(self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where branch='%s' and taluk='%s'" %(self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where branch='%s' and district='%s'" %(self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where branch='%s' and Year_Of_Passout='%s'" %(self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)    
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where taluk='%s' and district='%s'" %(self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "talukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where taluk='%s' and Year_Of_Passout='%s'" %(self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "talukyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and district='%s'" %(self.yopo,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      
                    # 3 combo

                      elif(self.course!="" and self.reli!="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and course='%s'" %(self.reli,self.cast,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastecourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and course='%s' and sex='%s'" %(self.reli,self.course,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncoursesex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and course='%s' and branch='%s'" %(self.reli,self.course,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncoursebranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and course='%s' and taluk='%s'" %(self.reli,self.course,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncoursetaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and course='%s' and district='%s'" %(self.reli,self.course,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncoursedistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and course='%s' and Year_Of_Passout='%s'" %(self.reli,self.course,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncourseyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast !="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s'" %(self.course,self.cast,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and branch='%s'" %(self.course,self.cast,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastebranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and taluk='%s'" %(self.course,self.cast,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastetaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and district='%s'" %(self.course,self.cast,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastedistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecasteyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s'" %(self.course,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s'" %(self.course,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and district='%s'" %(self.course,self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                        

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.course,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)   
                          
                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and branch='%s' and taluk='%s'" %(self.course,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                         
                        
                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and branch='%s' and district='%s'" %(self.course,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursebranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.course,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursebranchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and taluk='%s' and district='%s'" %(self.course,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursetalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursetalukyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast =="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursedistrictyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)


# 2 combo 21 completed


  
                      elif(self.course=="" and self.reli!="" and self.cast !="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s'" %(self.reli,self.cast,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast !="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s'" %(self.reli,self.cast,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s'" %(self.reli,self.cast,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s'" %(self.reli,self.cast,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast !="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncasteyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s'" %(self.reli,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and taluk='%s'" %(self.reli,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and district='%s'" %(self.reli,self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and taluk='%s' and branch='%s'" %(self.reli,self.tal,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontalukbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and branch='%s'" %(self.reli,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and branch='%s'" %(self.reli,self.yopo,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearfpassouttbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)    
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and taluk='%s' and district='%s'" %(self.reli,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.reli,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontalukyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and Year_Of_Passout='%s'" %(self.reli,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and branch='%s'" %(self.cast,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and taluk='%s'" %(self.cast,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and district='%s'" %(self.cast,self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.cast,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s' and branch='%s'" %(self.cast,self.tal,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetalukbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and district='%s' and branch='%s'" %(self.cast,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castedistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                     
                        
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and Year_Of_Passout='%s' and branch='%s'" %(self.cast,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "casteyopobranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      
                    
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s' and district='%s'" %(self.cast,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.cast,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetalukyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and Year_Of_Passout='%s' and district='%s'" %(self.cast,self.yopo,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "casteyearofpassoutdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where taluk='%s' and sex='%s' and branch='%s'" %(self.tal,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "taluksexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district='%s' and sex='%s' and branch='%s'" %(self.dis,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtsexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and sex='%s' and branch='%s'" %(self.yopo,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutsexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)    
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district='%s' and sex='%s' and taluk='%s'" %(self.dist,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtsextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and sex='%s' and taluk='%s'" %(self.yopo,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "YearOfPassoutsextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      
                        # seex and dis and yopo
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.seex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrictYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                        
                        
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district='%s' and taluk='%s' and branch='%s'" %(self.dis,self.tal,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districttalukbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and taluk='%s' and branch='%s'" %(self.yopo,self.tal,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "YearOfPassouttalukbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # bos and dis and yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and district='%s' and branch='%s'" %(self.yopo,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "YearOfPassoutdistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                    
                          
                          # tal and dis and yopo
            
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and district='%s' and taluk='%s'" %(self.yopo,self.dis,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "YearOfPassoutdistricttaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)    

                    # 4 combo


                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s' and course='%s'" %(self.reli,self.cast,self.seex,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesexcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and course='%s'" %(self.reli,self.cast,self.bos,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s' and course='%s'" %(self.reli,self.cast,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s' and course='%s'" %(self.reli,self.cast,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and Year_Of_Passout='%s' and course='%s'" %(self.reli,self.cast,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncasteyearofpassoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and course='%s'" %(self.reli,self.seex,self.bos,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexbranchcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and taluk='%s' and course='%s'" %(self.reli,self.seex,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsextalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and district='%s' and course='%s'" %(self.reli,self.seex,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and Year_Of_Passout='%s' and course='%s'" %(self.reli,self.seex,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexyearofpassoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and branch='%s' and taluk='%s' and course='%s'" %(self.reli,self.bos,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionbranchtalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and branch='%s' and district='%s' and course='%s'" %(self.reli,self.bos,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionbranchdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and branch='%s' and Year_Of_Passout='%s' and course='%s'" %(self.reli,self.bos,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionbranchyearofpassoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and taluk='%s' and district='%s' and course='%s'" %(self.reli,self.tal,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontalukdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and taluk='%s' and Year_Of_Passout='%s' and course='%s'" %(self.reli,self.tal,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiontalukYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and Year_Of_Passout='%s' and course='%s'" %(self.reli,self.dis,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and branch='%s' and course='%s'" %(self.cast,self.seex,self.bos,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexbranchcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and taluk='%s' and course='%s'" %(self.cast,self.seex,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesextalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and district='%s' and course='%s'" %(self.cast,self.seex,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and Year_Of_Passout='%s' and course='%s'" %(self.cast,self.seex,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesexYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and branch='%s' and taluk='%s' and course='%s'" %(self.cast,self.bos,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castebranchtalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and branch='%s' and district='%s' and course='%s'" %(self.cast,self.bos,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castebranchdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and branch='%s' and Year_Of_Passout='%s' and course='%s'" %(self.cast,self.bos,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castebranchYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s' and district='%s' and course='%s'" %(self.cast,self.tal,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetalukdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and taluk='%s' and Year_Of_Passout='%s' and course='%s'" %(self.cast,self.tal,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castetalukYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and district='%s' and Year_Of_Passout='%s' and course='%s'" %(self.cast,self.dis,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castedistrictYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and branch='%s' and taluk='%s' and course='%s'" %(self.seex,self.bos,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexbranchtalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)         
                          
                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and branch='%s' and district='%s' and course='%s'" %(self.seex,self.bos,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexbranchdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and branch='%s' and Year_Of_Passout='%s' and course='%s'" %(self.seex,self.bos,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexbranchYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and taluk='%s' and district='%s' and course='%s'" %(self.seex,self.tal,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sextalukdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and taluk='%s' and Year_Of_Passout='%s' and course='%s'" %(self.seex,self.tal,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sextalukYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s' and Year_Of_Passout='%s' and course='%s'" %(self.seex,self.dis,self.yopo,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrictYear_Of_Passoutcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where branch='%s' and district='%s' and taluk='%s' and course='%s'" %(self.bos,self.dis,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchdistricttalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where branch='%s' and Year_Of_Passout='%s' and taluk='%s' and course='%s'" %(self.bos,self.yopo,self.tal,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchYear_Of_Passouttalukcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where branch='%s' and Year_Of_Passout='%s' and district='%s' and course='%s'" %(self.bos,self.yopo,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchYear_Of_Passoutdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where taluk='%s' and Year_Of_Passout='%s' and district='%s' and course='%s'" %(self.tal,self.yopo,self.dis,self.course))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "talukYear_Of_Passoutdistrictcourse"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)



# 4 combo 35 combinations completed




                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s' and branch='%s'" %(self.reli,self.cast,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesexbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s' and taluk='%s'" %(self.reli,self.cast,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesextaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s' and district='%s'" %(self.reli,self.cast,self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastesexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s'" %(self.reli,self.cast,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and district='%s'" %(self.reli,self.cast,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s' and taluk='%s'" %(self.reli,self.cast,self.dis,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistricttaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      #reli,cast,tal,yopo    
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s'and Year_Of_Passout='%s'" %(self.reli,self.cast,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetalukYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                    #reli,cast,dis,yopo
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistrictYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                   
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and taluk='%s'" %(self.reli,self.seex,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                      
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and district='%s'" %(self.reli,self.seex,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexbranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.seex,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexbranchYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and district='%s' and taluk='%s'" %(self.reli,self.seex,self.dis,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexdistricttaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      #reli,seex,tal,yopo
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.reli,self.seex,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsextalukYearOfPassout"
                          wb.save(workbook_name + ".xlsx")                     
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      #reli,seex,dis,yopo
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and district='%s' and  Year_Of_Passout='%s'" %(self.reli,self.seex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsexdistrictYearOfPassout"
                          wb.save(workbook_name + ".xlsx")                          
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and branch='%s' and taluk='%s'" %(self.reli,self.dis,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and branch='%s' and taluk='%s'" %(self.reli,self.yopo,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyopobranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.dis,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictbranchyopo"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      #reli,tal,dis,yopo
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.reli,self.dis,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistricttalukYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                    # 4 combo continutation
                    
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and branch='%s' and taluk='%s'" %(self.seex,self.cast,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcastebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and branch='%s' and district='%s'" %(self.seex,self.cast,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcastebranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.seex,self.cast,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcastebranchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and district='%s' and taluk='%s'" %(self.seex,self.cast,self.dis,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcaste districtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      # cast,seex,tal,yopo
                      
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo !=""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and Year_Of_Passout='%s' and taluk='%s'" %(self.seex,self.cast,self.yopo,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcasteYearOfPassouttaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                      
                      
                      # cast,seex,dis,yopo

                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis !="" and self.yopo !=""):
                          db.execute("select *from studentdetails where sex='%s' and caste='%s' and Year_Of_Passout='%s' and district='%s'" %(self.seex,self.cast,self.yopo,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexcasteYearOfPassoutdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)   


                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district='%s' and caste='%s' and branch='%s' and taluk='%s'" %(self.dis,self.cast,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtcastebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and caste='%s' and branch='%s' and taluk='%s'" %(self.yopo,self.cast,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtcastebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          

                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal=="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where district='%s' and caste='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.dis,self.cast,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtcastebranchYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # cast,tal,dis,yopo
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal!="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where district='%s' and caste='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.dis,self.cast,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtcastetalukYearOfPassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s' and branch='%s' and taluk='%s'" %(self.seex,self.dis,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrictbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and Year_Of_Passout='%s' and branch='%s' and taluk='%s'" %(self.seex,self.yopo,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexYear_Of_Passoutbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.seex,self.dis,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistrictbranchYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                     
                        # seex,tal,dis,yopo
                        
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where sex='%s' and district='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.seex,self.dis,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "sexdistricttalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                        
                        
                        # bos,tal,dis,yopo
                        
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where branch='%s' and district='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.bos,self.dis,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "branchdistricttalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)   
                          
                          
                         # 5 combo
                         
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranch"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                         
                         
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and taluk='%s'" %(self.course,self.reli,self.cast,self.seex,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesextaluk"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)   
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and district='%s'" %(self.course,self.reli,self.cast,self.seex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
 
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                        
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and taluk='%s'" %(self.course,self.reli,self.cast,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and district='%s'" %(self.course,self.reli,self.cast,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.cast,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastetalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastetalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastedistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and taluk='%s'" %(self.course,self.reli,self.seex,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and district='%s'" %(self.course,self.reli,self.seex,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.seex,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsextalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsextalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)  
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionbranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
 
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionbranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)  
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionbranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligiontalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s'" %(self.course,self.cast,self.seex,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                         
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and district='%s'" %(self.course,self.cast,self.seex,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and taluk='%s' and district='%s'" %(self.course,self.cast,self.seex,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesextalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesextalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.cast,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastebranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastebranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastebranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastetalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.seex,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.seex,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.seex,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.seex,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesextalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursebranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                         
                         
# 5 combo 35 combinations completed                         
                         
                         
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s'" %(self.reli,self.cast,self.bos,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtaluksex"
                          wb.save(workbook_name + ".xlsx")
                          wb.system(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and district='%s' and sex='%s'" %(self.reli,self.cast,self.bos,self.dis,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchdistrictsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # reli,cast,seex,bos,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and Year_Of_Passout='%s' and sex='%s'" %(self.reli,self.cast,self.bos,self.yopo,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchyearofpassoutsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s' and taluk='%s' and sex='%s'" %(self.reli,self.cast,self.dis,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistricttaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and Year_Of_Passout='%s' and taluk='%s' and sex='%s'" %(self.reli,self.cast,self.yopo,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncasteYearOfPassouttaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and district='%s' and Year_Of_Passout='%s' and sex='%s'" %(self.reli,self.cast,self.dis,self.yopo,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastedistrictYearOfPassoutsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.reli,self.cast,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # reli,cast,bos,tal,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtalukyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # reli,cast,bos,dis,yopo

                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchdistrictyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                          # reli,cast,tal,dis,yopo

                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetalukdistrictyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and district='%s' and branch='%s' and taluk='%s' and sex='%s'" %(self.reli,self.dis,self.bos,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religiondistrictbranchtaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                          # reli,seex,bos,tal,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and branch='%s' and taluk='%s' and sex='%s'" %(self.reli,self.yopo,self.bos,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearofpassoutbranchtaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # reli,seex,bos,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and branch='%s' and district='%s' and sex='%s'" %(self.reli,self.yopo,self.bos,self.dis,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearofpassoutbranchdistrictsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                           
                          
                          # reli,seex,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and taluk='%s' and district='%s' and sex='%s'" %(self.reli,self.yopo,self.tal,self.dis,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearofpassouttalukdistrictsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                            
                          
                          # reli,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and Year_Of_Passout='%s' and taluk='%s' and district='%s' and branch='%s'" %(self.reli,self.yopo,self.tal,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionyearofpassouttalukdistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                           
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where district='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s'" %(self.dis,self.cast,self.bos,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "districtcastebranchtaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)    
                          
                          
                          # cast,seex,bos,tal,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s'" %(self.yopo,self.cast,self.bos,self.tal,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutcastebranchtaluksex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # cast,seex,bos,dis,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and caste='%s' and branch='%s' and district='%s' and sex='%s'" %(self.yopo,self.cast,self.bos,self.dis,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutcastebranchdistrictsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # cast,seex,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and caste='%s' and taluk='%s' and district='%s' and sex='%s'" %(self.yopo,self.cast,self.tal,self.dis,self.seex))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutcastetalukdistrictsex"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                           
                          
                          # cast,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and caste='%s' and taluk='%s' and district='%s' and branch='%s'" %(self.yopo,self.cast,self.tal,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutcastetalukdistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # seex,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where Year_Of_Passout='%s' and sex='%s' and taluk='%s' and district='%s' and branch='%s'" %(self.yopo,self.sex,self.tal,self.dis,self.bos))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "yearofpassoutsextalukdistrictbranch"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)            
                          
                          
                          
                          # 6 combo
                          
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.tal))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchtaluk"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and district='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
 
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.cast,self.seex,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesextalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesextalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.cast,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastetalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.seex,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsextalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionfbranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.cast,self.seex,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesexbranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastesextalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursecastebranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.seex,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursesexbranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          
                          
# 6 combo 21 combinations completed                          
                          
                                                    
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s' and district='%s'" %(self.reli,self.cast,self.bos,self.tal,self.sex,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtaluksexdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                          # reli,cast,seex,bos,tal,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.tal,self.sex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtaluksexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                         
                         
                          # reli,cast,seex,bos,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and district='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.dis,self.seex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchdistrictsexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                            
                          
                          # reli,cast,seex,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s' and district='%s' and sex='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.tal,self.dis,self.sex,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetalukdistrictsexyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                           
                          
                          # reli,cast,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and taluk='%s' and district='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.tal,self.dis,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastetalukdistrictbranchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # reli,seex,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and sex='%s' and taluk='%s' and district='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.reli,self.sex,self.tal,self.dis,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religionsextalukdistrictbranchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          # cast,seex,bos,tal,dis,yopo
                          
                      elif(self.course=="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where caste='%s' and sex='%s' and taluk='%s' and district='%s' and branch='%s' and Year_Of_Passout='%s'" %(self.cast,self.sex,self.tal,self.dis,self.bos,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "castesextalukdistrictbranchyearofpassout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                           
                          
                          
                          
                          # 7 combo



                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo==""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.tal,self.dis))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchtalukdistrict"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis=="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.tal,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchtalukYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal =="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and branch='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.bos,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesexbranchdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                    
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos==""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and sex='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.seex,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastesextalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)
                          
                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex ==""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir) 
                          
                      elif(self.course!="" and self.reli!="" and self.cast=="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.seex,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)

                      elif(self.course!="" and self.reli=="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and caste='%s' and sex='%s' and branch='%s' and taluk='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.cast,self.seex,self.bos,self.tal,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligionsexbranchtalukdistrictYear_Of_Passout"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                          
                          
# 7 combo 7 combinations completed                          
                          
                          
                      elif(self.course=="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where religion='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.reli,self.cast,self.bos,self.tal,self.sex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "religioncastebranchtaluksexdistrictyopo"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)


                    # 8 combo


                      elif(self.course!="" and self.reli!="" and self.cast!="" and self.seex !=""  and self.bos!=""  and self.tal !="" and self.dis!="" and self.yopo!=""):
                          db.execute("select *from studentdetails where course='%s' and religion='%s' and caste='%s' and branch='%s' and taluk='%s' and sex='%s' and district='%s' and Year_Of_Passout='%s'" %(self.course,self.reli,self.cast,self.bos,self.tal,self.sex,self.dis,self.yopo))
                          results = db.fetchall()
                          ws = wb.create_sheet(0)
                          #ws.title = studentdetails
                          ws.append(db.column_names)
                          for row in results:
                            ws.append(row)
                          workbook_name = "coursereligioncastebranchtaluksexdistrictyopo"
                          wb.save(workbook_name + ".xlsx")
                          v=os.path.abspath(workbook_name + ".xlsx")
                          dir = v.replace('\\','/')
                          webbrowser.open(dir)                          
                  
# 8 combo 1 combinations completed  
                  
                      


if __name__ == "__main__":
    firstgui=MyFirstGui()
    firstgui.mainloop()
